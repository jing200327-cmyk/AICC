from __future__ import annotations

import asyncio
import sys
import uuid
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from split_import.models import SplitJob, SplitOutputFile, SplitSourceFile
from split_import.service import SplitImportService

from .models import OutcallFile, OutcallJob
from .repository import OutcallJobRepository


BASE_URLS = {'test': 'https://uat.aidcc.cn', 'prod': 'https://service.aidcc.cn'}
ENVIRONMENT_NAMES = {'test': '测试环境', 'prod': '生产环境'}


class OutcallError(Exception):
    code = 'OUTCALL_ERROR'
    message = 'Outcall failed'

    def __init__(self, detail: str = ''):
        super().__init__(detail or self.message)
        self.detail = detail


class InvalidOutcallEnvironmentError(OutcallError):
    code = 'INVALID_OUTCALL_ENVIRONMENT'
    message = 'Outcall environment must be test or prod'


class InvalidOutcallModeError(OutcallError):
    code = 'INVALID_OUTCALL_MODE'
    message = 'Outcall mode must be test or formal'


class SplitJobNotFoundError(OutcallError):
    code = 'SPLIT_JOB_NOT_FOUND'
    message = 'Split preview job does not exist'


class SplitJobNotReadyError(OutcallError):
    code = 'SPLIT_JOB_NOT_READY'
    message = 'Split preview job is not completed'


class OutcallTenantNotFoundError(OutcallError):
    code = 'OUTCALL_TENANT_NOT_FOUND'
    message = 'Tenant is not configured in config.yaml'


class OutcallFileNotFoundError(OutcallError):
    code = 'OUTCALL_FILE_NOT_FOUND'
    message = 'No split output file is available for this outcall mode'


class DuplicateOutcallFileError(OutcallError):
    code = 'DUPLICATE_OUTCALL_FILE'
    message = 'Duplicate outcall file is already queued'


class OutcallQueuePausedError(OutcallError):
    code = 'OUTCALL_QUEUE_PAUSED'
    message = 'Outcall queue is paused; use resume pending batches'


class OutcallJobAlreadyRunningError(OutcallError):
    code = 'OUTCALL_JOB_ALREADY_RUNNING'
    message = 'An outcall job is already running for this store'


class OutcallPlatformStatusError(OutcallError):
    code = 'OUTCALL_PLATFORM_STATUS_ERROR'
    message = 'Unable to query outcall platform status'


class OutcallService:
    def __init__(
        self,
        config_path: Path,
        source_root: Path,
        split_service: SplitImportService,
        db_path: Path | None = None,
    ):
        self.config_path = Path(config_path)
        self.source_root = Path(source_root)
        self.split_service = split_service
        self.repository = OutcallJobRepository(
            db_path or (Path(__file__).resolve().parents[1] / 'storage' / 'aicc.sqlite3')
        )
        self.jobs: dict[str, OutcallJob] = {
            job.job_id: job for job in self.repository.load_jobs()
        }
        self.statuses: dict[str, Any] = {}
        self.tasks: dict[str, asyncio.Task] = {}
        self.file_signatures: dict[str, dict[str, tuple[tuple[str, ...], ...]]] = {}

    def start_job(self, store_code: str, environment: str, mode: str, split_job_id: str, force_restart: bool = False) -> OutcallJob:
        environment = self._validate_environment(environment)
        mode = self._validate_mode(mode)
        if mode == 'formal' and self.repository.is_queue_paused(
            store_code, environment, self._run_date()
        ):
            raise OutcallQueuePausedError(store_code)
        split_job = self._get_split_job(store_code, split_job_id)
        tenant, config, excel_cls, status_cls, process_tenant = self._load_tenant(split_job.store_name, environment)
        files = self._select_files(split_job, tenant.name, mode, excel_cls)
        file_signatures = {file.file_path.name: self._first_rows_signature(file.file_path) for _, file in files}
        if environment == 'prod' and not force_restart:
            self._assert_no_duplicate_outcall_files(tenant.name, files, file_signatures)

        return self._create_job(
            store_code=store_code,
            environment=environment,
            mode=mode,
            split_job_id=split_job_id,
            tenant=tenant,
            config=config,
            status_cls=status_cls,
            process_tenant=process_tenant,
            files=files,
            file_signatures=file_signatures,
            resume_existing=False,
        )

    def _create_job(
        self,
        *,
        store_code: str,
        environment: str,
        mode: str,
        split_job_id: str,
        tenant: Any,
        config: Any,
        status_cls: Any,
        process_tenant: Any,
        files: list[tuple[Any, Any]],
        file_signatures: dict[str, tuple[tuple[str, ...], ...]],
        resume_existing: bool,
    ) -> OutcallJob:
        now = datetime.now().isoformat()
        job = OutcallJob(
            job_id=f'outcall_{datetime.now().strftime("%Y%m%d")}_{uuid.uuid4().hex[:10]}',
            status='running',
            environment=environment,
            environment_name=ENVIRONMENT_NAMES[environment],
            base_url=config.base_url,
            store_code=store_code,
            store_name=tenant.name,
            mode=mode,
            split_job_id=split_job_id,
            resume_existing=resume_existing,
            files=[OutcallFile(
                batch_key=getattr(item, 'batch_key', ''),
                batch_name=file.batch_name,
                filename=file.file_path.name,
                path=str(file.file_path),
                row_count=getattr(item, 'row_count', 0),
            ) for item, file in files],
            created_at=now,
            updated_at=now,
        )
        status = status_cls(tenant.name)
        self.jobs[job.job_id] = job
        self.statuses[job.job_id] = status
        self.file_signatures[job.job_id] = file_signatures
        self.repository.save_job(job)
        self.tasks[job.job_id] = asyncio.create_task(
            self._run_job(job.job_id, config, tenant, [file for _, file in files], status, process_tenant)
        )
        self._sync_status(job.job_id)
        return self.jobs[job.job_id]

    def get_job(self, job_id: str) -> OutcallJob | None:
        self._sync_status(job_id)
        return self.jobs.get(job_id)

    def terminate_job(self, job_id: str) -> OutcallJob | None:
        status = self.statuses.get(job_id)
        queued_before_stop = list(
            getattr(status, 'queued_batches', []) if status else []
        )
        if status:
            status.request_terminate()
        task = self.tasks.get(job_id)
        if task and not task.done():
            task.cancel()
        job = self.jobs.get(job_id)
        if job:
            job.stop_requested = True
            job.stopped_batches = queued_before_stop or list(job.queued_batches)
            job.status = 'terminated'
            job.state = '已终止'
            job.message = '已停止本地后续批次调度；平台当前任务不会由此接口停止'
            job.updated_at = datetime.now().isoformat()
            self.repository.save_job(job)
        return job

    async def _run_job(self, job_id: str, config: Any, tenant: Any, files: list[Any], status: Any, process_tenant: Any):
        job = self.jobs[job_id]
        try:
            status.state = '启动中'
            status.message = f'{job.environment_name}，准备上传 {len(files)} 个批次'
            self._sync_status(job_id)
            await process_tenant(
                config,
                tenant,
                files,
                status,
                resume_existing=job.resume_existing,
            )
            self._sync_status(job_id)
            if job.status != 'terminated':
                job.status = 'completed' if status.state in ('全部完成', '处理完成', '无文件') else 'failed'
                if status.state == '全部完成':
                    job.status = 'completed'
                elif status.state == '处理完成':
                    failed_count = sum(1 for item in status.batch_results if not item.get('ok'))
                    job.status = 'failed' if failed_count else 'completed'
                elif status.state in (
                    '登录失败',
                    '配置缺失',
                    '上传失败',
                    '外呼失败',
                    '任务未创建',
                    '平台状态查询失败',
                ):
                    job.status = 'failed'
                job.updated_at = datetime.now().isoformat()
                self.repository.save_job(job)
        except asyncio.CancelledError:
            if job.stop_requested:
                job.status = 'terminated'
                job.state = '已终止'
                job.message = '已停止本地后续批次调度；平台当前任务不会由此接口停止'
            else:
                job.status = 'running'
                job.state = '等待恢复'
                job.message = '后端服务已关闭，下一次启动将根据平台状态恢复队列'
            job.updated_at = datetime.now().isoformat()
            self.repository.save_job(job)
            raise
        except Exception as exc:
            job.status = 'failed'
            job.state = '执行异常'
            job.message = str(exc)
            job.error = {'code': 'OUTCALL_EXECUTION_FAILED', 'message': 'Outcall execution failed', 'detail': str(exc)}
            job.updated_at = datetime.now().isoformat()
            self.repository.save_job(job)

    def _sync_status(self, job_id: str):
        job = self.jobs.get(job_id)
        status = self.statuses.get(job_id)
        if not job or not status:
            return
        snapshot = status.to_dict()
        job.state = snapshot.get('state') or status.state
        job.message = status.message or snapshot.get('msg', '')
        job.progress = snapshot.get('progress', '')
        job.task_id = status.task_id
        job.task_state = status.task_state
        job.current_batch = status.current_batch
        job.queued_batches = list(status.queued_batches)
        job.batch_results = list(status.batch_results)
        job.updated_at = datetime.now().isoformat()
        self.repository.save_job(job)

    async def restore_running_jobs(self) -> None:
        for job in list(self.jobs.values()):
            if job.status != 'running' or job.stop_requested or job.job_id in self.tasks:
                continue
            if self._run_date(job.created_at) != self._run_date():
                job.status = 'failed'
                job.state = '恢复失败'
                job.message = '外呼恢复任务已跨日，为防止误用次日文件未自动恢复'
                job.updated_at = datetime.now().isoformat()
                self.repository.save_job(job)
                continue
            if self.repository.is_queue_paused(
                job.store_code, job.environment, self._run_date(job.created_at)
            ):
                job.status = 'terminated'
                job.state = '已终止'
                job.message = '队列已暂停，后端重启后未恢复调度'
                job.updated_at = datetime.now().isoformat()
                self.repository.save_job(job)
                continue
            try:
                split_job = self._recover_split_job_from_outputs(
                    job.store_code, job.split_job_id
                )
                tenant, config, excel_cls, status_cls, process_tenant = self._load_tenant(
                    split_job.store_name, job.environment
                )
                files = self._select_files(split_job, tenant.name, job.mode, excel_cls)
                status = status_cls(tenant.name)
                job.resume_existing = True
                job.files = [
                    OutcallFile(
                        batch_key=getattr(item, 'batch_key', ''),
                        batch_name=file.batch_name,
                        filename=file.file_path.name,
                        path=str(file.file_path),
                        row_count=getattr(item, 'row_count', 0),
                    )
                    for item, file in files
                ]
                job.state = '恢复中'
                job.message = '后端重启，正在根据平台任务状态恢复批次'
                job.updated_at = datetime.now().isoformat()
                self.statuses[job.job_id] = status
                self.repository.save_job(job)
                self.tasks[job.job_id] = asyncio.create_task(
                    self._run_job(
                        job.job_id,
                        config,
                        tenant,
                        [file for _, file in files],
                        status,
                        process_tenant,
                    )
                )
            except OutcallError as exc:
                job.status = 'failed'
                job.state = '恢复失败'
                job.message = exc.detail or exc.message
                job.error = {
                    'code': exc.code,
                    'message': exc.message,
                    'detail': exc.detail,
                }
                job.updated_at = datetime.now().isoformat()
                self.repository.save_job(job)

    def resume_pending_batches(self, store_code: str, environment: str) -> OutcallJob:
        environment = self._validate_environment(environment)
        if self._has_running_job(store_code, environment):
            raise OutcallJobAlreadyRunningError(store_code)
        split_job_id = f'resume_{datetime.now().strftime("%Y%m%d")}_{store_code}'
        split_job = self._recover_split_job_from_outputs(store_code, split_job_id)
        tenant, config, excel_cls, status_cls, process_tenant = self._load_tenant(
            split_job.store_name, environment
        )
        files = self._select_files(split_job, tenant.name, 'formal', excel_cls)
        file_signatures = {
            file.file_path.name: self._first_rows_signature(file.file_path)
            for _, file in files
        }
        self.repository.set_queue_paused(
            store_code, environment, self._run_date(), False
        )
        return self._create_job(
            store_code=store_code,
            environment=environment,
            mode='formal',
            split_job_id=split_job_id,
            tenant=tenant,
            config=config,
            status_cls=status_cls,
            process_tenant=process_tenant,
            files=files,
            file_signatures=file_signatures,
            resume_existing=True,
        )

    async def stop_pending_batches(
        self, store_code: str, environment: str
    ) -> dict[str, Any]:
        environment = self._validate_environment(environment)
        self.split_service.get_store(store_code)
        self.repository.set_queue_paused(
            store_code, environment, self._run_date(), True
        )
        stopped_job_ids = []
        for job in list(self.jobs.values()):
            if (
                job.store_code == store_code
                and job.environment == environment
                and job.status == 'running'
            ):
                stopped_job_ids.append(job.job_id)
                self.terminate_job(job.job_id)

        status_query_error = ''
        try:
            state = await self.get_queue_state(store_code, environment)
        except OutcallPlatformStatusError as exc:
            status_query_error = exc.detail or exc.message
            store = self.split_service.get_store(store_code)
            fallback_batches = []
            for job_id in stopped_job_ids:
                job = self.jobs.get(job_id)
                if job:
                    fallback_batches.extend(job.stopped_batches)
            fallback_batches = list(dict.fromkeys(fallback_batches))
            state = {
                'store_code': store_code,
                'store_name': store.store_name,
                'environment': environment,
                'is_paused': True,
                'completed_batches': [],
                'running_batches': [],
                'pending_batches': [
                    {
                        'task_name': f"{store.store_name}{datetime.now().strftime('%m%d')}-{batch}",
                        'filename': '',
                        'batch_key': str(batch),
                        'row_count': 0,
                    }
                    for batch in fallback_batches
                ],
            }
        stopped_batches = [item['task_name'] for item in state['pending_batches']]
        for job_id in stopped_job_ids:
            job = self.jobs.get(job_id)
            if job:
                job.stopped_batches = stopped_batches
                self.repository.save_job(job)

        completed_names = [item['task_name'] for item in state['completed_batches']]
        running_names = [item['task_name'] for item in state['running_batches']]
        parts = [f"已经停止{state['store_name']}后续未外呼批次调度"]
        if completed_names:
            parts.append(f"当前已完成：{'、'.join(completed_names)}")
        if running_names:
            parts.append(
                f"平台当前仍在外呼：{'、'.join(running_names)}；如需立即停止请联系研发老师"
            )
        if stopped_batches:
            parts.append(f"已停止未推出批次：{'、'.join(stopped_batches)}")
        elif not running_names:
            parts.append('当前没有未推出批次')
        if status_query_error:
            parts.append(f'平台状态读取失败：{status_query_error}；本地后续调度已停止')
        state.update(
            {
                'status': 'paused',
                'stopped_job_ids': stopped_job_ids,
                'stopped_batches': stopped_batches,
                'requires_manual_platform_stop': bool(running_names),
                'status_query_error': status_query_error,
                'message': '；'.join(parts),
            }
        )
        return state

    async def get_queue_state(
        self, store_code: str, environment: str
    ) -> dict[str, Any]:
        environment = self._validate_environment(environment)
        store = self.split_service.get_store(store_code)
        tenant, config, _, _, _ = self._load_tenant(store.store_name, environment)
        env_config = tenant.get_env_config(environment)
        if not env_config.username or not env_config.password or not env_config.robot_id:
            raise OutcallPlatformStatusError(f'{tenant.name} {environment} 配置不完整')

        try:
            split_job = self._recover_split_job_from_outputs(
                store_code, f'state_{datetime.now().strftime("%Y%m%d")}_{store_code}'
            )
            formal_outputs = [
                output for output in split_job.outputs if output.batch_key != 'test'
            ]
        except SplitJobNotFoundError:
            formal_outputs = []

        from src.api import get_task_status, login
        import aiohttp

        async with aiohttp.ClientSession() as session:
            token = await login(
                session,
                config.base_url,
                tenant,
                env_config,
                client_id=config.settings.client_id,
            )
            if not token:
                raise OutcallPlatformStatusError(f'{tenant.name} 登录失败')
            result = await get_task_status(
                session,
                config.base_url,
                tenant,
                token,
                env_config,
                env_config.robot_id,
            )
        if not result:
            raise OutcallPlatformStatusError(f'{tenant.name} 任务状态查询失败')

        task_prefix = f"{tenant.name}{datetime.now().strftime('%m%d')}-"
        records = result.get('data', {}).get('records', []) or []
        platform_tasks = [
            record
            for record in records
            if str(record.get('taskName') or '').startswith(task_prefix)
            and (
                not record.get('createUserName')
                or record.get('createUserName') == env_config.username
            )
        ]
        platform_tasks.sort(key=lambda item: self._task_name_sort_key(item.get('taskName', '')))
        completed = [
            self._platform_task_item(record)
            for record in platform_tasks
            if record.get('state') == 2
        ]
        running = [
            self._platform_task_item(record)
            for record in platform_tasks
            if record.get('state') == 1
        ]
        existing_names = {str(record.get('taskName') or '') for record in platform_tasks}
        pending = []
        for output in sorted(formal_outputs, key=self._output_sort_key):
            batch_name = self._extract_batch_name(Path(output.path).stem, tenant.name)
            task_name = tenant.task_name_template.format(
                tenant=tenant.name,
                batch=batch_name,
                date=datetime.now().strftime('%m%d'),
            )
            if task_name not in existing_names:
                pending.append(
                    {
                        'task_name': task_name,
                        'filename': output.filename,
                        'batch_key': output.batch_key,
                        'row_count': output.row_count,
                    }
                )
        return {
            'store_code': store_code,
            'store_name': tenant.name,
            'environment': environment,
            'is_paused': self.repository.is_queue_paused(
                store_code, environment, self._run_date()
            ),
            'completed_batches': completed,
            'running_batches': running,
            'pending_batches': pending,
        }

    def _platform_task_item(self, record: dict[str, Any]) -> dict[str, Any]:
        return {
            'task_name': str(record.get('taskName') or ''),
            'task_id': str(record.get('id') or ''),
            'state': record.get('state'),
            'actual_count': int(record.get('actualCnt') or 0),
            'expected_count': int(record.get('expectCnt') or 0),
            'created_at': str(record.get('createTime') or ''),
            'updated_at': str(record.get('modifyTime') or ''),
        }

    def _task_name_sort_key(self, task_name: str):
        suffix = str(task_name).rsplit('-', 1)[-1]
        if suffix == '测试':
            return (0, 0)
        return (1, int(suffix)) if suffix.isdigit() else (2, suffix)

    def _has_running_job(self, store_code: str, environment: str) -> bool:
        return any(
            job.store_code == store_code
            and job.environment == environment
            and job.status == 'running'
            and not job.stop_requested
            for job in self.jobs.values()
        )

    def _run_date(self, timestamp: str = '') -> str:
        if timestamp:
            try:
                return datetime.fromisoformat(timestamp).strftime('%Y%m%d')
            except ValueError:
                pass
        return datetime.now().strftime('%Y%m%d')

    def _validate_environment(self, environment: str) -> str:
        if environment not in BASE_URLS:
            raise InvalidOutcallEnvironmentError(environment)
        return environment

    def _validate_mode(self, mode: str) -> str:
        if mode not in ('test', 'formal'):
            raise InvalidOutcallModeError(mode)
        return mode

    def _get_split_job(self, store_code: str, split_job_id: str):
        split_job = self.split_service.get_job(split_job_id)
        if not split_job:
            split_job = self._recover_split_job_from_outputs(store_code, split_job_id)
        if split_job.store_code != store_code:
            raise SplitJobNotFoundError(f'{split_job_id} does not belong to {store_code}')
        if split_job.status != 'completed':
            raise SplitJobNotReadyError(split_job.status)
        return split_job

    def _recover_split_job_from_outputs(self, store_code: str, split_job_id: str):
        store = self.split_service.get_store(store_code)
        date_dir = self.split_service.split_root / datetime.now().strftime('%y%m%d')
        store_dir = date_dir / store.store_name
        scan_dir = store_dir if store_dir.exists() else date_dir
        outputs = []
        if scan_dir.exists():
            for path in scan_dir.glob(f'{store.file_prefix}-*.xlsx'):
                if path.is_file() and not path.name.startswith('~$'):
                    outputs.append(self._output_from_file(path, store.file_prefix))
        outputs.sort(key=self._output_sort_key)
        if not outputs:
            raise SplitJobNotFoundError(f'{split_job_id}; 未找到当天分割输出文件：{store_dir}')
        source_path = self.split_service.split_root / f'{store.file_prefix}-模板.xlsx'
        source_file = SplitSourceFile(
            filename=source_path.name,
            path=str(source_path),
            size=source_path.stat().st_size if source_path.exists() else 0,
            updated_at=datetime.fromtimestamp(source_path.stat().st_mtime).isoformat() if source_path.exists() else '',
        )
        return SplitJob(
            job_id=split_job_id or f'recovered_{datetime.now().strftime("%Y%m%d")}_{store_code}',
            status='completed',
            store_code=store.store_code,
            store_name=store.store_name,
            source_file=source_file,
            output_dir=str(scan_dir),
            script_name=store.script_name,
            outputs=outputs,
            total_rows=sum(output.row_count for output in outputs),
            valid_rows=sum(output.row_count for output in outputs),
            invalid_rows=0,
            created_at=datetime.now().isoformat(),
        )

    def _output_from_file(self, path: Path, file_prefix: str) -> SplitOutputFile:
        stem = path.stem
        batch_name = self._extract_batch_name(stem, file_prefix)
        if batch_name == '测试':
            batch_key = 'test'
            display_name = '测试批次'
        else:
            batch_key = f'formal{batch_name}' if str(batch_name).isdigit() else f'formal_{batch_name}'
            display_name = f'正式批次 {batch_name}'
        return SplitOutputFile(
            batch_key=batch_key,
            batch_name=display_name,
            filename=path.name,
            path=str(path),
            row_count=self._count_excel_rows(path),
            columns=[],
            preview_rows=[],
        )

    def _output_sort_key(self, output: SplitOutputFile):
        if output.batch_key == 'test':
            return (0, 0)
        match = __import__('re').search(r'(\d+)$', output.batch_key)
        return (1, int(match.group(1)) if match else 9999)

    def _count_excel_rows(self, path: Path) -> int:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            return sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(cell is not None for cell in row))
        finally:
            wb.close()

    def _load_tenant(self, store_name: str, environment: str):
        venv_site_packages = self.source_root / '.venv' / 'Lib' / 'site-packages'
        if venv_site_packages.exists() and str(venv_site_packages) not in sys.path:
            sys.path.insert(0, str(venv_site_packages))
        if str(self.source_root) not in sys.path:
            sys.path.insert(0, str(self.source_root))
        from src.config import ExcelFile, load_config
        from src.tenant_processor import TenantStatus, process_tenant

        config = load_config(str(self.config_path))
        config.environment = environment
        config.base_url = BASE_URLS[environment]
        tenant = next((item for item in config.tenants if item.name == store_name), None)
        if tenant is None:
            raise OutcallTenantNotFoundError(store_name)
        return tenant, config, ExcelFile, TenantStatus, process_tenant

    def _select_files(self, split_job: Any, tenant_name: str, mode: str, excel_cls: Any):
        if mode == 'test':
            outputs = [output for output in split_job.outputs if output.batch_key == 'test']
        else:
            outputs = [output for output in split_job.outputs if output.batch_key != 'test']
        if not outputs:
            raise OutcallFileNotFoundError(mode)

        files = []
        for output in outputs:
            path = Path(output.path)
            if not path.exists() or not path.is_file():
                raise OutcallFileNotFoundError(str(path))
            batch_name = self._extract_batch_name(path.stem, tenant_name)
            files.append((output, excel_cls(
                file_path=path,
                tenant_name=tenant_name,
                batch_name=batch_name,
                tenant_prefix=tenant_name,
            )))
        return files

    def _assert_no_duplicate_outcall_files(
        self,
        tenant_name: str,
        files: list[tuple[Any, Any]],
        file_signatures: dict[str, tuple[tuple[str, ...], ...]],
    ):
        for job_id, job in list(self.jobs.items()):
            self._sync_status(job_id)
            if job.environment != 'prod' or job.store_name != tenant_name or job.status == 'terminated':
                continue

            existing_files = {item.filename for item in job.files}
            if not existing_files:
                continue

            existing_signatures = self.file_signatures.get(job_id)
            if existing_signatures is None:
                existing_signatures = self._job_file_signatures(job)
                self.file_signatures[job_id] = existing_signatures

            for _, file in files:
                filename = file.file_path.name
                if filename not in existing_files:
                    continue
                if existing_signatures.get(filename) == file_signatures.get(filename):
                    raise DuplicateOutcallFileError(f'{filename}文件已被外呼禁止重复推出')

    def _job_file_signatures(self, job: OutcallJob) -> dict[str, tuple[tuple[str, ...], ...]]:
        signatures = {}
        for file in job.files:
            path = Path(file.path)
            if path.exists() and path.is_file():
                signatures[file.filename] = self._first_rows_signature(path)
        return signatures

    def _first_rows_signature(self, path: Path, limit: int = 5) -> tuple[tuple[str, ...], ...]:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(cell is not None for cell in row):
                    continue
                rows.append(tuple(self._cell_signature(cell) for cell in row))
                if len(rows) >= limit:
                    break
            return tuple(rows)
        finally:
            wb.close()

    def _cell_signature(self, value: Any) -> str:
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.isoformat()
        return str(value).strip()

    def _extract_batch_name(self, stem: str, tenant_name: str) -> str:
        prefix = f'{tenant_name}-'
        if stem.startswith(prefix):
            return stem[len(prefix):]
        if '-' in stem:
            return stem.split('-', 1)[1]
        return stem


def job_to_dict(job: OutcallJob) -> dict[str, Any]:
    return asdict(job)
