from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ENTRY_PATTERN = re.compile(r'\[([^\]]+)\]([^\(]+)\(最高([\d.]+)%\)')
COLORS = {
    'bg': '#ffffff',
    'surface': '#f7f8fa',
    'border': '#dfe3e8',
    'text': '#15181d',
    'muted': '#667085',
    'blue': '#2563eb',
    'blue_soft': '#eaf2ff',
    'green': '#14866d',
    'green_soft': '#e7f6f1',
    'amber': '#c27608',
    'amber_soft': '#fff3df',
    'red': '#c83d3d',
    'red_soft': '#fdecec',
}


def parse_excel(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return {}

    days: dict[str, list[dict[str, Any]]] = {}
    for column_index, raw_day in enumerate(rows[0]):
        if raw_day in (None, ''):
            continue
        entries = []
        for row in rows[1:]:
            if column_index >= len(row) or row[column_index] in (None, ''):
                continue
            match = ENTRY_PATTERN.search(str(row[column_index]))
            if match:
                entries.append({
                    'id': match.group(1).strip(),
                    'name': match.group(2).strip(),
                    'pct': float(match.group(3)),
                })
        days[str(raw_day).strip()] = entries
    return days


def compute_stats(days: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    day_labels = list(days)
    day_count = len(day_labels)
    counts = {day: len(entries) for day, entries in days.items()}
    total = sum(counts.values())
    maximum = max(counts.values(), default=0)
    maximum_days = [day for day, count in counts.items() if count == maximum]

    frequencies = defaultdict(lambda: {'name': '', 'id': '', 'days': []})
    for day, entries in days.items():
        for entry in entries:
            key = entry['id']
            frequencies[key]['id'] = key
            frequencies[key]['name'] = entry['name']
            frequencies[key]['days'].append(day)

    frequent = [item for item in frequencies.values() if len(item['days']) >= 3]
    frequent.sort(key=lambda item: (-len(item['days']), item['name']))
    top3 = {
        day: sorted(entries, key=lambda item: -item['pct'])[:3]
        for day, entries in days.items()
    }
    sorted_counts = sorted(counts.values())
    if len(sorted_counts) >= 3:
        low_cut = sorted_counts[max(0, len(sorted_counts) // 3 - 1)]
        high_cut = sorted_counts[max(0, (2 * len(sorted_counts)) // 3 - 1)]
    else:
        low_cut = high_cut = sorted_counts[-1] if sorted_counts else 0

    def bar_tone(count: int) -> str:
        if count > high_cut:
            return 'red'
        if count > low_cut:
            return 'amber'
        return 'green'

    return {
        'day_labels': day_labels,
        'n_days': day_count,
        'counts': counts,
        'total': total,
        'avg': round(total / day_count, 1) if day_count else 0,
        'max_count': maximum,
        'max_days': maximum_days,
        'freq_list': frequent,
        'full_week_count': sum(
            1 for item in frequencies.values()
            if day_count and len(item['days']) == day_count
        ),
        'top3': top3,
        'bar_tones': {day: bar_tone(count) for day, count in counts.items()},
    }


def _font(size: int, bold: bool = False):
    candidates = [
        Path(r'C:\Windows\Fonts\msyhbd.ttc' if bold else r'C:\Windows\Fonts\msyh.ttc'),
        Path(r'C:\Windows\Fonts\simhei.ttf'),
        Path('/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc' if bold else '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    return ImageFont.load_default()


def _short_day(value: str) -> str:
    match = re.match(r'(\d+)月(\d+)[号日]?', value)
    return f'{int(match.group(1))}/{int(match.group(2))}' if match else value


def _text(draw, xy, value, size=18, color='text', bold=False):
    draw.text(xy, str(value), font=_font(size, bold), fill=COLORS.get(color, color))


def _card(draw, box, fill='surface'):
    draw.rounded_rectangle(box, radius=6, fill=COLORS[fill], outline=COLORS['border'], width=1)


def render_weekly(
    stats: dict[str, Any],
    output_path: str | Path,
    title: str = '外呼机器人用量周报',
    subtitle: str = '',
) -> Path:
    width = 1200
    padding = 48
    frequent_rows = max(1, len(stats['freq_list']))
    height = 690 + frequent_rows * 42 + max(1, stats['n_days']) * 42
    image = Image.new('RGB', (width, height), COLORS['bg'])
    draw = ImageDraw.Draw(image)
    y = padding
    draw.rectangle((padding, y, padding + 5, y + 56), fill=COLORS['blue'])
    _text(draw, (padding + 18, y), title, 32, bold=True)
    _text(draw, (padding + 18, y + 40), subtitle, 16, 'muted')
    y += 92

    labels = [
        ('全周超量总次数', stats['total'], 'blue_soft', 'blue'),
        ('日均超量机器人数', stats['avg'], 'green_soft', 'green'),
        ('单日最高超量数', stats['max_count'], 'amber_soft', 'amber'),
        ('连续全周超标机器人', stats['full_week_count'], 'red_soft', 'red'),
    ]
    gap = 14
    card_width = (width - padding * 2 - gap * 3) / 4
    for index, (label, value, fill, color) in enumerate(labels):
        left = padding + index * (card_width + gap)
        _card(draw, (left, y, left + card_width, y + 96), fill)
        _text(draw, (left + 16, y + 14), label, 15, color)
        _text(draw, (left + 16, y + 43), value, 30, color, bold=True)
    y += 126

    _text(draw, (padding, y), '1. 每日超量机器人数量', 22, bold=True)
    y += 38
    chart_height = 150
    baseline = y + chart_height
    days = stats['day_labels']
    maximum = max(stats['counts'].values(), default=1) or 1
    slot_width = (width - padding * 2) / max(1, len(days))
    for index, day in enumerate(days):
        count = stats['counts'][day]
        bar_height = (count / maximum) * 116
        center = padding + slot_width * (index + 0.5)
        tone = stats['bar_tones'][day]
        draw.rounded_rectangle(
            (center - 28, baseline - bar_height, center + 28, baseline),
            radius=4,
            fill=COLORS[tone],
        )
        _text(draw, (center - 9, baseline - bar_height - 26), count, 17, bold=True)
        _text(draw, (center - 20, baseline + 9), _short_day(day), 15, 'muted')
    draw.line((padding, baseline, width - padding, baseline), fill=COLORS['border'])
    y = baseline + 56

    _text(draw, (padding, y), '2. 出现3次及以上的机器人名单', 22, bold=True)
    y += 38
    _text(draw, (padding, y), '机器人', 16, 'muted')
    _text(draw, (padding + 470, y), '超标天数', 16, 'muted')
    _text(draw, (padding + 650, y), '出现日期', 16, 'muted')
    y += 30
    draw.line((padding, y, width - padding, y), fill=COLORS['border'])
    y += 12
    if not stats['freq_list']:
        _text(draw, (padding, y), '本周期暂无出现3次及以上的机器人', 16, 'muted')
        y += 42
    else:
        for item in stats['freq_list']:
            day_total = len(item['days'])
            color = 'red' if day_total == stats['n_days'] else ('amber' if day_total >= max(3, stats['n_days'] - 2) else 'muted')
            _text(draw, (padding, y), f"[{item['id']}] {item['name']}", 16)
            _text(draw, (padding + 470, y), f'{day_total}天', 16, color, bold=True)
            _text(draw, (padding + 650, y), '、'.join(_short_day(day) for day in item['days']), 15, color)
            y += 42

    y += 18
    _text(draw, (padding, y), '3. 每日超量 Top3 机器人', 22, bold=True)
    y += 38
    for day in days:
        entries = stats['top3'][day]
        summary = '  |  '.join(
            f"{index + 1}. {entry['name']} {entry['pct']:.2f}%"
            for index, entry in enumerate(entries)
        ) or '无超量机器人'
        _text(draw, (padding, y), _short_day(day), 16, bold=True)
        _text(draw, (padding + 100, y), summary, 15, 'muted')
        y += 42

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    image.crop((0, 0, width, min(height, y + padding))).save(output, 'PNG')
    return output


def render_daily(
    stats: dict[str, Any],
    output_path: str | Path,
    title: str = '外呼机器人用量日报',
) -> Path:
    day = stats['day_labels'][0] if stats['day_labels'] else '目标日期'
    entries = stats['top3'].get(day, [])
    all_entries = sorted(stats.get('all_entries', []), key=lambda item: -item['pct'])
    height = 330 + max(1, len(all_entries)) * 48
    image = Image.new('RGB', (1200, height), COLORS['bg'])
    draw = ImageDraw.Draw(image)
    _text(draw, (48, 44), title, 32, bold=True)
    _text(draw, (48, 88), f'统计日期：{day}', 17, 'muted')
    _card(draw, (48, 132, 344, 236), 'red_soft')
    _text(draw, (68, 150), '超量机器人数', 16, 'red')
    _text(draw, (68, 180), len(all_entries), 34, 'red', bold=True)
    _card(draw, (362, 132, 658, 236), 'amber_soft')
    _text(draw, (382, 150), '最高用量', 16, 'amber')
    _text(draw, (382, 180), f"{entries[0]['pct']:.2f}%" if entries else '0.00%', 34, 'amber', bold=True)
    _card(draw, (676, 132, 1152, 236), 'blue_soft')
    _text(draw, (696, 150), '最高用量机器人', 16, 'blue')
    _text(draw, (696, 184), entries[0]['name'] if entries else '无', 24, 'blue', bold=True)
    y = 278
    _text(draw, (48, y), '超量明细', 22, bold=True)
    y += 42
    if not all_entries:
        _text(draw, (48, y), '当日没有机器人超过配置限额', 17, 'green')
    else:
        for index, entry in enumerate(all_entries, start=1):
            fill = 'red_soft' if index == 1 else 'surface'
            _card(draw, (48, y - 8, 1152, y + 34), fill)
            _text(draw, (68, y), f"{index}. [{entry['id']}] {entry['name']}", 16, bold=index == 1)
            _text(draw, (994, y), f"{entry['pct']:.2f}%", 17, 'red', bold=True)
            y += 48

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    image.save(output, 'PNG')
    return output


def write_stats_json(stats: dict[str, Any], output_path: str | Path) -> Path:
    payload = {
        'counts': stats['counts'],
        'total': stats['total'],
        'avg': stats['avg'],
        'freq_list': [
            {'id': item['id'], 'name': item['name'], 'days': item['days']}
            for item in stats['freq_list']
        ],
        'top3': {
            day: [
                {'id': entry['id'], 'name': entry['name'], 'pct': entry['pct']}
                for entry in entries
            ]
            for day, entries in stats['top3'].items()
        },
    }
    path = Path(output_path)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    return path
