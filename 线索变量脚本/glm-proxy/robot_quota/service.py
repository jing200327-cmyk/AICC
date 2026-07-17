from __future__ import annotations

import ast
import asyncio
import importlib.util
import json
import os
import re
import subprocess
import sys
import uuid
from collections import Counter
from dataclasses import asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import RobotQuotaJob
from .report_renderer import compute_stats, parse_excel, render_daily, render_weekly, write_stats_json


class RobotQuotaError(Exception):
    code = 'ROBOT_QUOTA_ERROR'
    message = 'Robot quota operation failed'

    def __init__(self, detail: str = ''):
        super().__init__(detail or self.message)
        self.detail = detail


class InvalidRobotQuotaDateError(RobotQuotaError):
    code = 'INVALID_ROBOT_QUOTA_DATE'
    message = 'Report date must be YYMMDD'


class RobotQuotaRawDataError(RobotQuotaError):
    code = 'ROBOT_QUOTA_RAW_DATA_ERROR'
    message = 'Robot quota raw call data is unavailable'


class RobotQuotaOutputNotFoundError(RobotQuotaError):
    code = 'ROBOT_QUOTA_OUTPUT_NOT_FOUND'
    message = 'Robot quota output does not exist'


class RobotQuotaService:
    def __init__(
        self,
        daily_project_root: str | Path,
        output_root: str | Path,
        daily_runner: Callable[[str, str], None] | None = None,
        robot_id_resolver: Callable[[dict[str, Any], str, pd.DataFrame], dict[str, str]] | None = None,
    ):
        self.daily_project_root = Path(daily_project_root)
        self.output_root = Path(output_root)
        self.output_root.mkdir(parents=True, exist_ok=True)
        self.daily_runner = daily_runner or self._run_daily_chain
        self.robot_id_resolver = robot_id_resolver or self._resolve_robot_ids_from_platform
        self.jobs: dict[str, RobotQuotaJob] = {}
        self.tasks: dict[str, asyncio.Task] = {}

    def start_job(self, report_date: str | None = None) -> RobotQuotaJob:
        report_date = self._validate_date(report_date)
        now = datetime.now().isoformat()
        job = RobotQuotaJob(
            job_id=f'robot_quota_{report_date}_{uuid.uuid4().hex[:10]}',
            report_date=report_date,
            status='running',
            created_at=now,
            updated_at=now,
        )
        self.jobs[job.job_id] = job
        self.tasks[job.job_id] = asyncio.create_task(self._run_job(job.job_id))
        return job

    def get_job(self, job_id: str) -> RobotQuotaJob | None:
        return self.jobs.get(job_id)

    async def _run_job(self, job_id: str) -> None:
        job = self.jobs[job_id]
        try:
            result = await asyncio.to_thread(self.generate, job.report_date)
            job.status = 'completed'
            job.source_mode = result['source_mode']
            job.workbook_path = result['workbook_path']
            job.robots = result['robots']
            job.warnings = result['warnings']
        except RobotQuotaError as exc:
            job.status = 'failed'
            job.error = {'code': exc.code, 'message': exc.message, 'detail': exc.detail}
        except Exception as exc:
            job.status = 'failed'
            job.error = {
                'code': RobotQuotaError.code,
                'message': RobotQuotaError.message,
                'detail': str(exc),
            }
        finally:
            job.completed_at = datetime.now().isoformat()
            job.updated_at = job.completed_at

    def generate(self, report_date: str | None = None, prepare_raw: bool = True) -> dict[str, Any]:
        report_date = self._validate_date(report_date)
        accounts = [
            account for account in self._load_accounts()
            if self._account_active_on(account, report_date)
        ]
        if not accounts:
            raise RobotQuotaRawDataError('No active ACCOUNTS are configured for the selected date')

        reports_complete = self._daily_reports_complete(accounts, report_date)
        source_mode = 'daily_report_reuse' if reports_complete else 'raw_file_reuse'
        has_missing_raw = any(not self._call_path(account, report_date).exists() for account in accounts)
        if prepare_raw and (not reports_complete or has_missing_raw):
            for account in accounts:
                if (
                    not self._account_reports_complete(account, report_date)
                    or not self._call_path(account, report_date).exists()
                ):
                    self.daily_runner(report_date, str(account.get('name') or '').strip())
            source_mode = 'daily_chain'

        missing = [
            self._call_path(account, report_date)
            for account in accounts
            if not self._call_path(account, report_date).exists()
        ]
        if missing:
            raise RobotQuotaRawDataError(
                'Missing raw call files: ' + ', '.join(path.name for path in missing)
            )

        robot_rows: list[dict[str, Any]] = []
        warnings: list[str] = []
        cache = self._load_robot_id_cache()
        for account in accounts:
            account_rows, account_warnings = self._process_account(
                account,
                report_date,
                self._call_path(account, report_date),
                cache,
            )
            robot_rows.extend(account_rows)
            warnings.extend(account_warnings)
        self._save_robot_id_cache(cache)

        robot_rows.sort(key=lambda item: (-item['usage_rate'], item['store_name'], item['robot_name']))
        day_dir = self.output_root / 'data' / report_date
        day_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = day_dir / f'外呼机器人用量_{report_date}.xlsx'
        self._write_usage_workbook(workbook_path, report_date, robot_rows)

        result = {
            'report_date': report_date,
            'generated_at': datetime.now().isoformat(),
            'source_mode': source_mode,
            'robots': robot_rows,
            'robot_count': len(robot_rows),
            'over_quota_count': sum(1 for item in robot_rows if item['is_over_quota']),
            'workbook_name': workbook_path.name,
            'workbook_path': str(workbook_path),
            'warnings': warnings,
        }
        self._result_path(report_date).write_text(
            json.dumps(result, ensure_ascii=False, indent=2),
            encoding='utf-8',
        )
        self._write_run_log(result)
        return result

    def get_preview(self, report_date: str | None = None) -> dict[str, Any]:
        report_date = self._validate_date(report_date)
        result_path = self._result_path(report_date)
        workbook_path = self._workbook_path(report_date)
        if result_path.exists() and workbook_path.exists():
            result = json.loads(result_path.read_text(encoding='utf-8'))
            result['workbook_path'] = str(workbook_path)
            return result
        if not workbook_path.exists():
            raise RobotQuotaOutputNotFoundError(f'{report_date} robot quota workbook does not exist')
        return self._preview_from_workbook(report_date, workbook_path)

    def generate_daily_report_image(self, report_date: str | None = None) -> dict[str, Any]:
        report_date = self._validate_date(report_date)
        workbook_path = self._workbook_path(report_date)
        if not workbook_path.exists():
            raise RobotQuotaOutputNotFoundError(f'{report_date} robot quota workbook does not exist')
        days = parse_excel(workbook_path)
        stats = compute_stats(days)
        first_day = stats['day_labels'][0] if stats['day_labels'] else ''
        stats['all_entries'] = days.get(first_day, [])
        output_dir = self.output_root / 'reports' / 'daily' / report_date
        output_dir.mkdir(parents=True, exist_ok=True)
        image_path = output_dir / f'外呼机器人用量日报_{report_date}.png'
        stats_path = output_dir / f'外呼机器人用量日报_{report_date}_stats.json'
        render_daily(stats, image_path)
        write_stats_json(stats, stats_path)
        return {
            'report_date': report_date,
            'image_name': image_path.name,
            'image_path': str(image_path),
            'stats_path': str(stats_path),
        }

    def generate_weekly_report_image(self, period_start: str, period_end: str) -> dict[str, Any]:
        period_start = self._validate_date(period_start)
        period_end = self._validate_date(period_end)
        dates = self._date_range(period_start, period_end)
        if not dates or len(dates) > 7:
            raise InvalidRobotQuotaDateError('Weekly period must contain between 1 and 7 days')
        missing = [day for day in dates if not self._workbook_path(day).exists()]
        if missing:
            raise RobotQuotaOutputNotFoundError(
                'Missing daily robot quota workbooks: ' + ', '.join(missing)
            )

        output_dir = self.output_root / 'reports' / 'weekly' / f'{period_start}_{period_end}'
        output_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = output_dir / f'外呼机器人用量周报数据_{period_start}_{period_end}.xlsx'
        self._write_weekly_input(workbook_path, dates)
        days = parse_excel(workbook_path)
        stats = compute_stats(days)
        image_path = output_dir / f'外呼机器人用量周报_{period_start}_{period_end}.png'
        stats_path = output_dir / f'外呼机器人用量周报_{period_start}_{period_end}_stats.json'
        subtitle = f"统计周期：{stats['day_labels'][0]} - {stats['day_labels'][-1]}" if stats['day_labels'] else ''
        render_weekly(stats, image_path, subtitle=subtitle)
        write_stats_json(stats, stats_path)
        return {
            'period_start': period_start,
            'period_end': period_end,
            'days': list(days),
            'workbook_path': str(workbook_path),
            'image_name': image_path.name,
            'image_path': str(image_path),
            'stats_path': str(stats_path),
        }

    def daily_image_path(self, report_date: str | None = None) -> Path:
        report_date = self._validate_date(report_date)
        path = self.output_root / 'reports' / 'daily' / report_date / f'外呼机器人用量日报_{report_date}.png'
        if not path.exists():
            raise RobotQuotaOutputNotFoundError(path.name)
        return path

    def weekly_image_path(self, period_start: str, period_end: str) -> Path:
        period_start = self._validate_date(period_start)
        period_end = self._validate_date(period_end)
        path = self.output_root / 'reports' / 'weekly' / f'{period_start}_{period_end}' / f'外呼机器人用量周报_{period_start}_{period_end}.png'
        if not path.exists():
            raise RobotQuotaOutputNotFoundError(path.name)
        return path

    def workbook_path(self, report_date: str | None = None) -> Path:
        report_date = self._validate_date(report_date)
        path = self._workbook_path(report_date)
        if not path.exists():
            raise RobotQuotaOutputNotFoundError(path.name)
        return path

    def _process_account(
        self,
        account: dict[str, Any],
        report_date: str,
        call_path: Path,
        cache: dict[str, dict[str, str]],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        account_name = str(account.get('name') or '').strip()
        frame = pd.read_excel(call_path, engine='openpyxl', dtype=str)
        robot_column = self._column(frame, ['机器人', '机器人名称'])
        date_column = self._column(frame, ['结束时间', '开始时间'])
        status_column = self._column(frame, ['通话详情', '通话状态'])
        frame[robot_column] = frame[robot_column].fillna('').astype(str).str.strip()
        frame[date_column] = pd.to_datetime(frame[date_column], errors='coerce')
        selected_date = datetime.strptime(report_date, '%y%m%d').date()
        frame = frame[frame[date_column].dt.date == selected_date].copy()

        required = [
            str(value or '').strip()
            for value in account.get('required_group_values') or []
            if str(value or '').strip()
        ]
        available_robot_names = sorted(
            name for name in frame[robot_column].dropna().astype(str).str.strip().unique()
            if name
        )
        robot_names = required or available_robot_names
        if not required and len(available_robot_names) > 1 and account_name:
            account_matches = [name for name in available_robot_names if account_name in name]
            if len(account_matches) == 1:
                robot_names = account_matches
        if not robot_names and account_name:
            robot_names = [account_name]

        configured_ids = account.get('robot_ids') or {}
        display_names = account.get('group_display_names') or {}
        account_cache = cache.setdefault(account_name, {})
        platform_ids: dict[str, str] = {}
        platform_ids_attempted = False
        warnings = []
        rows = []

        for robot_name in robot_names:
            group = frame[frame[robot_column] == robot_name].copy()
            display_name = str(display_names.get(robot_name) or robot_name or account_name).strip()
            robot_id, id_source = self._configured_or_raw_robot_id(
                account,
                configured_ids,
                account_cache,
                robot_name,
                display_name,
                group,
                allow_account_robot_id=len(robot_names) == 1,
            )
            if not robot_id and not platform_ids_attempted:
                platform_ids_attempted = True
                try:
                    platform_ids = self.robot_id_resolver(account, report_date, frame) or {}
                except Exception as exc:
                    warnings.append(f'{account_name} 机器人ID查询失败：{type(exc).__name__}')
            if not robot_id:
                robot_id = str(platform_ids.get(robot_name) or platform_ids.get(display_name) or '').strip()
                if robot_id:
                    id_source = 'platform_task'
                    account_cache[robot_name] = robot_id
            if not robot_id:
                robot_id = '未识别'
                id_source = 'unresolved'
                warnings.append(f'{account_name}/{display_name} 未识别机器人ID，请检查平台任务列表或配置 robot_ids')

            status = group[status_column].fillna('').astype(str)
            line_limit_mask = status.str.contains('线路限制', na=False)
            call_count = int((~line_limit_mask).sum())
            excluded = int(line_limit_mask.sum())
            quota = self._quota_for(account, robot_name, display_name)
            rate = call_count / quota if quota else 0.0
            rows.append({
                'robot_id': robot_id,
                'robot_name': display_name,
                'raw_robot_name': robot_name,
                'store_name': account_name,
                'call_count': call_count,
                'quota': quota,
                'usage_rate': rate,
                'usage_percent': round(rate * 100, 2),
                'is_over_quota': rate > 1,
                'status': '超量' if rate > 1 else '正常',
                'excluded_line_limit_count': excluded,
                'robot_id_source': id_source,
                'source_file': call_path.name,
            })
        return rows, list(dict.fromkeys(warnings))

    def _configured_or_raw_robot_id(
        self,
        account: dict[str, Any],
        configured_ids: dict[str, Any],
        account_cache: dict[str, str],
        robot_name: str,
        display_name: str,
        group: pd.DataFrame,
        *,
        allow_account_robot_id: bool,
    ) -> tuple[str, str]:
        for candidate in ('机器人ID', '机器人id', 'robotId', 'robot_id'):
            if candidate in group.columns:
                values = [str(value).strip() for value in group[candidate].dropna().unique() if str(value).strip()]
                if values:
                    return Counter(values).most_common(1)[0][0], 'raw_file'
        configured = str(
            configured_ids.get(robot_name)
            or configured_ids.get(display_name)
            or (account.get('robot_id') if allow_account_robot_id else '')
            or ''
        ).strip()
        if configured:
            return configured, 'account_config'
        cached = str(account_cache.get(robot_name) or account_cache.get(display_name) or '').strip()
        return (cached, 'cache') if cached else ('', '')

    def _resolve_robot_ids_from_platform(
        self,
        account: dict[str, Any],
        report_date: str,
        frame: pd.DataFrame,
    ) -> dict[str, str]:
        recorder = self._load_recorder_module()
        token = recorder.login(account)
        report_day = datetime.strptime(report_date, '%y%m%d')
        task_to_robot: dict[str, str] = {}
        page = 1
        while page <= 100:
            payload = {
                'startTime': report_day.strftime('%Y-%m-01T00:00:00'),
                'endTime': report_day.strftime('%Y-%m-%dT23:59:59'),
                'robotId': '',
                'id': '',
                'taskName': '',
                'taskStatus': '',
                'isClue': 1,
                'expLimit': 1,
                'pageSize': 100,
                'currentPage': page,
                'actualPhoneNumber': '',
                'phoneNumber': '',
                'clueId': '',
            }
            status, response = recorder.post_json(
                f'{recorder.BASE_URL}/esl/v2/task/getPageCallTasks',
                headers=recorder.build_headers(token=token, content_type='application/json'),
                payload=payload,
                timeout=30,
                cookies=recorder.build_cookies(account),
            )
            if status >= 400 or response.get('code') != 200:
                raise RobotQuotaError('Platform task metadata query failed')
            data = response.get('data') or {}
            records = data.get('records') or []
            for record in records:
                task_name = str(record.get('taskName') or '').strip()
                robot_id = str(record.get('robotId') or '').strip()
                if task_name and robot_id:
                    task_to_robot[task_name] = robot_id
            page_count = int(data.get('pageCount') or 1)
            if page >= page_count or not records:
                break
            page += 1

        if '机器人' not in frame.columns or '任务名称' not in frame.columns:
            return {}
        mapping = {}
        for robot_name, group in frame.groupby('机器人'):
            ids = [
                task_to_robot.get(str(task).strip())
                for task in group['任务名称'].dropna().unique()
            ]
            ids = [value for value in ids if value]
            if ids:
                mapping[str(robot_name).strip()] = Counter(ids).most_common(1)[0][0]
        return mapping

    def _run_daily_chain(self, report_date: str, store: str) -> None:
        log_dir = self.output_root / 'logs' / report_date
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / f'{self._safe_name(store)}_daily_chain.log'
        env = os.environ.copy()
        env['REPORT_DATE'] = report_date
        env['REPORT_STORE'] = store
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            [sys.executable, str(self.daily_project_root / 'main.py')],
            cwd=str(self.daily_project_root),
            env=env,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
        )
        log_path.write_text((result.stdout or '') + (result.stderr or ''), encoding='utf-8')
        if result.returncode != 0:
            raise RobotQuotaRawDataError(f'{store} daily chain failed; see {log_path}')

    def _write_usage_workbook(self, path: Path, report_date: str, rows: list[dict[str, Any]]) -> None:
        workbook = Workbook()
        skill_sheet = workbook.active
        skill_sheet.title = '超量记录'
        report_day = datetime.strptime(report_date, '%y%m%d')
        skill_sheet.cell(1, 1, f'{report_day.month}月{report_day.day}号')
        over_quota = [item for item in rows if item['is_over_quota']]
        for row_index, item in enumerate(over_quota, start=2):
            skill_sheet.cell(
                row_index,
                1,
                f"• [{item['robot_id']}]{item['robot_name']}(最高{item['usage_percent']:.2f}%)",
            )
        skill_sheet.column_dimensions['A'].width = 58
        skill_sheet.freeze_panes = 'A2'

        detail = workbook.create_sheet('用量明细')
        headers = [
            '机器人ID', '机器人名称', '所属门店', '有效通话通次A', '每日限额',
            '用量B', '状态', '排除线路限制通次', '机器人ID来源', '原始文件',
        ]
        detail.append(headers)
        header_fill = PatternFill('solid', fgColor='2F6FED')
        over_fill = PatternFill('solid', fgColor='FDECEC')
        normal_fill = PatternFill('solid', fgColor='E8F5EF')
        border = Border(
            left=Side(style='thin', color='D6DBE3'),
            right=Side(style='thin', color='D6DBE3'),
            top=Side(style='thin', color='D6DBE3'),
            bottom=Side(style='thin', color='D6DBE3'),
        )
        for cell in detail[1]:
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        for item in rows:
            detail.append([
                item['robot_id'], item['robot_name'], item['store_name'], item['call_count'],
                item['quota'], item['usage_rate'], item['status'], item['excluded_line_limit_count'],
                item['robot_id_source'], item['source_file'],
            ])
            current = detail.max_row
            detail.cell(current, 6).number_format = '0.00%'
            row_fill = over_fill if item['is_over_quota'] else normal_fill
            for cell in detail[current]:
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(vertical='center')
        widths = [20, 38, 18, 18, 14, 14, 12, 20, 18, 42]
        for index, width in enumerate(widths, start=1):
            detail.column_dimensions[chr(64 + index)].width = width
        detail.freeze_panes = 'A2'
        detail.auto_filter.ref = detail.dimensions

        temp = path.with_name(f'{path.stem}.tmp{path.suffix}')
        workbook.save(temp)
        workbook.close()
        temp.replace(path)

    def _write_weekly_input(self, path: Path, dates: list[str]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '超量记录'
        for column, report_date in enumerate(dates, start=1):
            source = load_workbook(self._workbook_path(report_date), read_only=True, data_only=True)
            try:
                source_sheet = source['超量记录']
                for row_index in range(1, source_sheet.max_row + 1):
                    sheet.cell(row_index, column, source_sheet.cell(row_index, 1).value)
            finally:
                source.close()
            sheet.column_dimensions[chr(64 + column)].width = 58
        workbook.save(path)
        workbook.close()

    def _preview_from_workbook(self, report_date: str, path: Path) -> dict[str, Any]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook['用量明细']
            headers = [cell.value for cell in sheet[1]]
            items = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = dict(zip(headers, values))
                items.append({
                    'robot_id': str(row.get('机器人ID') or ''),
                    'robot_name': str(row.get('机器人名称') or ''),
                    'store_name': str(row.get('所属门店') or ''),
                    'call_count': int(row.get('有效通话通次A') or 0),
                    'quota': int(row.get('每日限额') or 0),
                    'usage_rate': float(row.get('用量B') or 0),
                    'usage_percent': round(float(row.get('用量B') or 0) * 100, 2),
                    'status': str(row.get('状态') or ''),
                    'is_over_quota': str(row.get('状态') or '') == '超量',
                    'excluded_line_limit_count': int(row.get('排除线路限制通次') or 0),
                })
        finally:
            workbook.close()
        return {
            'report_date': report_date,
            'source_mode': 'workbook',
            'robots': items,
            'robot_count': len(items),
            'over_quota_count': sum(1 for item in items if item['is_over_quota']),
            'workbook_name': path.name,
            'workbook_path': str(path),
            'warnings': [],
        }

    def _daily_reports_complete(self, accounts: list[dict[str, Any]], report_date: str) -> bool:
        return all(self._account_reports_complete(account, report_date) for account in accounts)

    def _account_reports_complete(self, account: dict[str, Any], report_date: str) -> bool:
        report_dir = self.daily_project_root / 'data' / report_date / '每日报告'
        return all(
            (report_dir / f'{name}_每日报告_{report_date}.txt').exists()
            for name in self._account_report_names(account)
        )

    def _account_report_names(self, account: dict[str, Any]) -> list[str]:
        account_name = str(account.get('name') or '').strip()
        required = [
            str(value or '').strip()
            for value in account.get('required_group_values') or []
            if str(value or '').strip()
        ]
        if not required:
            return [account_name]
        display_names = account.get('group_display_names') or {}
        return [str(display_names.get(value) or f'{account_name}-{value}').strip() for value in required]

    def _load_accounts(self) -> list[dict[str, Any]]:
        path = self.daily_project_root / 'tools' / 'recorder.py'
        if not path.exists():
            raise RobotQuotaRawDataError(f'ACCOUNTS configuration does not exist: {path}')
        try:
            module = ast.parse(path.read_text(encoding='utf-8'))
            assignment = next(
                node for node in module.body
                if isinstance(node, ast.Assign)
                and any(isinstance(target, ast.Name) and target.id == 'ACCOUNTS' for target in node.targets)
            )
            accounts = ast.literal_eval(assignment.value)
        except (StopIteration, SyntaxError, ValueError, OSError) as exc:
            raise RobotQuotaRawDataError(f'Failed to load ACCOUNTS: {type(exc).__name__}') from exc
        return list(accounts) if isinstance(accounts, list) else []

    def _load_recorder_module(self):
        path = self.daily_project_root / 'tools' / 'recorder.py'
        spec = importlib.util.spec_from_file_location(f'robot_quota_recorder_{id(self)}', path)
        if not spec or not spec.loader:
            raise RobotQuotaRawDataError('Cannot load daily recorder module')
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_robot_id_cache(self) -> dict[str, dict[str, str]]:
        path = self.output_root / 'config' / 'robot_ids.json'
        if not path.exists():
            return {}
        try:
            value = json.loads(path.read_text(encoding='utf-8'))
            return value if isinstance(value, dict) else {}
        except (OSError, ValueError):
            return {}

    def _save_robot_id_cache(self, cache: dict[str, dict[str, str]]) -> None:
        path = self.output_root / 'config' / 'robot_ids.json'
        path.parent.mkdir(parents=True, exist_ok=True)
        temp = path.with_suffix('.tmp')
        temp.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding='utf-8')
        temp.replace(path)

    def _write_run_log(self, result: dict[str, Any]) -> None:
        report_date = result['report_date']
        path = self.output_root / 'logs' / report_date / 'robot_quota.log'
        path.parent.mkdir(parents=True, exist_ok=True)
        lines = [
            f"generated_at={result['generated_at']}",
            f"source_mode={result['source_mode']}",
            f"robot_count={result['robot_count']}",
            f"over_quota_count={result['over_quota_count']}",
        ]
        lines.extend(f'warning={warning}' for warning in result['warnings'])
        path.write_text('\n'.join(lines) + '\n', encoding='utf-8')

    def _quota_for(self, account: dict[str, Any], robot_name: str, display_name: str) -> int:
        configured = account.get('robot_quotas') or {}
        value = configured.get(robot_name) or configured.get(display_name) or account.get('daily_call_quota')
        if value not in (None, ''):
            try:
                quota = int(value)
                if quota > 0:
                    return quota
            except (TypeError, ValueError):
                pass
        combined = f'{robot_name}{display_name}'
        return 400 if '番禺' in combined and '售后' in combined else 200

    def _column(self, frame: pd.DataFrame, candidates: list[str]) -> str:
        for candidate in candidates:
            if candidate in frame.columns:
                return candidate
        raise RobotQuotaRawDataError('Raw call file is missing required columns: ' + '/'.join(candidates))

    def _account_active_on(self, account: dict[str, Any], report_date: str) -> bool:
        start = str(account.get('mtd_start_date') or '').strip()
        return not re.fullmatch(r'\d{6}', start) or report_date >= start

    def _call_path(self, account: dict[str, Any], report_date: str) -> Path:
        name = str(account.get('name') or '').strip()
        return self.daily_project_root / 'data' / report_date / '原始数据' / f'{name}-aicc-话单-{report_date}.xlsx'

    def _workbook_path(self, report_date: str) -> Path:
        return self.output_root / 'data' / report_date / f'外呼机器人用量_{report_date}.xlsx'

    def _result_path(self, report_date: str) -> Path:
        return self.output_root / 'data' / report_date / f'外呼机器人用量_{report_date}.json'

    def _date_range(self, period_start: str, period_end: str) -> list[str]:
        start = datetime.strptime(period_start, '%y%m%d').date()
        end = datetime.strptime(period_end, '%y%m%d').date()
        if start > end:
            raise InvalidRobotQuotaDateError(f'{period_start}_{period_end}')
        result = []
        current = start
        while current <= end:
            result.append(current.strftime('%y%m%d'))
            current += timedelta(days=1)
        return result

    def _validate_date(self, value: str | None) -> str:
        value = str(value or datetime.now().strftime('%y%m%d')).strip()
        if not re.fullmatch(r'\d{6}', value):
            raise InvalidRobotQuotaDateError(value)
        try:
            datetime.strptime(value, '%y%m%d')
        except ValueError as exc:
            raise InvalidRobotQuotaDateError(value) from exc
        return value

    def _safe_name(self, value: str) -> str:
        return re.sub(r'[\\/:*?"<>|]+', '_', value).strip() or 'unknown'


def job_to_dict(job: RobotQuotaJob) -> dict[str, Any]:
    return asdict(job)
