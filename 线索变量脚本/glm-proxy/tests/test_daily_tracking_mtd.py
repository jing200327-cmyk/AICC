from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
TOOLS_DIR = APP_DIR.parents[1] / '龙星行报表工具_核心文件_260629' / 'tools'
if str(TOOLS_DIR) not in sys.path:
    sys.path.insert(0, str(TOOLS_DIR))

import daily_summary


def write_tracking(root: Path, company: str, records: dict) -> Path:
    month_dir = root / '2607'
    month_dir.mkdir(parents=True, exist_ok=True)
    path = month_dir / f'{company}.json'
    path.write_text(
        json.dumps(records, ensure_ascii=False),
        encoding='utf-8',
    )
    return path


def summary_value(path: Path, label: str, column: int = 3):
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for row in workbook.active.iter_rows(values_only=True):
            if row[0] == label:
                return row[column - 1]
    finally:
        workbook.close()
    raise AssertionError(f'Missing summary row: {label}')


def record(clues: int, connected: int, effective: int, calls: int) -> dict:
    return {
        '新增线索量': clues,
        '新增线索接通量': connected,
        '有效线索量': effective,
        '呼叫通次': calls,
    }


def write_existing_daily_summary(
    root: Path,
    report_date: str,
    company: str,
    *,
    clues: int,
    connected: int,
    effective: int,
    calls: int,
) -> None:
    summary_dir = root / 'data' / report_date / '汇总表'
    report_dir = root / 'data' / report_date / '每日报告'
    summary_dir.mkdir(parents=True, exist_ok=True)
    report_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['业务场景', '单位', company])
    sheet.append([f'MTD (截止：20{report_date[:2]}-{report_date[2:4]}-{report_date[4:6]})'])
    sheet.append([f'Daily Report (20{report_date[:2]}-{report_date[2:4]}-{report_date[4:6]})'])
    sheet.append(['新增线索量', '条', clues])
    sheet.append(['新增线索接通量', '条', connected])
    sheet.append(['新增线索呼叫通次', '次', calls])
    sheet.append(['已接通通话平均时长（向上取整）', '分钟', 2])
    sheet.append(['接通有效率（=有效线索量/新增线索接通量）', '%', 99.99])
    workbook.save(summary_dir / f'广州售后_日度月度汇总表_{report_date}.xlsx')

    (report_dir / f'{company}_每日报告_{report_date}.txt').write_text(
        f'【{company} {report_date}】推送新增线索量 {clues} 条，其中：\n'
        f'① 接通线索 {connected} 条\n'
        f'- 意向线索（{effective} 条，接通数占比0%）\n',
        encoding='utf-8',
    )


def write_mtd_summary(
    root: Path,
    report_date: str,
    company: str,
    *,
    clues: int,
    connected: int,
    effective: int,
    calls: int,
) -> None:
    path = (
        root / 'data' / report_date / '汇总表'
        / f'广州售后_日度月度汇总表_{report_date}.xlsx'
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['业务场景', '单位', company])
    sheet.append([
        f'MTD (截止：20{report_date[:2]}-{report_date[2:4]}-{report_date[4:6]})'
    ])
    sheet.append(['规模指标'])
    sheet.append(['累计线索量', '条', clues])
    sheet.append(['接通量', '条', connected])
    sheet.append(['有效线索量', '条', effective])
    sheet.append(['呼叫通次', '次', calls])
    sheet.append([
        f'Daily Report (20{report_date[:2]}-{report_date[2:4]}-{report_date[4:6]})'
    ])
    workbook.save(path)


def test_junyi_mtd_uses_month_tracking_file(monkeypatch, tmp_path):
    monkeypatch.setattr(daily_summary, 'TRACKING_DIR', tmp_path / '_tracking')
    tracking_path = write_tracking(
        daily_summary.TRACKING_DIR,
        '骏宜',
        {
            '260701': record(10, 4, 1, 15),
            '260702': record(20, 8, 2, 25),
        },
    )

    output = daily_summary.generate_daily_summary(
        company='骏宜',
        date='260703',
        data={
            'total': 30,
            'jietong_count': 12,
            'yixiang_count': 3,
            'call_count': 35,
        },
        output_dir=tmp_path / 'summary',
        monthly_stats={
            '累计线索量': 999,
            '累计接通量': 999,
            '累计有效线索量': 999,
            '累计呼叫通次': 999,
        },
    )

    assert tracking_path.exists()
    assert summary_value(output, '累计线索量') == 999
    assert summary_value(output, '接通量') == 999
    assert summary_value(output, '有效线索量') == 999
    assert summary_value(output, '呼叫通次') == 999
    tracking = json.loads(tracking_path.read_text(encoding='utf-8'))
    assert tracking['260703']['新增线索接通量'] == 12
    assert tracking['260703']['MTD累计接通量'] == 999


def test_tracking_snapshot_does_not_change_daily_metrics():
    tracking = {
        '260701': record(100, 40, 5, 150),
        '260702': {
            **record(50, 20, 2, 70),
            'MTD累计线索量': 150,
            'MTD累计接通量': 75,
            'MTD累计有效线索量': 8,
            'MTD累计呼叫通次': 260,
        },
    }

    stats, missing = daily_summary.aggregate_tracking_mtd(tracking, '260702')

    assert missing == []
    assert tracking['260702']['新增线索接通量'] == 20
    assert stats == {
        '累计线索量': 150,
        '累计接通量': 75,
        '累计有效线索量': 8,
        '累计呼叫通次': 260,
    }


def test_grouped_mtd_accumulates_each_robot_tracking(monkeypatch, tmp_path):
    monkeypatch.setattr(daily_summary, 'TRACKING_DIR', tmp_path / '_tracking')
    company = '广州龙星行-售后-续保提醒'
    write_tracking(
        daily_summary.TRACKING_DIR,
        company,
        {
            '260701': record(100, 60, 5, 150),
            '260702': record(100, 55, 4, 145),
        },
    )

    output = daily_summary.generate_multi_daily_summary(
        title='广州售后',
        date='260703',
        grouped_data=[{
            'company': company,
            'tracking_company': company,
            'data': {
                'total': 200,
                'jietong_count': 120,
                'yixiang_count': 10,
                'call_count': 300,
            },
            'monthly_stats': {'累计线索量': 500},
        }],
        output_dir=tmp_path / 'summary',
    )

    assert summary_value(output, '累计线索量') == 400
    assert summary_value(output, '接通量') == 235
    assert summary_value(output, '有效线索量') == 19
    assert summary_value(output, '呼叫通次') == 595


def test_tracking_mtd_reports_missing_calendar_dates():
    stats, missing = daily_summary.aggregate_tracking_mtd(
        {
            '260701': record(100, 60, 5, 150),
            '260703': record(200, 120, 10, 300),
            '260801': record(999, 999, 999, 999),
        },
        '260703',
    )

    assert stats['累计线索量'] == 300
    assert missing == ['260702']


def test_missing_tracking_date_is_backfilled_from_existing_daily_files(
    monkeypatch,
    tmp_path,
):
    monkeypatch.setattr(daily_summary, 'TRACKING_DIR', tmp_path / '_tracking')
    company = '广州龙星行-售后-续保提醒'
    tracking_path = write_tracking(
        daily_summary.TRACKING_DIR,
        company,
        {'260701': record(100, 60, 5, 150)},
    )
    write_existing_daily_summary(
        tmp_path,
        '260702',
        company,
        clues=120,
        connected=70,
        effective=9,
        calls=180,
    )

    output = daily_summary.generate_multi_daily_summary(
        title='广州售后',
        date='260703',
        grouped_data=[{
            'company': company,
            'tracking_company': company,
            'data': {
                'total': 130,
                'jietong_count': 80,
                'yixiang_count': 11,
                'call_count': 190,
            },
        }],
        output_dir=tmp_path / 'summary',
    )

    tracking = json.loads(tracking_path.read_text(encoding='utf-8'))
    assert tracking['260702'] == {
        '呼叫通次': 180,
        '新增线索量': 120,
        '新增线索接通量': 70,
        '有效线索量': 9,
        '已接通通话平均时长分钟': 2,
    }
    assert summary_value(output, '累计线索量') == 350
    assert summary_value(output, '接通量') == 210
    assert summary_value(output, '有效线索量') == 25
    assert summary_value(output, '呼叫通次') == 520


def test_missing_tracking_date_without_daily_source_is_not_fabricated(
    monkeypatch,
    tmp_path,
):
    monkeypatch.setattr(daily_summary, 'TRACKING_DIR', tmp_path / '_tracking')
    tracking = {
        '260701': record(100, 60, 5, 150),
        '260703': record(130, 80, 11, 190),
    }

    backfilled, unresolved = daily_summary.backfill_tracking_from_daily_reports(
        '广州龙星行-售后-续保提醒',
        '260703',
        tracking,
    )

    assert backfilled == []
    assert unresolved == ['260702']
    assert '260702' not in tracking


def test_tracking_mtd_uses_existing_summary_delta_for_missing_days(
    monkeypatch,
    tmp_path,
):
    monkeypatch.setattr(daily_summary, 'TRACKING_DIR', tmp_path / '_tracking')
    company = '广州龙星行-售后-续保提醒'
    tracking = {
        '260701': record(0, 0, 0, 0),
        '260702': record(0, 0, 0, 0),
        '260703': record(0, 0, 0, 0),
        '260706': record(100, 53, 3, 171),
    }
    write_mtd_summary(
        tmp_path,
        '260706',
        company,
        clues=300,
        connected=185,
        effective=14,
        calls=507,
    )

    backfilled, unresolved = (
        daily_summary.backfill_tracking_mtd_gaps_from_summaries(
            company,
            '260706',
            tracking,
        )
    )

    assert backfilled == ['260704-260705']
    assert unresolved == []
    adjustment = tracking[daily_summary.MTD_BACKFILL_KEY]['260704-260705']
    assert adjustment['新增线索量'] == 200
    assert adjustment['新增线索接通量'] == 132
    assert adjustment['有效线索量'] == 11
    assert adjustment['呼叫通次'] == 336
    assert adjustment['covered_dates'] == ['260704', '260705']
    assert adjustment['method'] == 'existing_summary_mtd_delta'

    stats, missing = daily_summary.aggregate_tracking_mtd(tracking, '260706')
    assert stats == {
        '累计线索量': 300,
        '累计接通量': 185,
        '累计有效线索量': 14,
        '累计呼叫通次': 507,
    }
    assert missing == []
