import sys
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from daily_report.api import create_router
from daily_report.service import DailyReportService, InvalidReportDateError, REPORT_PREVIEW_GROUPS

REPORT_TOOLS_DIR = next(APP_DIR.parents[1].glob("龙星行报表工具_核心文件_*/tools"))
if str(REPORT_TOOLS_DIR) not in sys.path:
    sys.path.insert(0, str(REPORT_TOOLS_DIR))

from process_clue_report import build_monthly_stats_by_group


def write_summary(root: Path, day: str, summary_name: str = "骏宜", base_dir: Path | None = None) -> Path:
    day_root = base_dir / day if base_dir is not None else root / "data" / day
    summary_dir = day_root / "汇总表"
    summary_dir.mkdir(parents=True, exist_ok=True)
    path = summary_dir / f"{summary_name}_日度月度汇总表_{day}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["业务场景", "单位", summary_name])
    sheet.append([f"MTD (截止：20{day[:2]}-{day[2:4]}-{day[4:]})", "", int(day[-2:])])
    sheet.append([f"Daily Report (20{day[:2]}-{day[2:4]}-{day[4:]})", "", int(day[-2:]) * 10])
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 14
    workbook.save(path)
    workbook.close()
    return path


def write_guangzhou_new_summary(root: Path, day: str, haizhu_value: int, panyu_value: int) -> Path:
    summary_dir = root / "data" / day / "汇总表"
    summary_dir.mkdir(parents=True, exist_ok=True)
    path = summary_dir / f"广州新车_日度月度汇总表_{day}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["业务场景", "单位", "广州龙星行-海珠新车首呼", "广州龙星行-番禺新车首呼"])
    sheet.append([f"MTD (截止：20{day[:2]}-{day[2:4]}-{day[4:]})", "条", haizhu_value, panyu_value])
    sheet.append([f"Daily Report (20{day[:2]}-{day[2:4]}-{day[4:]})", "条", haizhu_value * 10, panyu_value * 10])
    workbook.save(path)
    workbook.close()
    return path


def test_all_store_summary_uses_business_column_order(tmp_path):
    service = DailyReportService(tmp_path)
    report_date = '260720'
    for summary_name in [
        '骏宜',
        '长沙售后',
        '广州售后',
        '玉林新车首呼',
        '翔鹏',
        '南宁新车首呼',
        '长沙',
        '韶关',
    ]:
        write_summary(tmp_path, report_date, summary_name)
    write_guangzhou_new_summary(tmp_path, report_date, haizhu_value=13, panyu_value=29)

    service.generate_all_store_summary(report_date)

    output_path = (
        tmp_path
        / 'data'
        / report_date
        / '汇总表'
        / '0720全门店机器人汇总表.xlsx'
    )
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        headers = [cell.value for cell in workbook.active[1]]
    finally:
        workbook.close()
    assert headers == [
        '业务场景',
        '单位',
        '南宁新车首呼',
        '玉林新车首呼',
        '广州龙星行-海珠新车首呼',
        '广州龙星行-番禺新车首呼',
        '骏宜',
        '韶关',
        '翔鹏',
        '长沙',
        '长沙售后',
        '广州售后',
    ]



def test_grouped_mtd_status_uses_clue_terminal_status():
    clues = pd.DataFrame(
        {
            "线索ID": ["1", "2"],
            "通话状态": ["已接通", "已接通"],
            "线索状态": ["有效", "有效"],
        }
    )
    calls_to_date = pd.DataFrame(
        {
            "线索ID": ["1", "2"],
            "机器人": ["番禺", "番禺"],
            "结束时间": pd.to_datetime(["2026-06-03 10:00:00", "2026-06-03 11:00:00"]),
            "通话状态": ["已接通", "未接通-拒接"],
        }
    )

    stats = build_monthly_stats_by_group(clues, calls_to_date, "机器人")

    assert stats["番禺"]["累计线索量"] == 2
    assert stats["番禺"]["累计接通量"] == 2
    assert stats["番禺"]["累计有效线索量"] == 2
def test_guangzhou_new_supplement_uses_haizhu_account():
    group = next(item for item in REPORT_PREVIEW_GROUPS if item["key"] == "guangzhou_new")

    assert group["stores"] == ["海珠龙星行"]


def test_market_invitation_preview_is_after_guangzhou_new():
    new_index = next(
        index for index, group in enumerate(REPORT_PREVIEW_GROUPS)
        if group['key'] == 'guangzhou_new'
    )
    after_sales = REPORT_PREVIEW_GROUPS[new_index + 1]

    assert after_sales['key'] == 'guangzhou_after_sales'
    assert after_sales['reports'][-1] == '广州龙星行-市场-市场活动邀约'


def test_monthly_store_options_split_guangzhou_new_robots(tmp_path):
    service = DailyReportService(tmp_path)

    stores = service.list_monthly_summary_stores()
    store_codes = [item["code"] for item in stores]

    assert "guangzhou_new" not in store_codes
    assert "guangzhou_new_haizhu" in store_codes
    assert "guangzhou_new_panyu" in store_codes


def test_guangzhou_new_monthly_outputs_are_split_by_robot_column(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    monkeypatch.setattr(
        service,
        "_load_accounts",
        lambda: [{"name": "海珠龙星行", "mtd_start_date": "260603"}],
    )
    write_guangzhou_new_summary(tmp_path, "260603", haizhu_value=13, panyu_value=29)

    result = service.generate_monthly_summaries(
        ["guangzhou_new_haizhu", "guangzhou_new_panyu"],
        report_date="260603",
    )

    assert [item["key"] for item in result["groups"]] == [
        "guangzhou_new_haizhu",
        "guangzhou_new_panyu",
    ]
    expected = {
        "guangzhou_new_haizhu": ("广州龙星行-海珠新车首呼", 13, 130),
        "guangzhou_new_panyu": ("广州龙星行-番禺新车首呼", 29, 290),
    }
    for item in result["groups"]:
        output_path = Path(item["summary"]["path"])
        title, mtd_value, daily_value = expected[item["key"]]
        assert output_path.name == f"{title}_月度横向对比表.xlsx"
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            assert sheet.max_column == 3
            assert sheet.cell(row=1, column=3).value == f"{title}0603"
            assert sheet.cell(row=2, column=3).value == mtd_value
            assert sheet.cell(row=3, column=3).value == daily_value
        finally:
            workbook.close()

        image, filename = service.get_monthly_summary_image(
            result["period"],
            item["key"],
            output_folder=item["output_folder"],
        )
        assert image.startswith(bytes([137, 80, 78, 71]))
        assert filename.endswith("_monthly.png")


def test_split_guangzhou_monthly_groups_share_one_backfill(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    monkeypatch.setattr(
        service,
        "_load_accounts",
        lambda: [{"name": "海珠龙星行", "mtd_start_date": "260603"}],
    )
    write_guangzhou_new_summary(tmp_path, "260603", haizhu_value=13, panyu_value=29)
    backfill_calls = []

    def generate_missing(group, missing_dates, store_dir, source_date, mtd_start_date=None):
        backfill_calls.append((group["key"], tuple(missing_dates), source_date))
        return [
            ("260604", write_guangzhou_new_summary(tmp_path, "260604", haizhu_value=17, panyu_value=31))
        ]

    monkeypatch.setattr(service, "_generate_missing_daily_summaries", generate_missing)

    result = service.generate_monthly_summaries(
        ["guangzhou_new_haizhu", "guangzhou_new_panyu"],
        report_date="260604",
    )

    assert len(result["groups"]) == 2
    assert backfill_calls == [("guangzhou_new_haizhu", ("260604",), "260604")]


def test_recent_month_options_include_current_and_previous_two_months(tmp_path):
    service = DailyReportService(tmp_path)
    service._today = lambda: date(2026, 7, 13)

    months = service.list_monthly_summary_months()

    assert [item["value"] for item in months] == ["2607", "2606", "2605"]
    assert [item["label"] for item in months] == ["7月（当月）", "6月（前一个月）", "5月（前两个月）"]
    assert months[0]["period_end"] == "260713"
    assert months[1]["period_end"] == "260630"
    assert months[2]["period_end"] == "260531"


def test_monthly_status_reports_every_missing_calendar_date(tmp_path):
    service = DailyReportService(tmp_path)
    service._today = lambda: date(2026, 7, 13)
    write_summary(tmp_path, "260601")
    write_summary(tmp_path, "260630")

    status = service.get_monthly_summary_status(["junyi"], target_month="2606")

    assert status["period"] == "260601_260630"
    assert status["target_month"] == "2606"
    assert status["items"][0]["available_dates"] == ["260601", "260630"]
    assert len(status["items"][0]["missing_dates"]) == 28
    assert status["items"][0]["missing_dates"][0] == "260602"
    assert status["items"][0]["missing_dates"][-1] == "260629"
    assert status["items"][0]["output_folder"] == "骏宜_六月_缺失"


def test_custom_period_status_spans_months_and_reports_missing_dates(tmp_path):
    service = DailyReportService(tmp_path)
    write_summary(tmp_path, "260613")
    write_summary(tmp_path, "260713")

    status = service.get_monthly_summary_status(
        ["junyi"],
        period_start="260613",
        period_end="260713",
    )

    assert status["selection_mode"] == "custom"
    assert status["period"] == "260613_260713"
    assert status["period_label"] == "6月13日-7月13日"
    assert status["items"][0]["available_dates"] == ["260613", "260713"]
    assert status["items"][0]["missing_dates"][0] == "260614"
    assert status["items"][0]["missing_dates"][-1] == "260712"
    assert len(status["items"][0]["missing_dates"]) == 29
    assert status["items"][0]["output_folder"] == "骏宜_0613-0713_缺失"


def test_custom_period_requires_both_dates_and_ordered_range(tmp_path):
    service = DailyReportService(tmp_path)

    with pytest.raises(InvalidReportDateError):
        service.get_monthly_summary_status(["junyi"], period_start="260613")

    with pytest.raises(InvalidReportDateError):
        service.get_monthly_summary_status(
            ["junyi"],
            period_start="260713",
            period_end="260613",
        )


def test_guangzhou_new_monthly_status_skips_dates_before_launch(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    service._today = lambda: date(2026, 7, 13)
    monkeypatch.setattr(
        service,
        "_load_accounts",
        lambda: [{"name": "海珠龙星行", "mtd_start_date": "260603"}],
    )

    status = service.get_monthly_summary_status(["guangzhou_new_haizhu"], target_month="2606")
    item = status["items"][0]

    assert item["launch_date"] == "260603"
    assert item["effective_period_start"] == "260603"
    assert item["prelaunch_dates"] == ["260601", "260602"]
    assert item["missing_dates"][0] == "260603"
    assert item["expected_date_count"] == 28
    assert item["is_prelaunch_period"] is False


def test_month_entirely_before_launch_is_skipped_without_backfill(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    service._today = lambda: date(2026, 7, 13)
    monkeypatch.setattr(
        service,
        "_load_accounts",
        lambda: [{"name": "海珠龙星行", "mtd_start_date": "260603"}],
    )

    def unexpected_backfill(*args, **kwargs):
        raise AssertionError("pre-launch dates must not trigger backfill")

    monkeypatch.setattr(service, "_generate_missing_daily_summaries", unexpected_backfill)

    result = service.generate_monthly_summaries(["guangzhou_new_haizhu"], target_month="2605")

    assert result["groups"] == []
    assert result["errors"] == []
    assert result["skipped"] == [
        {
            "group": "guangzhou_new_haizhu",
            "name": "广州龙星行-海珠新车首呼",
            "launch_date": "260603",
            "reason": "广州龙星行-海珠新车首呼于260603上线，所选月份无可统计数据，已跳过",
        }
    ]


def test_generate_monthly_summary_merges_supplemented_days_in_date_order(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    service._today = lambda: date(2026, 7, 13)
    write_summary(tmp_path, "260601")
    write_summary(tmp_path, "260630")

    def generate_missing(group, missing_dates, store_dir, source_date, mtd_start_date=None):
        supplement_root = store_dir / "补数日报"
        return [
            (day, write_summary(tmp_path, day, group["summary"], supplement_root))
            for day in missing_dates
        ]

    monkeypatch.setattr(service, "_generate_missing_daily_summaries", generate_missing)

    result = service.generate_monthly_summaries(["junyi"], target_month="2606")
    group = result["groups"][0]
    output_path = Path(group["summary"]["path"])

    assert group["output_folder"] == "骏宜_六月_缺失"
    assert group["supplemented_dates"][0] == "260602"
    assert group["supplemented_dates"][-1] == "260629"
    assert len(group["source_files"]) == 30
    assert output_path.parent.name == "骏宜_六月_缺失"

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        assert sheet.max_column == 32
        assert sheet.cell(row=1, column=3).value == "骏宜0601"
        assert sheet.cell(row=1, column=32).value == "骏宜0630"
    finally:
        workbook.close()

    image, filename = service.get_monthly_summary_image(
        result["period"],
        "junyi",
        output_folder=group["output_folder"],
    )
    assert image.startswith(bytes([137, 80, 78, 71]))
    assert filename.endswith("_monthly.png")


def test_custom_period_generation_recalculates_every_day_from_period_start(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    for day in ("260630", "260701", "260702"):
        write_summary(tmp_path, day)
    captured = {}

    def recalculate(group, report_dates, store_dir, source_date, mtd_start_date=None):
        captured.update({
            "report_dates": list(report_dates),
            "source_date": source_date,
            "mtd_start_date": mtd_start_date,
        })
        return [
            (day, write_summary(tmp_path, day, group["summary"], store_dir / "补数日报"))
            for day in report_dates
        ]

    monkeypatch.setattr(service, "_generate_missing_daily_summaries", recalculate)

    result = service.generate_monthly_summaries(
        ["junyi"],
        period_start="260630",
        period_end="260702",
    )
    group = result["groups"][0]
    output_path = Path(group["summary"]["path"])

    assert result["selection_mode"] == "custom"
    assert result["period_label"] == "6月30日-7月2日"
    assert group["missing_dates"] == []
    assert group["recalculated_dates"] == ["260630", "260701", "260702"]
    assert captured == {
        "report_dates": ["260630", "260701", "260702"],
        "source_date": "260702",
        "mtd_start_date": "260630",
    }
    assert len(group["source_files"]) == 3
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        assert sheet.cell(row=1, column=3).value == "骏宜0630"
        assert sheet.cell(row=1, column=4).value == "骏宜0701"
        assert sheet.cell(row=1, column=5).value == "骏宜0702"
    finally:
        workbook.close()


def test_custom_period_backfill_uses_period_end_and_only_missing_dates(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    write_summary(tmp_path, "260630")
    write_summary(tmp_path, "260702")
    captured = {}

    def generate_missing(group, report_dates, store_dir, source_date, mtd_start_date=None):
        captured.update({
            "report_dates": list(report_dates),
            "source_date": source_date,
            "mtd_start_date": mtd_start_date,
        })
        return [
            (day, write_summary(tmp_path, day, group["summary"], store_dir / "补数日报"))
            for day in report_dates
        ]

    monkeypatch.setattr(service, "_generate_missing_daily_summaries", generate_missing)

    result = service.generate_monthly_summaries(
        ["junyi"],
        period_start="260630",
        period_end="260702",
    )

    assert captured == {
        "report_dates": ["260630", "260701", "260702"],
        "source_date": "260702",
        "mtd_start_date": "260630",
    }
    group = result["groups"][0]
    assert group["supplemented_dates"] == ["260701"]
    assert group["recalculated_dates"] == ["260630", "260701", "260702"]
    assert group["output_folder"] == "骏宜_0630-0702_缺失"

    image, filename = service.get_monthly_summary_image(
        result["period"],
        "junyi",
        output_folder=group["output_folder"],
    )
    assert image.startswith(bytes([137, 80, 78, 71]))
    assert filename.endswith("_monthly.png")



def test_monthly_backfill_uses_period_end_raw_snapshot_for_every_missing_day(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    account = {"name": "海珠龙星行", "mtd_start_date": "260603"}
    monkeypatch.setattr(service, "_load_accounts", lambda: [account])

    snapshot_dir = tmp_path / "data" / "260630" / "原始数据"
    snapshot_dir.mkdir(parents=True)
    clue_file = snapshot_dir / "海珠龙星行-outcall-线索明细-260630.xlsx"
    call_file = snapshot_dir / "海珠龙星行-aicc-话单-260630.xlsx"
    clue_file.touch()
    call_file.touch()

    captured = []

    def run_from_snapshot(
        account_config,
        report_date,
        output_dir,
        source_files,
        log_path,
        mtd_start_date=None,
    ):
        captured.append((report_date, source_files["clue"], source_files["call"]))
        generated = write_guangzhou_new_summary(
            tmp_path,
            report_date,
            haizhu_value=int(report_date[-2:]),
            panyu_value=int(report_date[-2:]) + 100,
        )
        target = output_dir / "汇总表" / generated.name
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(generated.read_bytes())

    monkeypatch.setattr(service, "_run_daily_processor_from_snapshot", run_from_snapshot)

    group = service._resolve_monthly_group("guangzhou_new_haizhu")
    generated = service._generate_missing_daily_summaries(
        group,
        ["260603", "260604"],
        tmp_path / "monthly",
        "260630",
    )

    assert [day for day, _ in generated] == ["260603", "260604"]
    assert captured == [
        ("260603", clue_file, call_file),
        ("260604", clue_file, call_file),
    ]


def test_cross_month_backfill_uses_each_months_snapshot(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    account = {"name": "广西龙星行"}
    monkeypatch.setattr(service, "_load_accounts", lambda: [account])
    resolved_snapshots = []
    processed_snapshots = []

    def resolve_snapshot(account_config, source_date, store_dir):
        resolved_snapshots.append(source_date)
        return {
            "clue": tmp_path / f"clue-{source_date}.xlsx",
            "call": tmp_path / f"call-{source_date}.xlsx",
        }

    def run_from_snapshot(
        account_config,
        report_date,
        output_dir,
        source_files,
        log_path,
        mtd_start_date=None,
    ):
        processed_snapshots.append((report_date, source_files["clue"].stem))
        write_summary(tmp_path, report_date, "南宁新车首呼", base_dir=output_dir.parent)

    monkeypatch.setattr(service, "_resolve_monthly_snapshot_files", resolve_snapshot)
    monkeypatch.setattr(service, "_run_daily_processor_from_snapshot", run_from_snapshot)
    group = {
        "key": "nanning",
        "title": "南宁",
        "summary": "南宁新车首呼",
        "stores": ["广西龙星行"],
    }

    generated = service._generate_missing_daily_summaries(
        group,
        ["260613", "260704"],
        tmp_path / "monthly",
        "260713",
    )

    assert resolved_snapshots == ["260630", "260713"]
    assert processed_snapshots == [
        ("260613", "clue-260630"),
        ("260704", "clue-260713"),
    ]
    assert [day for day, _ in generated] == ["260613", "260704"]


def test_account_mtd_start_date_limits_custom_period_backfill(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    account = {
        "name": "广西龙星行",
        "mtd_start_date": "260620",
    }
    monkeypatch.setattr(service, "_load_accounts", lambda: [account])
    captured = {}

    def resolve_period_snapshot(account_config, period_start, period_end, store_dir):
        captured["snapshot_period"] = (period_start, period_end)
        return {
            "clue": tmp_path / "clue.xlsx",
            "call": tmp_path / "call.xlsx",
        }

    def run_from_snapshot(
        account_config,
        report_date,
        output_dir,
        source_files,
        log_path,
        mtd_start_date=None,
    ):
        captured["processor_mtd_start_date"] = mtd_start_date
        write_summary(
            tmp_path,
            report_date,
            "南宁新车首呼",
            base_dir=output_dir.parent,
        )

    monkeypatch.setattr(service, "_resolve_period_snapshot_files", resolve_period_snapshot)
    monkeypatch.setattr(service, "_run_daily_processor_from_snapshot", run_from_snapshot)
    group = {
        "key": "nanning",
        "title": "南宁",
        "summary": "南宁新车首呼",
        "stores": ["广西龙星行"],
    }

    generated = service._generate_missing_daily_summaries(
        group,
        ["260621"],
        tmp_path / "monthly",
        "260713",
        mtd_start_date="260613",
    )

    assert [day for day, _ in generated] == ["260621"]
    assert captured == {
        "snapshot_period": ("260620", "260713"),
        "processor_mtd_start_date": "260620",
    }


def test_period_snapshot_combines_monthly_clues_and_calls(tmp_path):
    service = DailyReportService(tmp_path)
    account = {"name": "广西龙星行"}
    june_raw = tmp_path / "data" / "260630" / "原始数据"
    july_raw = tmp_path / "data" / "260713" / "原始数据"
    june_raw.mkdir(parents=True)
    july_raw.mkdir(parents=True)

    pd.DataFrame({"线索ID": ["1", "2"], "线索下发时间": ["2026-06-13", "2026-06-30"]}).to_excel(
        june_raw / "广西龙星行-outcall-线索明细-260630.xlsx",
        index=False,
    )
    pd.DataFrame({"线索ID": ["2", "3"], "线索下发时间": ["2026-06-30", "2026-07-01"]}).to_excel(
        july_raw / "广西龙星行-outcall-线索明细-260713.xlsx",
        index=False,
    )
    pd.DataFrame({"线索ID": ["1", "2"], "结束时间": ["2026-06-13", "2026-06-30"]}).to_excel(
        june_raw / "广西龙星行-aicc-话单-260630.xlsx",
        index=False,
    )
    pd.DataFrame({"线索ID": ["2", "3"], "结束时间": ["2026-07-01", "2026-07-13"]}).to_excel(
        july_raw / "广西龙星行-aicc-话单-260713.xlsx",
        index=False,
    )

    files = service._resolve_period_snapshot_files(
        account,
        "260613",
        "260713",
        tmp_path / "monthly",
    )

    clues = pd.read_excel(files["clue"], dtype={"线索ID": str})
    calls = pd.read_excel(files["call"], dtype={"线索ID": str})
    assert clues["线索ID"].tolist() == ["1", "2", "3"]
    assert calls["线索ID"].tolist() == ["1", "2", "2", "3"]

def test_monthly_generation_uses_source_when_redundant_copy_is_locked(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    write_summary(tmp_path, "260701")

    def locked_copy(*args, **kwargs):
        raise PermissionError("locked")

    monkeypatch.setattr("daily_report.service.shutil.copy2", locked_copy)

    result = service.generate_monthly_summaries(["junyi"], report_date="260701")

    group = result["groups"][0]
    assert Path(group["summary"]["path"]).exists()
    assert group["source_files"] == [str(tmp_path / "data" / "260701" / "汇总表" / "骏宜_日度月度汇总表_260701.xlsx")]
def test_target_month_must_be_one_of_recent_three_months(tmp_path):
    service = DailyReportService(tmp_path)
    service._today = lambda: date(2026, 7, 13)

    with pytest.raises(InvalidReportDateError):
        service.get_monthly_summary_status(["junyi"], target_month="2604")

def test_monthly_api_exposes_months_and_passes_target_month(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    service._today = lambda: date(2026, 7, 13)
    captured = {}

    def generate(**kwargs):
        captured.update(kwargs)
        return {
            "target_month": kwargs["target_month"],
            "period": "260601_260630",
            "groups": [],
            "errors": [],
        }

    monkeypatch.setattr(service, "generate_monthly_summaries", generate)
    app = FastAPI()
    app.include_router(create_router(service))
    client = TestClient(app)

    months_response = client.get("/api/daily-report/monthly-summary/months")
    generate_response = client.post(
        "/api/daily-report/monthly-summary",
        json={"groups": ["junyi"], "target_month": "2606", "force_overwrite": True},
    )

    assert months_response.status_code == 200
    assert [item["value"] for item in months_response.json()["months"]] == ["2607", "2606", "2605"]
    assert generate_response.status_code == 200
    assert generate_response.json()["target_month"] == "2606"
    assert captured["groups"] == ["junyi"]
    assert captured["target_month"] == "2606"
    assert captured["force_overwrite"] is True


def test_monthly_api_passes_custom_period(tmp_path, monkeypatch):
    service = DailyReportService(tmp_path)
    captured = {}

    def generate(**kwargs):
        captured.update(kwargs)
        return {
            "selection_mode": "custom",
            "period": "260613_260713",
            "groups": [],
            "errors": [],
        }

    monkeypatch.setattr(service, "generate_monthly_summaries", generate)
    app = FastAPI()
    app.include_router(create_router(service))
    client = TestClient(app)

    response = client.post(
        "/api/daily-report/monthly-summary",
        json={
            "groups": ["junyi"],
            "period_start": "260613",
            "period_end": "260713",
            "force_overwrite": False,
        },
    )

    assert response.status_code == 200
    assert captured["period_start"] == "260613"
    assert captured["period_end"] == "260713"
    assert captured["target_month"] is None
