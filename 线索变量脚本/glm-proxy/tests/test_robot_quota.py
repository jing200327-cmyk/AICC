from __future__ import annotations

import pprint
import sys
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from robot_quota.api import create_router
from robot_quota.models import RobotQuotaJob
from robot_quota.service import InvalidRobotQuotaDateError, RobotQuotaService
from task_records.service import TaskRecordService


def write_accounts(project_root: Path, accounts: list[dict]) -> None:
    tools_dir = project_root / 'tools'
    tools_dir.mkdir(parents=True, exist_ok=True)
    (tools_dir / 'recorder.py').write_text(
        'ACCOUNTS = ' + pprint.pformat(accounts, width=120, sort_dicts=False) + '\n',
        encoding='utf-8',
    )


def write_call_file(path: Path, rows: list[list[object]], include_robot_id: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    headers = ['机器人', '结束时间', '通话状态', '任务名称']
    if include_robot_id:
        headers.append('机器人ID')
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def make_service(tmp_path: Path, accounts: list[dict], **kwargs) -> RobotQuotaService:
    daily_root = tmp_path / 'daily'
    write_accounts(daily_root, accounts)
    return RobotQuotaService(
        daily_project_root=daily_root,
        output_root=tmp_path / 'quota',
        **kwargs,
    )


def test_usage_filters_selected_day_and_line_limit_calls(tmp_path):
    account = {
        'name': '广州售后',
        'username': 'user',
        'password': 'secret',
        'group_by_call_field': '机器人',
        'required_group_values': ['番禺售后', '活动招揽'],
        'group_display_names': {
            '番禺售后': '广州番禺售后',
            '活动招揽': '广州活动招揽',
        },
    }
    service = make_service(tmp_path, [account])
    call_path = service.daily_project_root / 'data' / '260710' / '原始数据' / '广州售后-aicc-话单-260710.xlsx'
    rows = []
    rows.extend([['番禺售后', '2026-07-10 09:00:00', '已接通', '番禺任务', '10038']] * 401)
    rows.extend([['番禺售后', '2026-07-10 10:00:00', '未接通-线路限制-403', '番禺任务', '10038']] * 3)
    rows.extend([['活动招揽', '2026-07-10 11:00:00', '未接通-拒接', '活动任务', '10039']] * 201)
    rows.append(['活动招揽', '2026-07-09 11:00:00', '已接通', '活动任务', '10039'])
    write_call_file(call_path, rows)

    result = service.generate('260710', prepare_raw=False)

    by_name = {item['robot_name']: item for item in result['robots']}
    assert by_name['广州番禺售后']['call_count'] == 401
    assert by_name['广州番禺售后']['excluded_line_limit_count'] == 3
    assert by_name['广州番禺售后']['quota'] == 400
    assert by_name['广州番禺售后']['usage_rate'] == pytest.approx(1.0025)
    assert by_name['广州番禺售后']['is_over_quota'] is True
    assert by_name['广州活动招揽']['call_count'] == 201
    assert by_name['广州活动招揽']['quota'] == 200
    assert result['over_quota_count'] == 2


def test_generation_respects_required_robot_groups_and_account_id_mapping(tmp_path):
    account = {
        'name': '韶关',
        'username': 'user',
        'password': 'secret',
        'group_by_call_field': '机器人',
        'required_group_values': ['目标机器人'],
        'group_display_names': {'目标机器人': '韶关'},
        'robot_ids': {'目标机器人': '88001'},
    }
    service = make_service(tmp_path, [account])
    call_path = service.daily_project_root / 'data' / '260710' / '原始数据' / '韶关-aicc-话单-260710.xlsx'
    write_call_file(
        call_path,
        [
            ['目标机器人', '2026-07-10 09:00:00', '已接通', '目标任务'],
            ['其它机器人', '2026-07-10 09:10:00', '已接通', '其它任务'],
        ],
        include_robot_id=False,
    )

    result = service.generate('260710', prepare_raw=False)

    assert [(item['robot_id'], item['robot_name'], item['call_count']) for item in result['robots']] == [
        ('88001', '韶关', 1),
    ]
    assert result['robots'][0]['robot_id_source'] == 'account_config'


def test_single_robot_account_ignores_unrelated_robot_when_name_matches(tmp_path):
    account = {'name': '骏宜', 'username': 'user', 'password': 'secret'}
    service = make_service(tmp_path, [account])
    call_path = service.daily_project_root / 'data' / '260710' / '原始数据' / '骏宜-aicc-话单-260710.xlsx'
    write_call_file(
        call_path,
        [
            ['龙星行-广州龙星骏宜', '2026-07-10 09:00:00', '已接通', '骏宜任务', '10001'],
            ['龙星行-新车首呼-广东韶关', '2026-07-10 09:10:00', '已接通', '韶关任务', '10002'],
        ],
    )

    result = service.generate('260710', prepare_raw=False)

    assert [(item['robot_name'], item['call_count']) for item in result['robots']] == [
        ('龙星行-广州龙星骏宜', 1),
    ]


def test_daily_report_completeness_reuses_raw_files_without_runner(tmp_path):
    account = {'name': '长沙', 'username': 'user', 'password': 'secret'}
    calls = []

    def daily_runner(*_args, **_kwargs):
        calls.append('called')

    service = make_service(tmp_path, [account], daily_runner=daily_runner)
    report_dir = service.daily_project_root / 'data' / '260710' / '每日报告'
    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / '长沙_每日报告_260710.txt').write_text('done', encoding='utf-8')
    call_path = service.daily_project_root / 'data' / '260710' / '原始数据' / '长沙-aicc-话单-260710.xlsx'
    write_call_file(call_path, [['长沙机器人', '2026-07-10 09:00:00', '已接通', '任务', '10001']])

    result = service.generate('260710')

    assert calls == []
    assert result['source_mode'] == 'daily_report_reuse'


def test_missing_daily_report_runs_daily_chain_before_processing(tmp_path):
    account = {'name': '长沙', 'username': 'user', 'password': 'secret'}
    calls = []

    def daily_runner(report_date: str, store: str) -> None:
        calls.append((report_date, store))
        call_path = service.daily_project_root / 'data' / report_date / '原始数据' / f'{store}-aicc-话单-{report_date}.xlsx'
        write_call_file(call_path, [['长沙机器人', '2026-07-10 09:00:00', '已接通', '任务', '10001']])

    service = make_service(tmp_path, [account], daily_runner=daily_runner)

    result = service.generate('260710')

    assert calls == [('260710', '长沙')]
    assert result['source_mode'] == 'daily_chain'


def test_missing_raw_file_runs_daily_chain_even_when_report_exists(tmp_path):
    account = {'name': '长沙', 'username': 'user', 'password': 'secret'}
    calls = []

    def daily_runner(report_date: str, store: str) -> None:
        calls.append((report_date, store))
        call_path = service.daily_project_root / 'data' / report_date / '原始数据' / f'{store}-aicc-话单-{report_date}.xlsx'
        write_call_file(call_path, [['长沙机器人', '2026-07-10 09:00:00', '已接通', '任务', '10001']])

    service = make_service(tmp_path, [account], daily_runner=daily_runner)
    report_dir = service.daily_project_root / 'data' / '260710' / '每日报告'
    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / '长沙_每日报告_260710.txt').write_text('done', encoding='utf-8')

    result = service.generate('260710')

    assert calls == [('260710', '长沙')]
    assert result['source_mode'] == 'daily_chain'


def test_platform_robot_id_resolver_is_called_once_per_account(tmp_path):
    account = {
        'name': '双机器人门店',
        'username': 'user',
        'password': 'secret',
        'required_group_values': ['机器人甲', '机器人乙'],
    }
    resolver_calls = []

    def resolver(*args):
        resolver_calls.append(args)
        return {}

    service = make_service(tmp_path, [account], robot_id_resolver=resolver)
    call_path = service.daily_project_root / 'data' / '260710' / '原始数据' / '双机器人门店-aicc-话单-260710.xlsx'
    write_call_file(
        call_path,
        [
            ['机器人甲', '2026-07-10 09:00:00', '已接通', '任务甲'],
            ['机器人乙', '2026-07-10 09:10:00', '已接通', '任务乙'],
        ],
        include_robot_id=False,
    )

    result = service.generate('260710', prepare_raw=False)

    assert len(resolver_calls) == 1
    assert all(item['robot_id'] == '未识别' for item in result['robots'])


def test_workbook_contains_skill_sheet_and_full_detail_sheet(tmp_path):
    account = {'name': '长沙', 'username': 'user', 'password': 'secret'}
    service = make_service(tmp_path, [account])
    call_path = service.daily_project_root / 'data' / '260710' / '原始数据' / '长沙-aicc-话单-260710.xlsx'
    write_call_file(
        call_path,
        [['长沙机器人', '2026-07-10 09:00:00', '已接通', '任务', '10001']] * 205,
    )

    result = service.generate('260710', prepare_raw=False)
    workbook = load_workbook(result['workbook_path'], data_only=False)
    try:
        skill_sheet = workbook['超量记录']
        detail_sheet = workbook['用量明细']
        assert skill_sheet['A1'].value == '7月10号'
        assert skill_sheet['A2'].value == '• [10001]长沙机器人(最高102.50%)'
        assert [cell.value for cell in detail_sheet[1]][:7] == [
            '机器人ID', '机器人名称', '所属门店', '有效通话通次A', '每日限额', '用量B', '状态',
        ]
        assert detail_sheet['F2'].value == pytest.approx(1.025)
        assert detail_sheet['F2'].number_format == '0.00%'
    finally:
        workbook.close()


def test_daily_and_weekly_images_use_skill_compatible_workbooks(tmp_path):
    account = {'name': '长沙', 'username': 'user', 'password': 'secret'}
    service = make_service(tmp_path, [account])
    for report_date, day_text, count in (
        ('260710', '2026-07-10 09:00:00', 205),
        ('260711', '2026-07-11 09:00:00', 210),
    ):
        call_path = service.daily_project_root / 'data' / report_date / '原始数据' / f'长沙-aicc-话单-{report_date}.xlsx'
        write_call_file(call_path, [['长沙机器人', day_text, '已接通', '任务', '10001']] * count)
        service.generate(report_date, prepare_raw=False)

    daily = service.generate_daily_report_image('260710')
    weekly = service.generate_weekly_report_image('260710', '260711')

    assert Path(daily['image_path']).read_bytes().startswith(b'\x89PNG')
    assert Path(weekly['image_path']).read_bytes().startswith(b'\x89PNG')
    assert Path(weekly['workbook_path']).exists()
    assert weekly['days'] == ['7月10号', '7月11号']


def test_api_returns_structured_invalid_date_error(tmp_path):
    service = make_service(tmp_path, [])
    app = FastAPI()
    app.include_router(create_router(service))
    client = TestClient(app)

    response = client.get('/api/robot-quota/preview', params={'report_date': '2026-07-10'})

    assert response.status_code == 422
    assert response.json()['error']['code'] == InvalidRobotQuotaDateError.code


def test_robot_quota_jobs_are_included_in_task_records(tmp_path):
    quota_service = make_service(
        tmp_path,
        [{'name': '长沙', 'username': 'user', 'password': 'secret'}],
    )
    output = quota_service.output_root / 'data' / '260710' / '外呼机器人用量_260710.xlsx'
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(b'xlsx')
    quota_service.jobs['robot_quota_260710_test'] = RobotQuotaJob(
        job_id='robot_quota_260710_test',
        report_date='260710',
        status='completed',
        created_at='2026-07-10T10:00:00',
        updated_at='2026-07-10T10:01:00',
        completed_at='2026-07-10T10:01:00',
        workbook_path=str(output),
    )
    empty = SimpleNamespace(jobs={})
    record_service = TaskRecordService(empty, empty, empty, empty, quota_service)

    result = record_service.list_records(task_type='robot_quota', task_date='2026-07-10')

    assert result['total'] == 1
    assert result['records'][0]['task_name'] == '20260710机器人用量监控'
    assert result['records'][0]['output_files'] == ['外呼机器人用量_260710.xlsx']
