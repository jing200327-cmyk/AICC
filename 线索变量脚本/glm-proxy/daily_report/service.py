from __future__ import annotations

import asyncio
import calendar
import html
import importlib
import json
import os
import re
import shutil
import subprocess
import stat
import sys
import time as time_module
import uuid
from copy import copy
from dataclasses import asdict
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from .models import DailyReportFile, DailyReportJob, DailyReportPreview


REPORT_PREVIEW_GROUPS = [
    {
        'key': 'junyi',
        'title': '\u9a8f\u5b9c',
        'reports': ['\u9a8f\u5b9c'],
        'summary': '\u9a8f\u5b9c',
        'stores': ['\u9a8f\u5b9c'],
    },
    {
        'key': 'changsha',
        'title': '\u957f\u6c99',
        'reports': ['\u957f\u6c99'],
        'summary': '\u957f\u6c99',
        'stores': ['\u957f\u6c99'],
    },
    {
        'key': 'shaoguan',
        'title': '\u97f6\u5173',
        'reports': ['\u97f6\u5173'],
        'summary': '\u97f6\u5173',
        'stores': ['\u97f6\u5173'],
    },
    {
        'key': 'xiangpeng',
        'title': '\u7fd4\u9e4f',
        'reports': ['\u7fd4\u9e4f'],
        'summary': '\u7fd4\u9e4f',
        'stores': ['\u7fd4\u9e4f'],
    },
    {
        'key': 'yulin',
        'title': '\u7389\u6797',
        'reports': ['\u5e7f\u897f\u9f99\u661f\u884c-\u7389\u6797\u65b0\u8f66\u9996\u547c'],
        'summary': '\u7389\u6797\u65b0\u8f66\u9996\u547c',
        'stores': ['\u5e7f\u897f\u9f99\u661f\u884c'],
    },
    {
        'key': 'nanning',
        'title': '\u5357\u5b81',
        'reports': ['\u5e7f\u897f\u9f99\u661f\u884c-\u5357\u5b81\u65b0\u8f66\u9996\u547c'],
        'summary': '\u5357\u5b81\u65b0\u8f66\u9996\u547c',
        'stores': ['\u5e7f\u897f\u9f99\u661f\u884c'],
    },
    {
        'key': 'guangzhou_new',
        'title': '\u5e7f\u5dde\u65b0\u8f66',
        'reports': [
            '\u5e7f\u5dde\u9f99\u661f\u884c-\u6d77\u73e0\u65b0\u8f66\u9996\u547c',
            '\u5e7f\u5dde\u9f99\u661f\u884c-\u756a\u79ba\u65b0\u8f66\u9996\u547c',
        ],
        'summary': '\u5e7f\u5dde\u65b0\u8f66',
        'stores': ['\u6d77\u73e0\u9f99\u661f\u884c'],
    },
    {
        'key': 'guangzhou_after_sales',
        'title': '\u5e7f\u5dde\u552e\u540e',
        'reports': [
            '\u5e7f\u5dde\u9f99\u661f\u884c-\u552e\u540e-\u6d3b\u52a8\u62db\u63fd',
            '\u5e7f\u5dde\u9f99\u661f\u884c-\u552e\u540e-\u7eed\u4fdd\u63d0\u9192',
        ],
        'summary': '\u5e7f\u5dde\u552e\u540e',
        'stores': ['\u5e7f\u5dde\u9f99\u661f\u884c'],
    },
]


MONTHLY_SUMMARY_GROUPS = []
for _preview_group in REPORT_PREVIEW_GROUPS:
    if _preview_group['key'] != 'guangzhou_new':
        MONTHLY_SUMMARY_GROUPS.append(_preview_group)
        continue
    MONTHLY_SUMMARY_GROUPS.extend([
        {
            'key': 'guangzhou_new_haizhu',
            'title': '广州龙星行-海珠新车首呼',
            'summary': '广州新车',
            'source_column': '广州龙星行-海珠新车首呼',
            'stores': ['海珠龙星行'],
        },
        {
            'key': 'guangzhou_new_panyu',
            'title': '广州龙星行-番禺新车首呼',
            'summary': '广州新车',
            'source_column': '广州龙星行-番禺新车首呼',
            'stores': ['海珠龙星行'],
        },
    ])


class DailyReportError(Exception):
    code = 'DAILY_REPORT_ERROR'
    message = 'Daily report failed'

    def __init__(self, detail: str = ''):
        super().__init__(detail or self.message)
        self.detail = detail


class InvalidReportDateError(DailyReportError):
    code = 'INVALID_REPORT_DATE'
    message = 'REPORT_DATE must be YYMMDD'


class InvalidReportStoreError(DailyReportError):
    code = 'INVALID_REPORT_STORE'
    message = 'Report store is not configured'


class InvalidReportGroupError(DailyReportError):
    code = 'INVALID_REPORT_GROUP'
    message = 'Report preview group is not configured'


class AllStoreSummaryExistsError(DailyReportError):
    code = 'ALL_STORE_SUMMARY_EXISTS'
    message = 'All-store summary already exists'


class DailyReportService:
    def __init__(self, project_root: Path):
        self.project_root = Path(project_root)
        self.jobs: dict[str, DailyReportJob] = {}
        self.tasks: dict[str, asyncio.Task] = {}

    def list_stores(self) -> list[dict[str, str]]:
        stores = [{'code': 'all', 'name': '所有'}]
        for account in self._load_accounts():
            name = str(account.get('name') or '').strip()
            if name:
                stores.append({'code': name, 'name': name})
        return stores

    def start_job(self, report_date: str | None = None, store: str = 'all', refresh_clue: bool = False) -> DailyReportJob:
        report_date = self._validate_report_date(report_date)
        store = (store or 'all').strip()
        store_name = self._validate_store(store)
        command = [sys.executable, str(self.project_root / 'main.py')]
        now = datetime.now().isoformat()
        job = DailyReportJob(
            job_id=f'daily_report_{report_date}_{uuid.uuid4().hex[:10]}',
            status='running',
            report_date=report_date,
            store=store,
            store_name=store_name,
            command=command,
            cwd=str(self.project_root),
            created_at=now,
            updated_at=now,
            refresh_clue=bool(refresh_clue),
        )
        self.jobs[job.job_id] = job
        self.tasks[job.job_id] = asyncio.create_task(self._run_job(job.job_id))
        return job

    def get_job(self, job_id: str) -> DailyReportJob | None:
        return self.jobs.get(job_id)

    def get_preview(self, report_date: str | None = None, store: str = 'all') -> dict[str, Any]:
        report_date = self._validate_preview_date(report_date)
        store = (store or 'all').strip()
        if not self._is_all_store(store):
            self._validate_store(store)

        reports = {item.name: asdict(item) for item in self._collect_reports(report_date, 'all', '')}
        summaries = self._collect_summaries(report_date)
        groups = []
        for group in REPORT_PREVIEW_GROUPS:
            if not self._preview_group_matches_store(group, store):
                continue
            report_items = []
            for report_name in group['reports']:
                report_items.append(reports.get(report_name) or {
                    'name': report_name,
                    'filename': '',
                    'path': '',
                    'content': '',
                    'missing': True,
                })
            summary = summaries.get(group['summary']) or {
                'name': group['summary'],
                'filename': '',
                'path': '',
                'columns': [],
                'rows': [],
                'content': '',
                'missing': True,
            }
            groups.append({
                'key': group['key'],
                'title': group['title'],
                'reports': report_items,
                'summary': summary,
            })

        return {
            'report_date': report_date,
            'groups': groups,
        }

    def get_summary_image(self, report_date: str | None = None, group: str = '') -> tuple[bytes, str]:
        report_date = self._validate_report_date(report_date)
        group_config = self._resolve_preview_group(group)
        summary_path = self._summary_path(report_date, group_config['summary'])
        if not summary_path.exists():
            raise DailyReportError(f'Summary file does not exist: {summary_path.name}')
        image_bytes = self._summary_image(summary_path)
        filename = f"{group_config['key']}_{report_date}.png"
        return image_bytes, filename

    def list_monthly_summary_stores(self) -> list[dict[str, str]]:
        return [
            {
                'code': group['key'],
                'name': group['title'],
                'summary': group['summary'],
            }
            for group in MONTHLY_SUMMARY_GROUPS
        ]

    def list_monthly_summary_months(self) -> list[dict[str, str]]:
        today = self._today()
        options = []
        for offset, suffix in enumerate(('当月', '前一个月', '前两个月')):
            month_index = today.year * 12 + today.month - 1 - offset
            year, zero_based_month = divmod(month_index, 12)
            month = zero_based_month + 1
            start = date(year, month, 1)
            if offset == 0:
                end = today
            else:
                end = date(year, month, calendar.monthrange(year, month)[1])
            options.append({
                'value': start.strftime('%y%m'),
                'label': f'{month}月（{suffix}）',
                'period_start': start.strftime('%y%m%d'),
                'period_end': end.strftime('%y%m%d'),
            })
        return options

    def get_monthly_summary_status(
        self,
        groups: list[str],
        report_date: str | None = None,
        target_month: str | None = None,
    ) -> dict[str, Any]:
        period_info = self._resolve_monthly_period(target_month, report_date)
        selected_groups = self._resolve_monthly_groups(groups)
        period_start = period_info['period_start']
        period_end = period_info['period_end']
        period = period_info['period']
        period_dir = self.project_root / 'data' / period
        items = []
        for group in selected_groups:
            display_name = group['title']
            schedule = self._monthly_group_schedule(group, period_start, period_end)
            expected_dates = schedule['expected_dates']
            expected_date_set = set(expected_dates)
            source_files = [
                (day, path)
                for day, path in self._monthly_summary_source_files(period_end, group['summary'])
                if day in expected_date_set
            ]
            available_dates = [day for day, _ in source_files]
            missing_dates = [day for day in expected_dates if day not in set(available_dates)]
            output_folder = self._monthly_output_folder(display_name, period_info['month'], bool(missing_dates))
            store_dir = period_dir / output_folder
            output_path = store_dir / f'{display_name}_月度横向对比表.xlsx'
            existing_dirs = [
                path for path in self._monthly_store_dir_candidates(period_dir, display_name, period_info['month'])
                if path.exists()
            ]
            items.append({
                'key': group['key'],
                'name': display_name,
                'summary_name': group['summary'],
                'period': period,
                'period_dir_exists': period_dir.exists(),
                'store_dir_exists': bool(existing_dirs),
                'existing_dirs': [str(path) for path in existing_dirs],
                'output_folder': output_folder,
                'output_exists': output_path.exists(),
                'store_dir': str(store_dir),
                'output_path': str(output_path) if output_path.exists() else '',
                'available_dates': available_dates,
                'missing_dates': missing_dates,
                'expected_date_count': len(expected_dates),
                'expected_dates': expected_dates,
                'launch_date': schedule['launch_date'],
                'effective_period_start': schedule['effective_period_start'],
                'prelaunch_dates': schedule['prelaunch_dates'],
                'is_prelaunch_period': schedule['is_prelaunch_period'],
            })
        return {
            'report_date': period_end,
            'target_month': period_info['month'],
            'month_label': period_info['month_label'],
            'period_start': period_start,
            'period_end': period_end,
            'period': period,
            'period_dir': str(period_dir),
            'has_existing': any(item['store_dir_exists'] for item in items),
            'items': items,
        }

    def generate_monthly_summaries(
        self,
        groups: list[str],
        report_date: str | None = None,
        force_overwrite: bool = False,
        target_month: str | None = None,
    ) -> dict[str, Any]:
        status = self.get_monthly_summary_status(groups, report_date=report_date, target_month=target_month)
        selected_groups = self._resolve_monthly_groups(groups)
        status_by_key = {item['key']: item for item in status['items']}
        period_dir = Path(status['period_dir'])
        generated = []
        errors = []
        skipped = []
        supplement_cache: dict[tuple[str, tuple[str, ...], str], list[tuple[str, Path]]] = {}

        existing_items = [item for item in status['items'] if item['store_dir_exists']]
        if existing_items and not force_overwrite:
            names = '、'.join(item['name'] for item in existing_items)
            raise DailyReportError(f'{names} 已存在月度横向汇总目录，请确认是否再次生成')
        if force_overwrite:
            for item in existing_items:
                for existing_dir in item['existing_dirs']:
                    self._clear_monthly_store_dir(Path(existing_dir))

        for group in selected_groups:
            item = status_by_key[group['key']]
            summary_name = group['summary']
            display_name = group['title']
            if item['is_prelaunch_period']:
                skipped.append({
                    'group': group['key'],
                    'name': display_name,
                    'launch_date': item['launch_date'],
                    'reason': f"{display_name}于{item['launch_date']}上线，所选月份无可统计数据，已跳过",
                })
                continue

            store_dir = Path(item['store_dir'])
            store_dir.mkdir(parents=True, exist_ok=True)

            expected_date_set = set(item['expected_dates'])
            standard_files = [
                (day, path)
                for day, path in self._monthly_summary_source_files(status['period_end'], summary_name)
                if day in expected_date_set
            ]
            supplemented_files: list[tuple[str, Path]] = []
            try:
                if item['missing_dates']:
                    cache_key = (summary_name, tuple(item['missing_dates']), status['period_end'])
                    if cache_key not in supplement_cache:
                        supplement_cache[cache_key] = self._generate_missing_daily_summaries(
                            group,
                            item['missing_dates'],
                            store_dir,
                            status['period_end'],
                        )
                    supplemented_files = supplement_cache[cache_key]
            except DailyReportError as exc:
                errors.append({
                    'group': group['key'],
                    'name': display_name,
                    'message': exc.detail or str(exc),
                })
                continue

            files_by_date = {day: path for day, path in standard_files}
            files_by_date.update({day: path for day, path in supplemented_files})
            missing_after_supplement = [
                day for day in item['expected_dates']
                if day not in files_by_date
            ]
            if missing_after_supplement:
                errors.append({
                    'group': group['key'],
                    'name': display_name,
                    'message': f"{display_name} 补数后仍缺少日期：{','.join(missing_after_supplement)}",
                })
                continue

            source_files = sorted(files_by_date.items(), key=lambda item: item[0])
            copied_files = []
            for _, source_path in source_files:
                target_path = store_dir / source_path.name
                copied_path = target_path
                try:
                    if source_path.resolve() != target_path.resolve():
                        shutil.copy2(source_path, target_path)
                except PermissionError:
                    # Root-level copies are only for convenient access; calculation uses source_files directly.
                    copied_path = source_path
                copied_files.append(str(copied_path))

            output_path = store_dir / f'{display_name}_月度横向对比表.xlsx'
            try:
                self._build_monthly_horizontal_summary(
                    source_files,
                    output_path,
                    display_name,
                    status['period_end'],
                    source_column=group.get('source_column'),
                )
            except PermissionError as exc:
                raise DailyReportError(f'{output_path.name} 文件正在被占用，请关闭已打开的 Excel 文件后重试') from exc
            preview = self._summary_preview(output_path, display_name)
            generated.append({
                'key': group['key'],
                'title': f"{display_name}{status['period_start'][-4:]}-{status['period_end'][-4:]}汇总",
                'name': display_name,
                'summary_name': summary_name,
                'period': status['period'],
                'target_month': status['target_month'],
                'month_label': status['month_label'],
                'output_folder': item['output_folder'],
                'output_dir': str(store_dir),
                'source_files': copied_files,
                'missing_dates': item['missing_dates'],
                'supplemented_dates': [day for day, _ in supplemented_files],
                'summary': preview,
            })

        if not generated and errors:
            detail = '; '.join(item['message'] for item in errors)
            raise DailyReportError(detail)

        return {
            'report_date': status['period_end'],
            'target_month': status['target_month'],
            'month_label': status['month_label'],
            'period_start': status['period_start'],
            'period_end': status['period_end'],
            'period': status['period'],
            'output_dir': str(period_dir),
            'groups': generated,
            'errors': errors,
            'skipped': skipped,
        }

    def get_monthly_summary_image(
        self,
        period: str,
        group: str,
        output_folder: str = '',
    ) -> tuple[bytes, str]:
        if not re.fullmatch(r'\d{6}_\d{6}', str(period or '').strip()):
            raise InvalidReportDateError(period)
        if str(group or '').strip() in {'guangzhou_new', '广州新车'}:
            group_config = self._resolve_preview_group(group)
        else:
            group_config = self._resolve_monthly_group(str(group or '').strip())
        display_name = group_config['title']
        month = period[:4]
        allowed_folders = {
            display_name,
            self._monthly_output_folder(display_name, month, True),
        }
        folder = str(output_folder or '').strip() or display_name
        if folder not in allowed_folders:
            raise InvalidReportGroupError(folder)
        summary_path = self.project_root / 'data' / period / folder / f'{display_name}_月度横向对比表.xlsx'
        if not summary_path.exists() and not output_folder:
            fallback_folder = self._monthly_output_folder(display_name, month, True)
            fallback = self.project_root / 'data' / period / fallback_folder / f'{display_name}_月度横向对比表.xlsx'
            if fallback.exists():
                summary_path = fallback
        if not summary_path.exists():
            raise DailyReportError(f'Monthly summary file does not exist: {summary_path.name}')
        image_bytes = self._summary_image(summary_path)
        return image_bytes, f'{display_name}_{period}_monthly.png'

    def _today(self) -> date:
        return date.today()

    def _resolve_monthly_period(
        self,
        target_month: str | None,
        report_date: str | None,
    ) -> dict[str, str]:
        if target_month:
            value = str(target_month).strip()
            options = {item['value']: item for item in self.list_monthly_summary_months()}
            if value not in options:
                raise InvalidReportDateError(value)
            selected = options[value]
            return {
                'month': selected['value'],
                'month_label': f"{int(selected['value'][-2:])}月",
                'period_start': selected['period_start'],
                'period_end': selected['period_end'],
                'period': f"{selected['period_start']}_{selected['period_end']}",
            }

        end = self._validate_report_date(report_date)
        start = f'{end[:4]}01'
        return {
            'month': end[:4],
            'month_label': f'{int(end[2:4])}月',
            'period_start': start,
            'period_end': end,
            'period': f'{start}_{end}',
        }

    def _date_range(self, start: str, end: str) -> list[str]:
        start_date = datetime.strptime(start, '%y%m%d').date()
        end_date = datetime.strptime(end, '%y%m%d').date()
        result = []
        current = start_date
        while current <= end_date:
            result.append(current.strftime('%y%m%d'))
            current += timedelta(days=1)
        return result

    def _monthly_group_schedule(
        self,
        group: dict[str, Any],
        period_start: str,
        period_end: str,
    ) -> dict[str, Any]:
        try:
            configured_accounts = self._load_accounts()
        except (ImportError, ModuleNotFoundError):
            configured_accounts = []
        accounts = {
            str(account.get('name') or '').strip(): account
            for account in configured_accounts
        }
        launch_dates = []
        has_unrestricted_account = False
        for account_name in group.get('stores', []):
            account = accounts.get(account_name)
            if not account:
                continue
            launch_date = self._account_launch_date(account)
            if launch_date:
                launch_dates.append(launch_date)
            else:
                has_unrestricted_account = True

        launch_date = ''
        if launch_dates and not has_unrestricted_account:
            launch_date = min(launch_dates)

        effective_period_start = max(period_start, launch_date) if launch_date else period_start
        expected_dates = (
            self._date_range(effective_period_start, period_end)
            if effective_period_start <= period_end
            else []
        )
        prelaunch_dates = []
        if launch_date and launch_date > period_start:
            prelaunch_end = min(
                datetime.strptime(period_end, '%y%m%d').date(),
                datetime.strptime(launch_date, '%y%m%d').date() - timedelta(days=1),
            )
            if datetime.strptime(period_start, '%y%m%d').date() <= prelaunch_end:
                prelaunch_dates = self._date_range(period_start, prelaunch_end.strftime('%y%m%d'))

        return {
            'launch_date': launch_date,
            'effective_period_start': effective_period_start,
            'expected_dates': expected_dates,
            'prelaunch_dates': prelaunch_dates,
            'is_prelaunch_period': bool(launch_date and launch_date > period_end),
        }

    def _account_launch_date(self, account: dict[str, Any]) -> str:
        value = str(account.get('mtd_start_date') or '').strip()
        if not re.fullmatch(r'[0-9]{6}', value):
            return ''
        try:
            datetime.strptime(value, '%y%m%d')
        except ValueError:
            return ''
        return value

    def _account_is_active_on(self, account: dict[str, Any], report_date: str) -> bool:
        launch_date = self._account_launch_date(account)
        return not launch_date or report_date >= launch_date

    def _monthly_output_folder(self, display_name: str, month: str, has_missing: bool) -> str:
        if not has_missing:
            return display_name
        month_names = {
            1: '一月', 2: '二月', 3: '三月', 4: '四月', 5: '五月', 6: '六月',
            7: '七月', 8: '八月', 9: '九月', 10: '十月', 11: '十一月', 12: '十二月',
        }
        return f"{display_name}_{month_names[int(month[-2:])]}_缺失"

    def _monthly_store_dir_candidates(self, period_dir: Path, display_name: str, month: str) -> list[Path]:
        return [
            period_dir / display_name,
            period_dir / self._monthly_output_folder(display_name, month, True),
        ]

    def _generate_missing_daily_summaries(
        self,
        group: dict[str, Any],
        missing_dates: list[str],
        store_dir: Path,
        source_date: str,
    ) -> list[tuple[str, Path]]:
        accounts = {
            str(account.get('name') or '').strip(): account
            for account in self._load_accounts()
        }
        account_names = [
            name for name in group.get('stores', [])
            if name in accounts
        ]
        if not account_names:
            raise DailyReportError(f"{group['title']} 未找到可用于补跑的账户配置")

        supplement_root = store_dir / '补数日报'
        log_root = store_dir / '补数日志'
        supplement_root.mkdir(parents=True, exist_ok=True)
        log_root.mkdir(parents=True, exist_ok=True)

        snapshot_files = {
            account_name: self._resolve_monthly_snapshot_files(
                accounts[account_name],
                source_date,
                store_dir,
            )
            for account_name in account_names
        }
        generated = []

        for report_date in missing_dates:
            eligible_account_names = [
                name for name in account_names
                if self._account_is_active_on(accounts[name], report_date)
            ]
            if not eligible_account_names:
                continue

            day_dir = supplement_root / report_date
            for account_name in eligible_account_names:
                log_path = log_root / f'{report_date}_{account_name}.log'
                self._run_daily_processor_from_snapshot(
                    accounts[account_name],
                    report_date,
                    day_dir,
                    snapshot_files[account_name],
                    log_path,
                )

            summary_dir = day_dir / '汇总表'
            suffix = self._summary_suffix(report_date)
            direct = summary_dir / f"{group['summary']}{suffix}.xlsx"
            if direct.exists():
                generated.append((report_date, direct))
                continue
            matches = sorted(
                path for path in summary_dir.glob(f"{group['summary']}*{suffix}.xlsx")
                if path.is_file() and not path.name.startswith('~$')
            ) if summary_dir.exists() else []
            if not matches:
                raise DailyReportError(
                    f"{group['title']} {report_date} 补跑完成但未生成 {group['summary']} 汇总表"
                )
            generated.append((report_date, matches[0]))
        return generated

    def _resolve_monthly_snapshot_files(
        self,
        account: dict[str, Any],
        source_date: str,
        store_dir: Path,
    ) -> dict[str, Path]:
        account_name = str(account.get('name') or '').strip()

        def files_in(root: Path) -> dict[str, Path] | None:
            raw_dir = root / '原始数据'
            files = {
                'clue': raw_dir / f'{account_name}-outcall-线索明细-{source_date}.xlsx',
                'call': raw_dir / f'{account_name}-aicc-话单-{source_date}.xlsx',
            }
            return files if all(path.exists() for path in files.values()) else None

        standard_files = files_in(self.project_root / 'data' / source_date)
        if standard_files:
            return standard_files

        snapshot_dir = store_dir / '补数基准' / source_date
        cached_files = files_in(snapshot_dir)
        if cached_files:
            return cached_files

        snapshot_dir.mkdir(parents=True, exist_ok=True)
        env = os.environ.copy()
        env['REPORT_DATE'] = source_date
        env['REPORT_STORE'] = account_name
        env['REPORT_OUTPUT_DIR'] = str(snapshot_dir)
        env['REPORT_REFRESH_CLUE'] = '1'
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            [sys.executable, str(self.project_root / 'main.py')],
            cwd=str(self.project_root),
            env=env,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
        )
        log_path = store_dir / '补数日志' / f'{source_date}_{account_name}_完整快照.log'
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text((result.stdout or '') + (result.stderr or ''), encoding='utf-8')
        if result.returncode != 0:
            raise DailyReportError(
                f'{account_name} {source_date} 完整月度快照下载失败，详见 {log_path}'
            )

        generated_files = files_in(snapshot_dir)
        if not generated_files:
            raise DailyReportError(
                f'{account_name} {source_date} 完整月度快照缺少线索明细或话单，详见 {log_path}'
            )
        return generated_files

    def _run_daily_processor_from_snapshot(
        self,
        account: dict[str, Any],
        report_date: str,
        output_dir: Path,
        source_files: dict[str, Path],
        log_path: Path,
    ) -> None:
        account_name = str(account.get('name') or '').strip()
        command = [
            sys.executable,
            str(self.project_root / 'tools' / 'process_clue_report.py'),
            '--call_file', str(source_files['call']),
            '--xiansuo_file', str(source_files['clue']),
            '--company', account_name,
            '--date', report_date,
            '--output_dir', str(output_dir),
        ]
        scalar_options = {
            'group_by_call_field': '--group_by_call_field',
            'mtd_start_date': '--mtd_start_date',
            'merge_summary_title': '--merge_summary_title',
            'unmatched_group_name': '--unmatched_group_name',
            'clue_match_limit': '--clue_match_limit',
        }
        for key, flag in scalar_options.items():
            value = account.get(key)
            if value not in (None, ''):
                command.extend([flag, str(value)])
        list_options = {
            'exclude_clue_ids': '--exclude_clue_ids',
            'required_group_values': '--required_group_values',
        }
        for key, flag in list_options.items():
            values = account.get(key) or []
            if values:
                command.extend([flag, ','.join(str(value) for value in values)])
        json_options = {
            'group_display_names': '--group_display_names',
            'group_summary_names': '--group_summary_names',
        }
        for key, flag in json_options.items():
            values = account.get(key) or {}
            if values:
                command.extend([flag, json.dumps(values, ensure_ascii=False)])

        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        result = subprocess.run(
            command,
            cwd=str(self.project_root),
            env=env,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
        )
        log_path.parent.mkdir(parents=True, exist_ok=True)
        source_note = (
            f"月度补跑统一快照：{source_files['clue']}\n"
            f"月度补跑统一话单：{source_files['call']}\n\n"
        )
        log_path.write_text(
            source_note + (result.stdout or '') + (result.stderr or ''),
            encoding='utf-8',
        )
        if result.returncode != 0:
            raise DailyReportError(
                f'{account_name} {report_date} 按完整月度快照补跑失败，详见 {log_path}'
            )

    def _clear_monthly_store_dir(self, path: Path) -> None:
        if not path.exists():
            return
        for child in path.iterdir():
            try:
                if child.is_dir():
                    self._remove_tree(child)
                else:
                    child.unlink()
            except PermissionError:
                # Some Windows accounts can overwrite files but cannot delete them; keep going and overwrite known outputs.
                continue
    def _remove_tree(self, path: Path) -> None:
        def handle_remove_error(func, failed_path, exc_info):
            os.chmod(failed_path, stat.S_IWRITE)
            func(failed_path)

        last_error = None
        for attempt in range(4):
            try:
                shutil.rmtree(path, onerror=handle_remove_error)
                return
            except PermissionError as exc:
                last_error = exc
                time_module.sleep(0.35 * (attempt + 1))
        if last_error:
            raise last_error

    def _resolve_monthly_groups(self, groups: list[str]) -> list[dict[str, Any]]:
        resolved = []
        seen = set()
        for raw_value in groups or []:
            value = str(raw_value or '').strip()
            if value in {'guangzhou_new', '广州新车'}:
                matched_groups = [
                    group for group in MONTHLY_SUMMARY_GROUPS
                    if group['key'] in {'guangzhou_new_haizhu', 'guangzhou_new_panyu'}
                ]
            else:
                matched_groups = [self._resolve_monthly_group(value)]
            for group in matched_groups:
                if group['key'] not in seen:
                    resolved.append(group)
                    seen.add(group['key'])
        if not resolved:
            raise InvalidReportStoreError('请选择至少一个门店')
        return resolved

    def _resolve_monthly_group(self, value: str) -> dict[str, Any]:
        matches = [
            group for group in MONTHLY_SUMMARY_GROUPS
            if value in {group['key'], group['title']}
        ]
        if len(matches) == 1:
            return matches[0]
        summary_matches = [
            group for group in MONTHLY_SUMMARY_GROUPS
            if value == group['summary']
        ]
        if len(summary_matches) == 1:
            return summary_matches[0]
        raise InvalidReportGroupError(value)

    def _monthly_summary_source_files(self, report_date: str, summary_name: str) -> list[tuple[str, Path]]:
        data_dir = self.project_root / 'data'
        if not data_dir.exists():
            return []
        period_start = f'{report_date[:4]}01'
        suffix_template = self._summary_suffix
        source_files = []
        for day_dir in sorted(data_dir.iterdir(), key=lambda item: item.name):
            day = day_dir.name
            if not day_dir.is_dir() or not re.fullmatch(r'\d{6}', day):
                continue
            if day[:4] != report_date[:4] or day < period_start or day > report_date:
                continue
            summary_dir = self._data_child_dir(day, '汇总表')
            suffix = suffix_template(day)
            direct = summary_dir / f'{summary_name}{suffix}.xlsx'
            if direct.exists():
                source_files.append((day, direct))
                continue
            if summary_dir.exists():
                matches = sorted(path for path in summary_dir.glob(f'{summary_name}*{suffix}.xlsx') if path.is_file() and not path.name.startswith('~$'))
                if matches:
                    source_files.append((day, matches[0]))
        return source_files

    def _build_monthly_horizontal_summary(
        self,
        source_files: list[tuple[str, Path]],
        output_path: Path,
        display_name: str,
        report_date: str,
        source_column: str | None = None,
    ) -> None:
        template_path = source_files[-1][1]
        workbook = load_workbook(template_path, read_only=False, data_only=True)
        try:
            sheet = workbook.active
            self._expand_merged_summary_rows(sheet)
            template_source_column = self._summary_column_index(sheet, source_column)
            template_source_width = sheet.column_dimensions[get_column_letter(template_source_column)].width or 12
            source_values = [
                self._summary_column_values(path, source_column)
                for _, path in source_files
            ]
            max_rows = max([sheet.max_row or 1] + [len(values) for values in source_values])

            report_date_text = datetime.strptime(report_date, '%y%m%d').strftime('%Y-%m-%d')
            for row_index in range(1, max_rows + 1):
                label_cell = sheet.cell(row=row_index, column=1)
                label_text = str(label_cell.value or '')
                if label_text.startswith('MTD'):
                    label_cell.value = f'MTD (截止：{report_date_text})'
                elif label_text.startswith('Daily Report'):
                    label_cell.value = f'Daily Report ({report_date_text})'

                source_style_cell = sheet.cell(row=row_index, column=template_source_column)
                for col_offset, (day, _) in enumerate(source_files):
                    target_cell = sheet.cell(row=row_index, column=3 + col_offset)
                    self._copy_cell_style(source_style_cell, target_cell)
                    if row_index == 1:
                        target_cell.value = f'{display_name}{day[-4:]}'
                        continue
                    values = source_values[col_offset]
                    if row_index - 1 < len(values):
                        value, number_format = values[row_index - 1]
                        target_cell.value = value
                        if number_format:
                            target_cell.number_format = number_format

            if sheet.max_column > 2 + len(source_files):
                sheet.delete_cols(3 + len(source_files), sheet.max_column - 2 - len(source_files))

            for col_index in range(3, 3 + len(source_files)):
                letter = get_column_letter(col_index)
                sheet.column_dimensions[letter].width = max(12, template_source_width)

            workbook.save(output_path)
        finally:
            workbook.close()

    def _expand_merged_summary_rows(self, sheet) -> None:
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.max_col <= 2:
                continue
            source_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            style_source = self._cell_style_snapshot(source_cell)
            sheet.unmerge_cells(str(merged_range))
            for row_index in range(merged_range.min_row, merged_range.max_row + 1):
                for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                    self._apply_cell_style_snapshot(sheet.cell(row=row_index, column=col_index), style_source)

    def _summary_column_index(self, sheet, source_column: str | None) -> int:
        if not source_column:
            return 3
        expected = str(source_column).strip()
        for column_index in range(3, sheet.max_column + 1):
            actual = str(sheet.cell(row=1, column=column_index).value or '').strip()
            if actual == expected:
                return column_index
        raise DailyReportError(f'汇总表中未找到机器人列：{expected}')

    def _summary_column_values(
        self,
        path: Path,
        source_column: str | None = None,
    ) -> list[tuple[Any, str]]:
        workbook = load_workbook(path, read_only=False, data_only=True)
        try:
            sheet = workbook.active
            column_index = self._summary_column_index(sheet, source_column)
            return [
                (
                    sheet.cell(row=row_index, column=column_index).value,
                    str(sheet.cell(row=row_index, column=column_index).number_format or ''),
                )
                for row_index in range(1, sheet.max_row + 1)
            ]
        finally:
            workbook.close()

    def _copy_cell_style(self, source, target) -> None:
        style_snapshot = self._cell_style_snapshot(source)
        self._apply_cell_style_snapshot(target, style_snapshot)

    def _cell_style_snapshot(self, cell) -> dict[str, Any]:
        return {
            'style': copy(cell._style),
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'protection': copy(cell.protection),
            'number_format': cell.number_format,
        }

    def _apply_cell_style_snapshot(self, cell, snapshot: dict[str, Any]) -> None:
        cell._style = copy(snapshot['style'])
        cell.font = copy(snapshot['font'])
        cell.fill = copy(snapshot['fill'])
        cell.border = copy(snapshot['border'])
        cell.alignment = copy(snapshot['alignment'])
        cell.protection = copy(snapshot['protection'])
        cell.number_format = snapshot['number_format']

    def get_all_store_summary_status(self, report_date: str | None = None) -> dict[str, Any]:
        report_date = self._validate_report_date(report_date)
        summary_dir = self._data_child_dir(report_date, '汇总表')
        output_path = summary_dir / f'{report_date[-4:]}全门店机器人汇总表.xlsx'
        return {
            'report_date': report_date,
            'filename': output_path.name,
            'exists': output_path.exists(),
            'output_dir': str(summary_dir),
            'output_path': str(output_path) if output_path.exists() else '',
        }

    def generate_all_store_summary(self, report_date: str | None = None, force_overwrite: bool = False) -> dict[str, Any]:
        report_date = self._validate_report_date(report_date)
        summary_dir = self._data_child_dir(report_date, '汇总表')
        source_files = self._all_store_summary_source_files(report_date)
        if not source_files:
            raise DailyReportError(f'{report_date} 未找到可汇总的门店汇总表')
        output_path = summary_dir / f'{report_date[-4:]}全门店机器人汇总表.xlsx'
        if output_path.exists() and not force_overwrite:
            raise AllStoreSummaryExistsError(f'已有{output_path.name}，请确认是否重新生成')
        try:
            self._build_all_store_summary(source_files, output_path, report_date)
        except PermissionError as exc:
            raise DailyReportError(f'{output_path.name} 文件正在被占用，请关闭已打开的 Excel 文件后重试') from exc
        preview = self._summary_preview(output_path, f'{report_date[-4:]}全门店汇总')
        return {
            'report_date': report_date,
            'title': f'{report_date[-4:]}全门店汇总',
            'output_dir': str(summary_dir),
            'source_files': [str(path) for path in source_files],
            'summary': preview,
        }

    def get_all_store_summary_image(self, report_date: str | None = None) -> tuple[bytes, str]:
        report_date = self._validate_report_date(report_date)
        summary_dir = self._data_child_dir(report_date, '汇总表')
        summary_path = summary_dir / f'{report_date[-4:]}全门店机器人汇总表.xlsx'
        if not summary_path.exists():
            raise DailyReportError(f'All-store summary file does not exist: {summary_path.name}')
        return self._summary_image(summary_path), f'all_store_{report_date}.png'

    def _all_store_summary_source_files(self, report_date: str) -> list[Path]:
        summary_dir = self._data_child_dir(report_date, '汇总表')
        if not summary_dir.exists():
            return []
        suffix = self._summary_suffix(report_date)
        files = []
        for path in sorted(summary_dir.glob(f'*{suffix}.xlsx')):
            if path.name.startswith('~$') or '全门店机器人汇总表' in path.name:
                continue
            files.append(path)
        return files

    def _build_all_store_summary(self, source_files: list[Path], output_path: Path, report_date: str) -> None:
        template_path = max(source_files, key=lambda path: self._worksheet_dimensions(path)[0])
        workbook = load_workbook(template_path, read_only=False, data_only=True)
        source_workbooks = []
        try:
            sheet = workbook.active
            template_styles = {
                row_index: self._cell_style_snapshot(sheet.cell(row=row_index, column=3))
                for row_index in range(1, sheet.max_row + 1)
            }
            template_width = sheet.column_dimensions['C'].width or 12
            if sheet.max_column > 2:
                sheet.delete_cols(3, sheet.max_column - 2)
            report_date_text = datetime.strptime(report_date, '%y%m%d').strftime('%Y-%m-%d')
            for row_index in range(1, sheet.max_row + 1):
                label_cell = sheet.cell(row=row_index, column=1)
                label_text = str(label_cell.value or '')
                if label_text.startswith('MTD'):
                    label_cell.value = f'MTD (截止：{report_date_text})'
                elif label_text.startswith('Daily Report'):
                    label_cell.value = f'Daily Report ({report_date_text})'

            target_row_keys = self._summary_row_keys(sheet)
            target_col = 3
            for source_path in source_files:
                source_wb = load_workbook(source_path, read_only=False, data_only=True)
                source_workbooks.append(source_wb)
                source_sheet = source_wb.active
                source_row_keys = self._summary_row_keys(source_sheet)
                source_rows_by_key = {key: row_index for row_index, key in source_row_keys.items()}
                for source_col in range(3, source_sheet.max_column + 1):
                    header = source_sheet.cell(row=1, column=source_col).value
                    if header in (None, '') and not any(source_sheet.cell(row=row_index, column=source_col).value not in (None, '') for row_index in range(2, source_sheet.max_row + 1)):
                        continue
                    for target_row in range(1, sheet.max_row + 1):
                        target_cell = sheet.cell(row=target_row, column=target_col)
                        if target_row in template_styles:
                            self._apply_cell_style_snapshot(target_cell, template_styles[target_row])
                        source_row = source_rows_by_key.get(target_row_keys.get(target_row))
                        if source_row is None:
                            target_cell.value = None
                            continue
                        source_cell = source_sheet.cell(row=source_row, column=source_col)
                        self._copy_cell_style(source_cell, target_cell)
                        target_cell.value = source_cell.value
                    target_letter = get_column_letter(target_col)
                    sheet.column_dimensions[target_letter].width = source_sheet.column_dimensions[get_column_letter(source_col)].width or template_width
                    target_col += 1
            workbook.save(output_path)
        finally:
            for source_wb in source_workbooks:
                source_wb.close()
            workbook.close()

    def _summary_row_keys(self, sheet) -> dict[int, tuple[str, str, int]]:
        row_keys: dict[int, tuple[str, str, int]] = {}
        section = 'sheet'
        counts: dict[tuple[str, str], int] = {}
        for row_index in range(1, sheet.max_row + 1):
            label = str(sheet.cell(row=row_index, column=1).value or '').strip()
            if row_index == 1:
                row_keys[row_index] = ('sheet', 'header', 1)
                continue
            if label.startswith('MTD'):
                section = 'mtd'
                counts = {}
                row_keys[row_index] = (section, 'section', 1)
                continue
            if label.startswith('Daily Report'):
                section = 'daily'
                counts = {}
                row_keys[row_index] = (section, 'section', 1)
                continue

            metric = self._summary_metric_key(label)
            if not metric:
                continue
            count_key = (section, metric)
            counts[count_key] = counts.get(count_key, 0) + 1
            row_keys[row_index] = (section, metric, counts[count_key])
        return row_keys

    def _summary_metric_key(self, value: str) -> str:
        text = str(value or '').strip()
        text = re.sub(r'[（(].*?[）)]', '', text)
        text = re.sub(r'\s+', '', text)
        return text

    def _worksheet_dimensions(self, path: Path) -> tuple[int, int]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            return sheet.max_row or 1, sheet.max_column or 1
        finally:
            workbook.close()

    def get_raw_files_status(self, report_date: str | None = None, store: str = 'all') -> dict[str, Any]:
        report_date = self._validate_report_date(report_date)
        store = (store or 'all').strip()
        store_name = self._validate_store(store)
        accounts = self._load_accounts()
        if store not in ('', 'all', '所有'):
            accounts = [account for account in accounts if str(account.get('name') or '').strip() == store]

        raw_dir = self.project_root / 'data' / report_date / '原始数据'
        items = []
        for account in accounts:
            name = str(account.get('name') or '').strip()
            clue_path = raw_dir / f'{name}-outcall-线索明细-{report_date}.xlsx'
            call_path = raw_dir / f'{name}-aicc-话单-{report_date}.xlsx'
            items.append({
                'store': name,
                'clue_exists': clue_path.exists(),
                'clue_path': str(clue_path) if clue_path.exists() else '',
                'call_exists': call_path.exists(),
                'call_path': str(call_path) if call_path.exists() else '',
            })

        return {
            'report_date': report_date,
            'store': store,
            'store_name': store_name,
            'raw_dir': str(raw_dir),
            'has_existing_clue': any(item['clue_exists'] for item in items),
            'items': items,
        }

    async def _run_job(self, job_id: str):
        job = self.jobs[job_id]
        env = os.environ.copy()
        env['REPORT_DATE'] = job.report_date
        if job.refresh_clue:
            env['REPORT_REFRESH_CLUE'] = '1'
        else:
            env.pop('REPORT_REFRESH_CLUE', None)
        if job.store != 'all':
            env['REPORT_STORE'] = job.store
        else:
            env.pop('REPORT_STORE', None)
        env['PYTHONIOENCODING'] = 'utf-8'

        try:
            proc = await asyncio.create_subprocess_exec(
                *job.command,
                cwd=job.cwd,
                env=env,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await proc.communicate()
            output = stdout.decode('utf-8', errors='replace')
            error_output = stderr.decode('utf-8', errors='replace')
            job.returncode = proc.returncode
            job.log = output + error_output
            job.reports = self._collect_reports(job.report_date, job.store, job.log)
            job.outputs = self._collect_outputs(job.report_date)
            job.status = 'completed' if proc.returncode == 0 else 'failed'
            if proc.returncode != 0:
                job.error = {
                    'code': 'DAILY_REPORT_EXECUTION_FAILED',
                    'message': 'Daily report script failed',
                    'detail': self._tail(job.log),
                }
        except Exception as exc:
            job.status = 'failed'
            job.log = f'{job.log}\n{exc}'.strip()
            job.error = {
                'code': 'DAILY_REPORT_EXECUTION_FAILED',
                'message': 'Daily report script failed',
                'detail': str(exc),
            }
        finally:
            job.completed_at = datetime.now().isoformat()
            job.updated_at = job.completed_at

    def _validate_report_date(self, report_date: str | None) -> str:
        value = (report_date or datetime.now().strftime('%y%m%d')).strip()
        if not re.fullmatch(r'\d{6}', value):
            raise InvalidReportDateError(value)
        try:
            datetime.strptime(value, '%y%m%d')
        except ValueError as exc:
            raise InvalidReportDateError(value) from exc
        return value

    def _validate_store(self, store: str) -> str:
        if store in ('', 'all', '所有'):
            return '所有'
        names = {str(account.get('name') or '').strip() for account in self._load_accounts()}
        if store not in names:
            raise InvalidReportStoreError(store)
        return store

    def _load_accounts(self) -> list[dict[str, Any]]:
        tools_dir = self.project_root / 'tools'
        if str(tools_dir) not in sys.path:
            sys.path.insert(0, str(tools_dir))
        import recorder
        recorder = importlib.reload(recorder)
        return list(recorder.ACCOUNTS)

    def _validate_preview_date(self, report_date: str | None) -> str:
        value = (report_date or '').strip()
        if value:
            return self._validate_report_date(value)
        return self._latest_report_date()

    def _latest_report_date(self) -> str:
        data_dir = self.project_root / 'data'
        if data_dir.exists():
            dates = sorted(
                path.name for path in data_dir.iterdir()
                if path.is_dir() and re.fullmatch(r'\d{6}', path.name)
            )
            if dates:
                return dates[-1]
        return self._validate_report_date(None)

    def _is_all_store(self, store: str) -> bool:
        return store in ('', 'all', '\u6240\u6709', '鎵€鏈?')

    def _preview_group_matches_store(self, group: dict[str, Any], store: str) -> bool:
        if self._is_all_store(store):
            return True
        return store in group.get('stores', []) or store == group.get('title')

    def _data_child_dir(self, report_date: str, dirname: str) -> Path:
        data_dir = self.project_root / 'data' / report_date
        direct = data_dir / dirname
        if direct.exists():
            return direct
        if data_dir.exists():
            for path in data_dir.iterdir():
                if path.is_dir() and dirname in path.name:
                    return path
        return direct

    def _daily_report_suffix(self, report_date: str) -> str:
        return f'_\u6bcf\u65e5\u62a5\u544a_{report_date}'

    def _summary_suffix(self, report_date: str) -> str:
        return f'_\u65e5\u5ea6\u6708\u5ea6\u6c47\u603b\u8868_{report_date}'

    def _collect_summaries(self, report_date: str) -> dict[str, dict[str, Any]]:
        summary_dir = self._data_child_dir(report_date, '\u6c47\u603b\u8868')
        summaries: dict[str, dict[str, Any]] = {}
        if not summary_dir.exists():
            return summaries
        suffix = self._summary_suffix(report_date)
        for path in sorted(summary_dir.glob(f'*{suffix}.xlsx')):
            if path.name.startswith('~$'):
                continue
            name = path.stem[:-len(suffix)] if path.stem.endswith(suffix) else path.stem
            summaries[name] = self._summary_preview(path, name)
        return summaries

    def _resolve_preview_group(self, group: str) -> dict[str, Any]:
        value = (group or '').strip()
        for item in REPORT_PREVIEW_GROUPS:
            if value in {item['key'], item['title'], item['summary']}:
                return item
        raise InvalidReportGroupError(value)

    def _summary_path(self, report_date: str, summary_name: str) -> Path:
        summary_dir = self._data_child_dir(report_date, '\u6c47\u603b\u8868')
        suffix = self._summary_suffix(report_date)
        direct = summary_dir / f'{summary_name}{suffix}.xlsx'
        if direct.exists():
            return direct
        matches = sorted(summary_dir.glob(f'{summary_name}*{suffix}.xlsx')) if summary_dir.exists() else []
        return matches[0] if matches else direct

    def _summary_image(self, path: Path) -> bytes:
        workbook = load_workbook(path, read_only=False, data_only=True)
        try:
            sheet = workbook.active
            image = self._worksheet_image(sheet)
        finally:
            workbook.close()
        buffer = BytesIO()
        image.save(buffer, format='PNG')
        return buffer.getvalue()

    def _worksheet_image(self, sheet) -> Image.Image:
        min_row = sheet.min_row or 1
        max_row = sheet.max_row or 1
        min_col = sheet.min_column or 1
        max_col = sheet.max_column or 1
        col_widths = [self._column_image_width(sheet, col_index) for col_index in range(min_col, max_col + 1)]
        row_heights = [self._row_image_height(sheet, row_index) for row_index in range(min_row, max_row + 1)]
        left_pad = 0
        top_pad = 0
        width = sum(col_widths) + left_pad * 2 + 2
        height = sum(row_heights) + top_pad * 2 + 2
        image = Image.new('RGB', (width, height), '#ffffff')
        draw = ImageDraw.Draw(image)

        x_offsets = [left_pad]
        for col_width in col_widths[:-1]:
            x_offsets.append(x_offsets[-1] + col_width)
        y_offsets = [top_pad]
        for row_height in row_heights[:-1]:
            y_offsets.append(y_offsets[-1] + row_height)

        merged_starts: dict[tuple[int, int], tuple[int, int]] = {}
        merged_children: set[tuple[int, int]] = set()
        for merged_range in sheet.merged_cells.ranges:
            min_c, min_r, max_c, max_r = merged_range.bounds
            merged_starts[(min_r, min_c)] = (max_r - min_r + 1, max_c - min_c + 1)
            for row_index in range(min_r, max_r + 1):
                for col_index in range(min_c, max_c + 1):
                    if (row_index, col_index) != (min_r, min_c):
                        merged_children.add((row_index, col_index))

        for row_index in range(min_row, max_row + 1):
            for col_index in range(min_col, max_col + 1):
                if (row_index, col_index) in merged_children:
                    continue
                rel_row = row_index - min_row
                rel_col = col_index - min_col
                rowspan, colspan = merged_starts.get((row_index, col_index), (1, 1))
                x1 = x_offsets[rel_col]
                y1 = y_offsets[rel_row]
                x2 = x1 + sum(col_widths[rel_col:rel_col + colspan])
                y2 = y1 + sum(row_heights[rel_row:rel_row + rowspan])
                cell = sheet.cell(row=row_index, column=col_index)
                fill = self._cell_fill_color(cell)
                draw.rectangle([x1, y1, x2, y2], fill=fill)
                self._draw_cell_text(draw, cell, x1, y1, x2, y2)
                self._draw_cell_border(draw, cell, x1, y1, x2, y2)

        return image

    def _column_image_width(self, sheet, col_index: int) -> int:
        letter = get_column_letter(col_index)
        width = sheet.column_dimensions[letter].width or 10
        return max(76, min(int(float(width) * 13 + 18), 520))

    def _row_image_height(self, sheet, row_index: int) -> int:
        height = sheet.row_dimensions[row_index].height
        if height:
            return max(24, int(float(height) * 1.55))
        return 28

    def _cell_fill_color(self, cell) -> str:
        fill = cell.fill
        if fill and fill.fill_type and fill.fill_type != 'none':
            return self._excel_color(fill.fgColor) or '#ffffff'
        return '#ffffff'

    def _draw_cell_text(self, draw: ImageDraw.ImageDraw, cell, x1: int, y1: int, x2: int, y2: int) -> None:
        text = self._display_cell_value(cell)
        if not text:
            return
        font = self._image_font(cell.font)
        color = self._excel_color(cell.font.color) if cell.font else ''
        color = color or '#000000'
        padding_x = 7
        padding_y = 4
        max_width = max(10, x2 - x1 - padding_x * 2)
        lines = []
        for part in str(text).split('\n'):
            lines.extend(self._wrap_text(draw, part, font, max_width) or [''])
        line_boxes = [draw.textbbox((0, 0), line, font=font) for line in lines]
        line_heights = [box[3] - box[1] for box in line_boxes]
        line_gap = 3
        total_height = sum(line_heights) + max(0, len(lines) - 1) * line_gap
        alignment = cell.alignment
        vertical = alignment.vertical if alignment else None
        horizontal = alignment.horizontal if alignment else None
        if vertical == 'center':
            y = y1 + max(padding_y, (y2 - y1 - total_height) // 2)
        elif vertical == 'bottom':
            y = y2 - total_height - padding_y
        else:
            y = y1 + padding_y
        for index, line in enumerate(lines):
            box = line_boxes[index]
            text_width = box[2] - box[0]
            if horizontal in {'center', 'centerContinuous'}:
                x = x1 + max(padding_x, (x2 - x1 - text_width) // 2)
            elif horizontal == 'right':
                x = x2 - text_width - padding_x
            else:
                x = x1 + padding_x
            draw.text((x, y), line, fill=color, font=font)
            y += line_heights[index] + line_gap

    def _wrap_text(self, draw: ImageDraw.ImageDraw, text: str, font, max_width: int) -> list[str]:
        if not text:
            return ['']
        lines: list[str] = []
        current = ''
        for char in text:
            candidate = current + char
            if current and draw.textlength(candidate, font=font) > max_width:
                lines.append(current)
                current = char
            else:
                current = candidate
        if current:
            lines.append(current)
        return lines

    def _draw_cell_border(self, draw: ImageDraw.ImageDraw, cell, x1: int, y1: int, x2: int, y2: int) -> None:
        border = cell.border
        if not border:
            draw.rectangle([x1, y1, x2, y2], outline='#d0d5dd', width=1)
            return
        sides = {
            'left': ((x1, y1), (x1, y2)),
            'right': ((x2, y1), (x2, y2)),
            'top': ((x1, y1), (x2, y1)),
            'bottom': ((x1, y2), (x2, y2)),
        }
        for edge, points in sides.items():
            side = getattr(border, edge)
            if side and side.style:
                color = self._excel_color(side.color) or '#b8c0cc'
                width = 2 if side.style in {'medium', 'thick'} else 1
            else:
                color = '#cfd6df'
                width = 1
            draw.line(points, fill=color, width=width)

    def _image_font(self, font):
        size = 13
        if font and font.sz:
            size = max(11, int(float(font.sz) * 1.35))
        font_path = self._font_path(bool(font and font.bold))
        if font_path:
            try:
                return ImageFont.truetype(font_path, size=size)
            except OSError:
                pass
        return ImageFont.load_default(size=size)

    def _font_path(self, bold: bool = False) -> str:
        candidates = [
            r'C:\Windows\Fonts\msyhbd.ttc' if bold else r'C:\Windows\Fonts\msyh.ttc',
            r'C:\Windows\Fonts\simhei.ttf',
            r'C:\Windows\Fonts\simsun.ttc',
            r'C:\Windows\Fonts\arialbd.ttf' if bold else r'C:\Windows\Fonts\arial.ttf',
        ]
        for candidate in candidates:
            if candidate and Path(candidate).exists():
                return candidate
        return ''

    def _summary_preview(self, path: Path, name: str) -> dict[str, Any]:
        workbook = load_workbook(path, read_only=False, data_only=True)
        try:
            sheet = workbook.active
            rows: list[list[str]] = []
            for raw_row in sheet.iter_rows(values_only=False):
                row = [self._display_cell_value(cell) for cell in raw_row]
                while row and row[-1] == '':
                    row.pop()
                if row:
                    rows.append(row)
            html_table = self._worksheet_html(sheet)
        finally:
            workbook.close()

        max_cols = max((len(row) for row in rows), default=0)
        normalized = [row + [''] * (max_cols - len(row)) for row in rows]
        columns = normalized[0] if normalized else []
        body = normalized[1:] if len(normalized) > 1 else []
        content = '\n'.join('\t'.join(row) for row in normalized)
        return {
            'name': name,
            'filename': path.name,
            'path': str(path),
            'columns': columns,
            'rows': body,
            'content': content,
            'html': html_table,
            'missing': False,
        }

    def _worksheet_html(self, sheet) -> str:
        min_row = sheet.min_row or 1
        max_row = sheet.max_row or 1
        min_col = sheet.min_column or 1
        max_col = sheet.max_column or 1
        merged_starts: dict[tuple[int, int], tuple[int, int]] = {}
        merged_children: set[tuple[int, int]] = set()
        for merged_range in sheet.merged_cells.ranges:
            min_c, min_r, max_c, max_r = merged_range.bounds
            merged_starts[(min_r, min_c)] = (max_r - min_r + 1, max_c - min_c + 1)
            for row_index in range(min_r, max_r + 1):
                for col_index in range(min_c, max_c + 1):
                    if (row_index, col_index) != (min_r, min_c):
                        merged_children.add((row_index, col_index))

        parts = [
            '<table class="summary-table excel-summary-table" style="border-collapse:collapse;border-spacing:0;table-layout:fixed;min-width:100%;">',
            '<colgroup>',
        ]
        for col_index in range(min_col, max_col + 1):
            letter = get_column_letter(col_index)
            width = sheet.column_dimensions[letter].width or 10
            pixels = max(42, min(int(float(width) * 7 + 8), 260))
            parts.append(f'<col style="width:{pixels}px;">')
        parts.append('</colgroup><tbody>')

        for row_index in range(min_row, max_row + 1):
            row_height = sheet.row_dimensions[row_index].height
            row_style = f' style="height:{int(row_height * 1.333)}px;"' if row_height else ''
            parts.append(f'<tr{row_style}>')
            for col_index in range(min_col, max_col + 1):
                if (row_index, col_index) in merged_children:
                    continue
                cell = sheet.cell(row=row_index, column=col_index)
                rowspan, colspan = merged_starts.get((row_index, col_index), (1, 1))
                attrs = []
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')
                style = self._cell_inline_style(cell)
                if style:
                    attrs.append(f'style="{html.escape(style, quote=True)}"')
                value = html.escape(self._display_cell_value(cell), quote=False).replace('\n', '<br>')
                attr_text = (' ' + ' '.join(attrs)) if attrs else ''
                parts.append(f'<td{attr_text}>{value}</td>')
            parts.append('</tr>')
        parts.append('</tbody></table>')
        return ''.join(parts)

    def _cell_inline_style(self, cell) -> str:
        styles = [
            'border:1px solid #d0d5dd',
            'padding:4px 6px',
            'font-size:12px',
            'line-height:1.35',
            'vertical-align:middle',
            'white-space:pre-wrap',
            'word-break:break-word',
            'background:#ffffff',
            'color:#111827',
        ]

        font = cell.font
        if font:
            if font.name:
                styles.append(f'font-family:{self._css_string(font.name)}')
            if font.sz:
                styles.append(f'font-size:{float(font.sz):g}pt')
            if font.bold:
                styles.append('font-weight:700')
            if font.italic:
                styles.append('font-style:italic')
            if font.underline:
                styles.append('text-decoration:underline')
            color = self._excel_color(font.color)
            if color:
                styles.append(f'color:{color}')

        fill = cell.fill
        if fill and fill.fill_type and fill.fill_type != 'none':
            color = self._excel_color(fill.fgColor)
            if color:
                styles.append(f'background-color:{color}')

        alignment = cell.alignment
        if alignment:
            if alignment.horizontal:
                horizontal = {'centerContinuous': 'center'}.get(alignment.horizontal, alignment.horizontal)
                styles.append(f'text-align:{horizontal}')
            if alignment.vertical:
                styles.append(f'vertical-align:{alignment.vertical}')
            if alignment.wrap_text:
                styles.append('white-space:pre-wrap')
            else:
                styles.append('white-space:normal')

        border = cell.border
        if border:
            for edge in ('left', 'right', 'top', 'bottom'):
                side_style = self._css_border(getattr(border, edge), edge)
                if side_style:
                    styles.append(side_style)

        number_format = str(cell.number_format or '').strip()
        if number_format and number_format != 'General':
            styles.append(f'mso-number-format:{self._css_string(number_format)}')

        return ';'.join(styles)

    def _css_border(self, side, edge: str) -> str:
        if not side or not side.style:
            return ''
        style = str(side.style)
        width = {
            'hair': '1px',
            'thin': '1px',
            'medium': '2px',
            'thick': '3px',
        }.get(style, '1px')
        line = {
            'dashed': 'dashed',
            'dashDot': 'dashed',
            'dashDotDot': 'dashed',
            'dotted': 'dotted',
            'double': 'double',
            'mediumDashed': 'dashed',
        }.get(style, 'solid')
        color = self._excel_color(side.color) or '#d0d5dd'
        return f'border-{edge}:{width} {line} {color}'

    def _excel_color(self, color) -> str:
        if not color or not getattr(color, 'type', None):
            return ''
        if color.type == 'rgb' and color.rgb:
            value = str(color.rgb)[-6:]
            if value:
                return f'#{value}'
        return ''

    def _css_string(self, value: str) -> str:
        return '"' + str(value).replace('"', '\\"') + '"'

    def _display_cell_value(self, cell) -> str:
        value = cell.value
        if value is None:
            return ''
        if isinstance(value, datetime):
            if value.time() == time(0, 0):
                return value.strftime('%Y-%m-%d')
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, time):
            return value.strftime('%H:%M:%S')
        if isinstance(value, (int, float)):
            number_format = str(cell.number_format or '')
            if '%' in number_format:
                decimals = 0
                match = re.search(r'0\.([0#]+)', number_format)
                if match:
                    decimals = len(match.group(1))
                unquoted_format = re.sub(r'"[^"]*"', '', number_format)
                display_value = value * 100 if '%' in unquoted_format else value
                return f'{display_value:.{decimals}f}%'
            match = re.search(r'0\.([0]+)', number_format)
            if match and 'E+' not in number_format:
                decimals = len(match.group(1))
                return f'{value:.{decimals}f}'
        return str(value)

    def _collect_reports(self, report_date: str, store: str, log: str) -> list[DailyReportPreview]:
        report_dir = self.project_root / 'data' / report_date / '每日报告'
        if not report_dir.exists():
            return []

        filenames = []
        for match in re.finditer(r'报告:\s*([^\r\n]+?\.txt)', log):
            filenames.append(match.group(1).strip())

        paths: list[Path]
        if filenames:
            seen = set()
            paths = []
            for filename in filenames:
                path = report_dir / filename
                if path.exists() and path.name not in seen:
                    paths.append(path)
                    seen.add(path.name)
        elif store not in ('all', '所有'):
            paths = sorted(report_dir.glob(f'{store}_每日报告_{report_date}.txt'))
        else:
            paths = sorted(report_dir.glob(f'*_每日报告_{report_date}.txt'))

        return [self._report_preview(path, report_date) for path in paths]

    def _report_preview(self, path: Path, report_date: str) -> DailyReportPreview:
        content = path.read_text(encoding='utf-8', errors='replace').strip()
        suffix = f'_每日报告_{report_date}'
        name = path.stem[:-len(suffix)] if path.stem.endswith(suffix) else path.stem
        return DailyReportPreview(name=name, filename=path.name, path=str(path), content=content)

    def _collect_outputs(self, report_date: str) -> list[DailyReportFile]:
        out_dir = self.project_root / 'data' / report_date
        if not out_dir.exists():
            return []
        outputs = []
        for path in sorted(out_dir.rglob('*')):
            if path.is_file() and not path.name.startswith('~$'):
                kind = 'report' if '每日报告' in path.parts else ('summary' if '汇总表' in path.parts else 'data')
                outputs.append(DailyReportFile(filename=path.name, path=str(path), kind=kind, size=path.stat().st_size))
        return outputs

    def _tail(self, text: str, limit: int = 1200) -> str:
        return text[-limit:] if len(text) > limit else text


def job_to_dict(job: DailyReportJob) -> dict[str, Any]:
    return asdict(job)