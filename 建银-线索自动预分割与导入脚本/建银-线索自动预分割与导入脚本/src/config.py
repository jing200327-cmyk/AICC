"""
配置加载与 Excel 文件扫描模块
"""
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import re
import yaml


@dataclass
class EnvironmentConfig:
    username: str
    password: str
    robot_id: str
    dealer_id: str = ""


@dataclass
class TenantConfig:
    name: str
    prefixes: List[str]
    environments: Dict[str, EnvironmentConfig]
    task_name_template: str = "{tenant}-{batch}"

    def get_env_config(self, env: str) -> EnvironmentConfig:
        return self.environments[env]


@dataclass
class Settings:
    client_id: str = "110011"
    max_leads_per_window: int = 100
    time_window_minutes: int = 30
    poll_interval_seconds: int = 30
    request_timeout_seconds: int = 30
    max_retries: int = 3


@dataclass
class ExcelFile:
    file_path: Path
    tenant_name: str
    batch_name: str
    tenant_prefix: str


@dataclass
class AppConfig:
    environment: str
    settings: Settings
    tenants: List[TenantConfig]
    base_url: str = ""


def load_config(path: str) -> AppConfig:
    with open(path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f)

    environment = raw.get("environment", "test")
    settings_data = raw.get("settings", {})
    settings = Settings(**{k: v for k, v in settings_data.items() if v is not None})

    tenants = []
    for t in raw.get("tenants", []):
        env_configs = {}
        for env_key in ("test", "prod"):
            env_data = t.get("environments", {}).get(env_key, {})
            if env_data:
                env_configs[env_key] = EnvironmentConfig(
                    username=env_data.get("username", ""),
                    password=env_data.get("password", ""),
                    robot_id=env_data.get("robot_id", ""),
                    dealer_id=env_data.get("dealer_id", ""),
                )
        prefixes = t.get("prefixes", [t.get("prefix", t["name"])])
        if isinstance(prefixes, str):
            prefixes = [prefixes]
        tenant = TenantConfig(
            name=t["name"],
            prefixes=prefixes,
            environments=env_configs,
            task_name_template=t.get("task_name_template", "{tenant}-{batch}"),
        )
        tenants.append(tenant)

    _BASE_URLS = {"test": "https://uat.aidcc.cn", "prod": "https://service.aidcc.cn"}
    base_url = _BASE_URLS.get(environment, _BASE_URLS["test"])

    return AppConfig(
        environment=environment,
        settings=settings,
        tenants=tenants,
        base_url=base_url,
    )


def _extract_batch_name(filename_stem: str, prefix: str) -> Optional[str]:
    pattern = re.compile(rf"^{re.escape(prefix)}-(.+)$")
    m = pattern.match(filename_stem)
    if m:
        return m.group(1)
    if filename_stem == prefix:
        return "default"
    return None


def _batch_sort_key(batch_name: str):
    if batch_name == "测试":
        return (0, batch_name)
    try:
        return (1, int(batch_name))
    except ValueError:
        return (2, batch_name)


def _find_date_dirs(base_dir: Path) -> List[Path]:
    dirs = []
    for d in base_dir.iterdir():
        if d.is_dir() and re.match(r"^\d{6}$", d.name):
            dirs.append(d)
    dirs.sort(reverse=True)
    return dirs


def _find_today_presplit_dir(base_dir: Path) -> Optional[Path]:
    today_dir_name = datetime.now().strftime("%y%m%d")
    scan_dir = base_dir / "线索预分割" / today_dir_name
    if scan_dir.is_dir():
        return scan_dir
    return None


def _resolve_scan_dirs(data_dir: Optional[str], settings_data_dir: Optional[str]) -> List[Path]:
    if data_dir:
        scan_dir = Path(data_dir)
        if scan_dir.is_dir():
            return [scan_dir]
        print(f"警告: 指定目录不存在: {data_dir}")
        return []
    if settings_data_dir:
        scan_dir = Path(settings_data_dir)
        if scan_dir.is_dir():
            return [scan_dir]
        print(f"警告: 配置中的 data_dir 不存在: {settings_data_dir}")
        return []
    today_presplit_dir = _find_today_presplit_dir(Path.cwd())
    if today_presplit_dir:
        return [today_presplit_dir]
    date_dirs = _find_date_dirs(Path.cwd())
    if date_dirs:
        return [date_dirs[0]]
    return [Path.cwd()]


def scan_files(config: AppConfig, data_dir: Optional[str] = None) -> Dict[str, List[ExcelFile]]:
    scan_dirs = _resolve_scan_dirs(data_dir, config.settings.data_dir if hasattr(config.settings, 'data_dir') else None)

    tenant_files: Dict[str, List[ExcelFile]] = {t.name: [] for t in config.tenants}

    for scan_dir in scan_dirs:
        for xlsx_file in scan_dir.glob("*.xlsx"):
            stem = xlsx_file.stem
            matched = False
            for tenant in config.tenants:
                for prefix in tenant.prefixes:
                    batch_name = _extract_batch_name(stem, prefix)
                    if batch_name is not None:
                        ef = ExcelFile(
                            file_path=xlsx_file,
                            tenant_name=tenant.name,
                            batch_name=batch_name,
                            tenant_prefix=prefix,
                        )
                        tenant_files[tenant.name].append(ef)
                        matched = True
                        break
                if matched:
                    break

    for name in tenant_files:
        tenant_files[name].sort(key=lambda f: _batch_sort_key(f.batch_name))

    return tenant_files


def count_leads_in_excel(file_path: Path) -> int:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            count += 1
    wb.close()
    return count
