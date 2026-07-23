# -*- coding: utf-8 -*-
import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HEADER_FILL = PatternFill(start_color="4E83FD", end_color="4E83FD", fill_type="solid")
SECTION_FILL = PatternFill(start_color="BACEFD", end_color="BACEFD", fill_type="solid")
SUBSECTION_FILL = PatternFill(start_color="DEE0E3", end_color="DEE0E3", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
DATA_FONT = Font(size=10)
WHITE_FONT = Font(bold=True, size=11, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

TRACKING_DIR = Path(
    os.environ.get('REPORT_TRACKING_DIR')
    or Path(__file__).parent.parent / '_tracking'
)
MTD_BACKFILL_KEY = "_mtd_backfill"
MTD_SNAPSHOT_FIELDS = {
    '累计线索量': 'MTD累计线索量',
    '累计接通量': 'MTD累计接通量',
    '累计有效线索量': 'MTD累计有效线索量',
    '累计呼叫通次': 'MTD累计呼叫通次',
}


def _safe_div(a, b):
    return round(a / b * 100, 2) if b else 0.0


def _load_tracking(company: str, year_month: str) -> dict:
    TRACKING_DIR.mkdir(exist_ok=True)
    month_dir = TRACKING_DIR / year_month
    fp = month_dir / f"{company}.json"
    legacy_fp = TRACKING_DIR / f"{company}_{year_month}.json"
    if fp.exists():
        return json.loads(fp.read_text(encoding='utf-8'))
    if legacy_fp.exists():
        return json.loads(legacy_fp.read_text(encoding='utf-8'))
    return {}


def _load_tracking_with_aliases(company: str, year_month: str, aliases: list[str] | None = None) -> dict:
    records = _load_tracking(company, year_month)
    if records:
        return records

    for alias in aliases or []:
        if alias and alias != company:
            records = _load_tracking(alias, year_month)
            if records:
                return records
    return {}


def _save_tracking(company: str, year_month: str, records: dict):
    TRACKING_DIR.mkdir(exist_ok=True)
    month_dir = TRACKING_DIR / year_month
    month_dir.mkdir(exist_ok=True)
    (month_dir / f"{company}.json").write_text(
        json.dumps(records, ensure_ascii=False, indent=2), encoding='utf-8')


def _previous_date_key(date: str) -> str:
    return (datetime.strptime(date, "%y%m%d") - timedelta(days=1)).strftime("%y%m%d")


def _load_previous_tracking(company: str, date: str, aliases: list[str] | None = None) -> dict:
    prev_date = _previous_date_key(date)
    prev_tracking = _load_tracking_with_aliases(company, prev_date[:4], aliases)
    return prev_tracking.get(prev_date, {})


def _tracking_metric(record: dict, key: str) -> int:
    value = record.get(key, 0)
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def aggregate_tracking_mtd(
    tracking: dict,
    report_date: str,
) -> tuple[dict[str, int], list[str]]:
    """Aggregate calendar-month MTD metrics from daily tracking records."""
    report_dt = datetime.strptime(report_date, '%y%m%d')
    month_start = report_dt.replace(day=1)
    records = {
        key: value
        for key, value in tracking.items()
        if isinstance(value, dict)
        and len(key) == 6
        and key.isdigit()
        and key[:4] == report_date[:4]
        and key <= report_date
    }
    expected_dates = []
    current = month_start
    while current <= report_dt:
        expected_dates.append(current.strftime('%y%m%d'))
        current += timedelta(days=1)
    totals = {
        '累计线索量': sum(
            _tracking_metric(item, '新增线索量') for item in records.values()
        ),
        '累计接通量': sum(
            _tracking_metric(item, '新增线索接通量') for item in records.values()
        ),
        '累计有效线索量': sum(
            _tracking_metric(item, '有效线索量') for item in records.values()
        ),
        '累计呼叫通次': sum(
            _tracking_metric(item, '呼叫通次') for item in records.values()
        ),
    }

    report_record = records.get(report_date, {})
    has_snapshot = all(
        field in report_record for field in MTD_SNAPSHOT_FIELDS.values()
    )
    if has_snapshot:
        totals = {
            metric: _tracking_metric(report_record, field)
            for metric, field in MTD_SNAPSHOT_FIELDS.items()
        }

    covered_dates = set()
    for adjustment in tracking.get(MTD_BACKFILL_KEY, {}).values():
        if not isinstance(adjustment, dict):
            continue
        adjustment_dates = adjustment.get('covered_dates') or []
        if not adjustment_dates or not all(
            isinstance(key, str)
            and len(key) == 6
            and key.isdigit()
            and key[:4] == report_date[:4]
            and key <= report_date
            for key in adjustment_dates
        ):
            continue
        covered_dates.update(adjustment_dates)
        if has_snapshot:
            continue
        totals['累计线索量'] += _tracking_metric(adjustment, '新增线索量')
        totals['累计接通量'] += _tracking_metric(
            adjustment, '新增线索接通量'
        )
        totals['累计有效线索量'] += _tracking_metric(
            adjustment, '有效线索量'
        )
        totals['累计呼叫通次'] += _tracking_metric(adjustment, '呼叫通次')

    missing_dates = [
        key
        for key in expected_dates
        if key not in records and key not in covered_dates
    ]
    return totals, missing_dates


def _tracking_record(data: dict, monthly_stats: dict | None = None) -> dict:
    record = {
        "呼叫通次": data.get("call_count", 0),
        "新增线索量": data.get("total", 0),
        "新增线索接通量": data.get("jietong_count", 0),
        "有效线索量": data.get("yixiang_count", 0),
        "已接通通话平均时长分钟": data.get(
            "connected_avg_duration_minutes", 0
        ),
    }
    for metric, field in MTD_SNAPSHOT_FIELDS.items():
        if monthly_stats and metric in monthly_stats:
            record[field] = _tracking_metric(monthly_stats, metric)
    return record


def _summary_candidates(report_date: str) -> list[Path]:
    data_root = TRACKING_DIR.parent / 'data'
    direct_dir = data_root / report_date / '汇总表'
    candidates = list(direct_dir.glob(f'*_{report_date}.xlsx'))
    if not data_root.exists():
        return candidates

    for period_dir in data_root.glob(f'{report_date[:4]}*_*'):
        parts = period_dir.name.split('_', 1)
        if len(parts) != 2 or not (parts[0] <= report_date <= parts[1]):
            continue
        candidates.extend(period_dir.rglob(f'*_{report_date}.xlsx'))
    return list(dict.fromkeys(candidates))


def _percentage_points(value) -> float | None:
    if value is None or value == '':
        return None
    if isinstance(value, str):
        text = value.strip().rstrip('%')
        try:
            return float(text)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _effective_count_from_report(
    report_date: str,
    company_names: list[str],
) -> int | None:
    report_dir = TRACKING_DIR.parent / 'data' / report_date / '每日报告'
    for company in company_names:
        path = report_dir / f'{company}_每日报告_{report_date}.txt'
        if not path.exists():
            continue
        text = path.read_text(encoding='utf-8', errors='replace')
        match = re.search(r'意向线索[（(]\s*(\d+)\s*条', text)
        if match:
            return int(match.group(1))
    return None


def _tracking_record_from_summary(
    path: Path,
    report_date: str,
    company_names: list[str],
) -> dict | None:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        column_index = next(
            (
                index
                for index, value in enumerate(header)
                if str(value or '').strip() in company_names
            ),
            None,
        )
        if column_index is None:
            return None

        values = {}
        in_daily_section = False
        for row in sheet.iter_rows(values_only=True):
            label = str(row[0] or '').strip()
            if label.startswith('Daily Report'):
                in_daily_section = True
                continue
            if not in_daily_section:
                continue
            if column_index >= len(row):
                continue
            if label in {
                '新增线索量',
                '新增线索接通量',
                '新增线索呼叫通次',
                '已接通通话平均时长（向上取整）',
            }:
                values[label] = row[column_index]
            elif label.startswith('接通有效率'):
                values['接通有效率'] = row[column_index]

        required = {'新增线索量', '新增线索接通量', '新增线索呼叫通次'}
        if not required.issubset(values):
            return None

        effective = _effective_count_from_report(report_date, company_names)
        if effective is None:
            rate = _percentage_points(values.get('接通有效率'))
            connected = _tracking_metric(values, '新增线索接通量')
            effective = round(connected * rate / 100) if rate is not None else 0
        return {
            '呼叫通次': _tracking_metric(values, '新增线索呼叫通次'),
            '新增线索量': _tracking_metric(values, '新增线索量'),
            '新增线索接通量': _tracking_metric(values, '新增线索接通量'),
            '有效线索量': int(effective),
            '已接通通话平均时长分钟': _tracking_metric(
                values,
                '已接通通话平均时长（向上取整）',
            ),
        }
    finally:
        workbook.close()


def _mtd_record_from_summary(
    path: Path,
    company_names: list[str],
) -> dict | None:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        column_index = next(
            (
                index
                for index, value in enumerate(header)
                if str(value or '').strip() in company_names
            ),
            None,
        )
        if column_index is None:
            return None

        labels = {
            '累计线索量': '新增线索量',
            '接通量': '新增线索接通量',
            '有效线索量': '有效线索量',
            '呼叫通次': '呼叫通次',
        }
        values = {}
        for row in sheet.iter_rows(values_only=True):
            label = str(row[0] or '').strip()
            if label.startswith('Daily Report'):
                break
            target = labels.get(label)
            if target and column_index < len(row):
                values[target] = row[column_index]
        if not set(labels.values()).issubset(values):
            return None
        return {
            key: _tracking_metric(values, key)
            for key in labels.values()
        }
    finally:
        workbook.close()


def _tracking_source_label(path: Path) -> str:
    try:
        return str(path.relative_to(TRACKING_DIR.parent))
    except ValueError:
        return str(path)


def backfill_tracking_mtd_gaps_from_summaries(
    company: str,
    report_date: str,
    tracking: dict,
    aliases: list[str] | None = None,
) -> tuple[list[str], list[str]]:
    """Backfill unresolved MTD gaps from the earliest later MTD snapshot.

    Consecutive days without their own daily report cannot be split reliably.
    Store the verified aggregate delta with provenance instead of inventing a
    per-day distribution.
    """
    tracking.pop(MTD_BACKFILL_KEY, None)
    tracking_dates = [
        key
        for key, value in tracking.items()
        if isinstance(value, dict)
        and len(key) == 6
        and key.isdigit()
        and key[:4] == report_date[:4]
    ]
    backfill_date = max([report_date, *tracking_dates])
    _, unresolved = aggregate_tracking_mtd(tracking, backfill_date)
    if not unresolved:
        return [], []

    company_names = [company] + [
        alias for alias in aliases or [] if alias and alias != company
    ]
    remaining = set(unresolved)
    adjustments = {}
    report_dt = datetime.strptime(backfill_date, '%y%m%d')

    while remaining:
        first_missing = min(remaining)
        anchor_dt = datetime.strptime(first_missing, '%y%m%d')
        anchor_record = None
        anchor_path = None
        anchor_date = None
        while anchor_dt <= report_dt:
            candidate_date = anchor_dt.strftime('%y%m%d')
            for path in _summary_candidates(candidate_date):
                record = _mtd_record_from_summary(path, company_names)
                if record is not None:
                    anchor_record = record
                    anchor_path = path
                    anchor_date = candidate_date
                    break
            if anchor_record is not None:
                break
            anchor_dt += timedelta(days=1)

        if anchor_record is None or anchor_date is None or anchor_path is None:
            break

        covered_dates = sorted(key for key in remaining if key <= anchor_date)
        known, _ = aggregate_tracking_mtd(tracking, anchor_date)
        residual = {
            '新增线索量': (
                anchor_record['新增线索量'] - known['累计线索量']
            ),
            '新增线索接通量': (
                anchor_record['新增线索接通量'] - known['累计接通量']
            ),
            '有效线索量': (
                anchor_record['有效线索量'] - known['累计有效线索量']
            ),
            '呼叫通次': (
                anchor_record['呼叫通次'] - known['累计呼叫通次']
            ),
        }
        if any(value < 0 for value in residual.values()):
            print(
                f'  MTD tracking缺口无法回填: {company} {first_missing}，'
                f'锚点汇总小于现有tracking累计 ({anchor_path.name})'
            )
            break

        key = (
            covered_dates[0]
            if len(covered_dates) == 1
            else f'{covered_dates[0]}-{covered_dates[-1]}'
        )
        adjustments[key] = {
            **residual,
            'covered_dates': covered_dates,
            'source': _tracking_source_label(anchor_path),
            'source_mtd_date': anchor_date,
            'method': 'existing_summary_mtd_delta',
        }
        tracking[MTD_BACKFILL_KEY] = adjustments
        remaining.difference_update(covered_dates)
        print(
            f'  MTD tracking区间回填: {company} {key}，'
            f'新增线索 {residual["新增线索量"]} 条，'
            f'依据 {anchor_path.name}'
        )

    _, unresolved = aggregate_tracking_mtd(tracking, backfill_date)
    return (
        list(adjustments),
        [key for key in unresolved if key <= report_date],
    )


def backfill_tracking_from_daily_reports(
    company: str,
    report_date: str,
    tracking: dict,
    aliases: list[str] | None = None,
) -> tuple[list[str], list[str]]:
    tracking_without_adjustments = dict(tracking)
    tracking_without_adjustments.pop(MTD_BACKFILL_KEY, None)
    _, missing_dates = aggregate_tracking_mtd(
        tracking_without_adjustments,
        report_date,
    )
    company_names = [company] + [
        alias for alias in aliases or [] if alias and alias != company
    ]
    backfilled = []
    for missing_date in missing_dates:
        record = None
        for path in _summary_candidates(missing_date):
            record = _tracking_record_from_summary(
                path,
                missing_date,
                company_names,
            )
            if record is not None:
                break
        if record is None:
            continue
        tracking[missing_date] = record
        backfilled.append(missing_date)
        print(
            f'  MTD tracking回填: {company} {missing_date} '
            f'新增线索 {record["新增线索量"]} 条'
        )

    _, unresolved = aggregate_tracking_mtd(tracking, report_date)
    return backfilled, unresolved


def _print_tracking_mtd_status(
    company: str,
    report_date: str,
    missing_dates: list[str],
) -> None:
    tracking_path = TRACKING_DIR / report_date[:4] / f'{company}.json'
    print(f'  MTD tracking累计: {tracking_path}')
    if missing_dates:
        print(
            '  MTD tracking缺失日期: '
            + '、'.join(missing_dates)
            + '；缺失日期不会被自动按0伪造，请补齐tracking后重跑。'
        )


def generate_daily_summary(company: str, date: str, data: dict, output_dir: str | Path,
                           monthly_stats: dict | None = None) -> Path:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    year_month = date[:4]

    # === Tracking ===
    tracking = _load_tracking(company, year_month)
    tracking[date] = _tracking_record(data, monthly_stats)
    backfill_tracking_from_daily_reports(
        company,
        date,
        tracking,
    )
    _, missing_dates = backfill_tracking_mtd_gaps_from_summaries(
        company,
        date,
        tracking,
    )
    _save_tracking(company, year_month, tracking)

    mtd_stats, _ = aggregate_tracking_mtd(tracking, date)
    _print_tracking_mtd_status(company, date, missing_dates)
    mtd_total = mtd_stats['累计线索量']
    mtd_jietong = mtd_stats['累计接通量']
    mtd_yixiang = mtd_stats['累计有效线索量']
    mtd_call_count = mtd_stats['累计呼叫通次']

    # 昨日对比数据：月初时需要从上月 tracking 文件读取前一天。
    yesterday = _load_previous_tracking(company, date)

    # ===== 生成 Excel =====
    wb = Workbook()
    ws = wb.active
    ws.title = company

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 16

    for col, val in [(1, "业务场景"), (2, "单位"), (3, company)]:
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    def _section_row(row_num, label):
        for c in range(1, 4):
            ws.cell(row=row_num, column=c).fill = SECTION_FILL
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        ws.cell(row=row_num, column=1, value=label).font = HEADER_FONT

    def _subsection_row(row_num, label):
        for c in range(1, 4):
            ws.cell(row=row_num, column=c).fill = SUBSECTION_FILL
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        ws.cell(row=row_num, column=1, value=label).font = HEADER_FONT

    def _data_row(row_num, name, unit, value):
        ws.cell(row=row_num, column=1, value=name).font = DATA_FONT
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=unit).font = DATA_FONT
        ws.cell(row=row_num, column=2).border = THIN_BORDER
        cell = ws.cell(row=row_num, column=3, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER

    def _pct_row(row_num, name, unit, value):
        ws.cell(row=row_num, column=1, value=name).font = DATA_FONT
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=unit).font = DATA_FONT
        ws.cell(row=row_num, column=2).border = THIN_BORDER
        cell = ws.cell(row=row_num, column=3, value=value)
        cell.font = DATA_FONT
        cell.number_format = '0.00"%"'
        cell.border = THIN_BORDER

    r = 1

    date_display = f"20{date[:2]}-{date[2:4]}-{date[4:6]}"

    # === MTD Section ===
    r += 1; _section_row(r, f"MTD (截止：{date_display})")

    mtd_avg_call = round(mtd_call_count / mtd_total, 2) if mtd_total else 0.0
    mtd_jietong_rate = _safe_div(mtd_jietong, mtd_total)
    mtd_youxiao_rate = _safe_div(mtd_yixiang, mtd_jietong)
    mtd_overall_rate = _safe_div(mtd_yixiang, mtd_total)

    r += 1; _subsection_row(r, "规模指标")
    r += 1; _data_row(r, "累计线索量", "条", mtd_total)
    r += 1; _data_row(r, "接通量", "条", mtd_jietong)
    r += 1; _data_row(r, "有效线索量", "条", mtd_yixiang)
    r += 1; _data_row(r, "呼叫通次", "次", mtd_call_count)
    r += 1; _data_row(r, "平均呼叫通次", "次/条", mtd_avg_call)

    r += 1; _subsection_row(r, "检测指标")
    r += 1; _pct_row(r, "接通率", "%", mtd_jietong_rate)
    r += 1; _pct_row(r, "接通有效率", "%", mtd_youxiao_rate)
    r += 1; _pct_row(r, "整体有效率", "%", mtd_overall_rate)

    # === Daily Report Section ===
    r += 1; _section_row(r, f"Daily Report ({date_display})")

    today_total = data.get("total", 0)
    today_jietong = data.get("jietong_count", 0)
    today_call = data.get("call_count", 0)
    today_yixiang = data.get("yixiang_count", 0)
    today_connected_avg_duration_minutes = data.get("connected_avg_duration_minutes")

    yest_total = yesterday.get("新增线索量", 0) if yesterday else 0
    yest_jietong = yesterday.get("新增线索接通量", 0) if yesterday else 0
    yest_yixiang = yesterday.get("有效线索量", 0) if yesterday else 0

    r += 1; _subsection_row(r, "规模指标")
    r += 1; _data_row(r, "新增线索量", "条", today_total)
    r += 1; _data_row(r, "昨日全天对比", "条", yest_total if yesterday else "")
    r += 1; _data_row(r, "新增线索接通量", "条", today_jietong)
    r += 1; _data_row(r, "新增线索呼叫通次", "次", today_call)
    if today_connected_avg_duration_minutes is not None:
        r += 1; _data_row(r, "已接通通话平均时长（向上取整）", "分钟", today_connected_avg_duration_minutes)
    r += 1; _data_row(r, "平均呼叫通次", "次/条", round(today_call / today_total, 2) if today_total else 0.0)

    r += 1; _subsection_row(r, "检测指标")

    today_jietong_rate = _safe_div(today_jietong, today_total)
    yest_jietong_rate = _safe_div(yest_jietong, yest_total)
    today_youxiao_rate = _safe_div(today_yixiang, today_jietong)
    today_overall_rate = _safe_div(today_yixiang, today_total)
    yest_overall_rate = _safe_div(yest_yixiang, yest_total)

    r += 1; _pct_row(r, "接通率", "%", today_jietong_rate)
    r += 1; _pct_row(r, "昨日全天对比", "%", yest_jietong_rate if yesterday else "")
    r += 1; _pct_row(r, "接通有效率", "%", today_youxiao_rate)
    r += 1; _pct_row(r, "整体有效率", "%", today_overall_rate)
    r += 1; _pct_row(r, "昨日全天对比", "%", yest_overall_rate if yesterday else "")

    filepath = out_dir / f"{company}_日度月度汇总表_{date}.xlsx"
    wb.save(filepath)
    return filepath


def generate_multi_daily_summary(title: str, date: str, grouped_data: list[dict],
                                 output_dir: str | Path) -> Path:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    year_month = date[:4]
    date_display = f"20{date[:2]}-{date[2:4]}-{date[4:6]}"

    prepared = []
    for item in grouped_data:
        company = item["company"]
        tracking_company = item.get("tracking_company", company)
        data = item["data"]
        tracking_aliases = item.get("tracking_aliases") or []
        tracking = _load_tracking_with_aliases(tracking_company, year_month, tracking_aliases)
        tracking[date] = _tracking_record(
            data,
            item.get("monthly_stats") or {},
        )
        backfill_tracking_from_daily_reports(
            tracking_company,
            date,
            tracking,
            aliases=tracking_aliases,
        )
        _, missing_dates = backfill_tracking_mtd_gaps_from_summaries(
            tracking_company,
            date,
            tracking,
            aliases=tracking_aliases,
        )
        _save_tracking(tracking_company, year_month, tracking)

        mtd_stats, _ = aggregate_tracking_mtd(tracking, date)
        _print_tracking_mtd_status(tracking_company, date, missing_dates)

        yesterday = _load_previous_tracking(tracking_company, date, tracking_aliases)

        mtd_total = mtd_stats['累计线索量']
        mtd_jietong = mtd_stats['累计接通量']
        mtd_yixiang = mtd_stats['累计有效线索量']
        mtd_call_count = mtd_stats['累计呼叫通次']

        today_total = data.get("total", 0)
        today_jietong = data.get("jietong_count", 0)
        today_call = data.get("call_count", 0)
        today_yixiang = data.get("yixiang_count", 0)
        today_connected_avg_duration_minutes = data.get("connected_avg_duration_minutes")
        yest_total = yesterday.get("新增线索量", 0) if yesterday else 0
        yest_jietong = yesterday.get("新增线索接通量", 0) if yesterday else 0
        yest_yixiang = yesterday.get("有效线索量", 0) if yesterday else 0

        prepared.append({
            "company": company,
            "mtd_total": mtd_total,
            "mtd_jietong": mtd_jietong,
            "mtd_yixiang": mtd_yixiang,
            "mtd_call_count": mtd_call_count,
            "mtd_avg_call": round(mtd_call_count / mtd_total, 2) if mtd_total else 0.0,
            "mtd_jietong_rate": _safe_div(mtd_jietong, mtd_total),
            "mtd_youxiao_rate": _safe_div(mtd_yixiang, mtd_jietong),
            "mtd_overall_rate": _safe_div(mtd_yixiang, mtd_total),
            "today_total": today_total,
            "yest_total": yest_total if yesterday else "",
            "today_jietong": today_jietong,
            "today_small_round_count": data.get("small_round_count"),
            "today_hangup_after_connect_count": data.get("hangup_after_connect_count"),
            "today_call": today_call,
            "today_connected_avg_duration_minutes": today_connected_avg_duration_minutes,
            "today_avg_call": round(today_call / today_total, 2) if today_total else 0.0,
            "today_jietong_rate": _safe_div(today_jietong, today_total),
            "yest_jietong_rate": _safe_div(yest_jietong, yest_total) if yesterday else "",
            "today_youxiao_rate": _safe_div(today_yixiang, today_jietong),
            "today_overall_rate": _safe_div(today_yixiang, today_total),
            "yest_overall_rate": _safe_div(yest_yixiang, yest_total) if yesterday else "",
        })

    wb = Workbook()
    ws = wb.active
    ws.title = title

    total_cols = 2 + len(prepared)
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 10
    for idx in range(3, total_cols + 1):
        ws.column_dimensions[chr(64 + idx)].width = 30

    headers = ["业务场景", "单位"] + [item["company"] for item in prepared]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    def _section_row(row_num, label):
        for c in range(1, total_cols + 1):
            ws.cell(row=row_num, column=c).fill = SECTION_FILL
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        ws.cell(row=row_num, column=1, value=label).font = HEADER_FONT

    def _subsection_row(row_num, label):
        for c in range(1, total_cols + 1):
            ws.cell(row=row_num, column=c).fill = SUBSECTION_FILL
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        ws.cell(row=row_num, column=1, value=label).font = HEADER_FONT

    def _data_row(row_num, name, unit, key):
        ws.cell(row=row_num, column=1, value=name).font = DATA_FONT
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        ws.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_num, column=2, value=unit).font = DATA_FONT
        ws.cell(row=row_num, column=2).border = THIN_BORDER
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        for idx, item in enumerate(prepared, start=3):
            cell = ws.cell(row=row_num, column=idx, value=item[key])
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

    def _label_row(row_num, name):
        ws.cell(row=row_num, column=1, value=name).font = DATA_FONT
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        ws.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_num, column=2, value="").font = DATA_FONT
        ws.cell(row=row_num, column=2).border = THIN_BORDER
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        for idx in range(3, total_cols + 1):
            cell = ws.cell(row=row_num, column=idx, value="")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

    def _pct_row(row_num, name, unit, key):
        ws.cell(row=row_num, column=1, value=name).font = DATA_FONT
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        ws.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_num, column=2, value=unit).font = DATA_FONT
        ws.cell(row=row_num, column=2).border = THIN_BORDER
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        for idx, item in enumerate(prepared, start=3):
            cell = ws.cell(row=row_num, column=idx, value=item[key])
            cell.font = DATA_FONT
            cell.number_format = '0.00"%"'
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

    r = 1
    r += 1; _section_row(r, f"MTD (截止：{date_display})")
    r += 1; _subsection_row(r, "规模指标")
    r += 1; _data_row(r, "累计线索量", "条", "mtd_total")
    r += 1; _data_row(r, "接通量", "条", "mtd_jietong")
    r += 1; _data_row(r, "有效线索量", "条", "mtd_yixiang")
    r += 1; _data_row(r, "呼叫通次", "次", "mtd_call_count")
    r += 1; _data_row(r, "平均呼叫通次", "次/条", "mtd_avg_call")

    r += 1; _subsection_row(r, "检测指标")
    r += 1; _pct_row(r, "接通率", "%", "mtd_jietong_rate")
    r += 1; _pct_row(r, "接通有效率", "%", "mtd_youxiao_rate")
    r += 1; _pct_row(r, "整体有效率", "%", "mtd_overall_rate")

    r += 1; _section_row(r, f"Daily Report ({date_display})")
    r += 1; _subsection_row(r, "规模指标")
    r += 1; _data_row(r, "新增线索量", "条", "today_total")
    r += 1; _data_row(r, "昨日全天对比", "条", "yest_total")
    r += 1; _data_row(r, "新增线索接通量", "条", "today_jietong")
    if title == "广州新车":
        r += 1; _label_row(r, "其中：")
        r += 1; _data_row(r, "  沟通小于等于2轮", "条", "today_small_round_count")
        r += 1; _data_row(r, "  接通后挂断", "条", "today_hangup_after_connect_count")
    r += 1; _data_row(r, "新增线索呼叫通次", "次", "today_call")
    if any(item.get("today_connected_avg_duration_minutes") is not None for item in prepared):
        r += 1; _data_row(r, "已接通通话平均时长（向上取整）", "分钟", "today_connected_avg_duration_minutes")
    r += 1; _data_row(r, "平均呼叫通次", "次/条", "today_avg_call")

    r += 1; _subsection_row(r, "检测指标")
    r += 1; _pct_row(r, "接通率（=新增线索接通量/新增线索量）", "%", "today_jietong_rate")
    r += 1; _pct_row(r, "昨日全天对比", "%", "yest_jietong_rate")
    r += 1; _pct_row(r, "接通有效率（=有效线索量/新增线索接通量）", "%", "today_youxiao_rate")
    r += 1; _pct_row(r, "整体有效率（=有效线索量/新增线索量）", "%", "today_overall_rate")
    r += 1; _pct_row(r, "昨日全天对比", "%", "yest_overall_rate")

    filepath = out_dir / f"{title}_日度月度汇总表_{date}.xlsx"
    wb.save(filepath)
    return filepath
